#!/usr/bin/env python3
"""
Parse a TAGLYZ Consolidated Profit and Loss workbook into a normalised JSON payload.

Written to be resilient to the ways the CPA's file shifts month to month:
  - property columns are located by header text, not fixed column letters
  - line items are located by row label, not fixed row numbers
  - section subtotals are read from the workbook's own subtotal rows AND
    recomputed from the line items, so discrepancies are surfaced, not hidden
"""
import json
import re
import sys
from collections import OrderedDict

import openpyxl

MONTH_RE = re.compile(r"^(\d{2})\.\s*(\w+)", re.I)
MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]
ABBR = {m: m[:3] for m in MONTHS}

# Row labels that are structure, not data.
STRUCTURE = {
    "income", "revenue", "total for revenue", "total for income",
    "cost of goods sold", "gross profit", "expenses",
    "administrative expenses", "total for administrative expenses",
    "rental expenses", "total for rental expenses",
    "utilities", "total for utilities", "total for expenses",
    "net operating income", "other income", "other expenses",
    "net other income", "net income", "capital improvements",
    "debt payments", "cash flow", "net cash flow",
}
INCOME_ITEMS = {"rental income", "laundry", "interest income", "other income"}

# Canonical grouping for the expense-mix view.
GROUPS = {
    "Taxes & insurance": {"property taxes", "insurance"},
    "Repairs & handyman": {"repairs & maintenance", "handyman expense",
                           "commission expense"},
    "Utilities": {"electric", "gas", "water", "trash", "sewage/stormwater",
                  "landscaping", "security", "cable & internet"},
    "Management & admin": {"property manager", "accounting", "legal",
                           "bank fees", "advertising", "investor interest"},
}


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def find_header_row(ws):
    """The header row is the one carrying the property/entity column names."""
    for r in range(1, 15):
        filled = sum(1 for c in range(2, ws.max_column + 1)
                     if isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip())
        if filled >= 5:
            return r
    return 5


def property_columns(ws, hdr):
    """
    Map column index -> property name, and capture entity structure.
    Skips entity-name columns (which head a group) and 'Total for ...' columns.
    """
    props, entities, current = OrderedDict(), OrderedDict(), None
    for c in range(2, ws.max_column + 1):
        raw = ws.cell(hdr, c).value
        if not isinstance(raw, str) or not raw.strip():
            continue
        name = raw.strip()
        low = name.lower()
        if low.startswith("total"):   # 'Total for <entity>' and the grand 'Total'
            continue
        # An entity header column is empty in the data rows; it labels the group.
        has_data = any(isinstance(ws.cell(r, c).value, (int, float))
                       for r in range(hdr + 1, min(hdr + 45, ws.max_row + 1)))
        if not has_data and re.match(r"^(taglyz|.*\bllc\b)", low):
            current = name
            entities.setdefault(current, [])
            continue
        if re.match(r"^(taglyz\b|.*\bllc$)", low):
            current = name
            entities.setdefault(current, [])
            continue
        props[c] = name
        if current:
            entities.setdefault(current, []).append(name)
    return props, entities


def parse_month(ws):
    hdr = find_header_row(ws)
    props, entities = property_columns(ws, hdr)

    # Locate the boundary rows we care about.
    rows = {}
    for r in range(hdr + 1, ws.max_row + 1):
        lab = norm(ws.cell(r, 1).value)
        if lab and lab not in rows:
            rows[lab] = r
    stop = rows.get("total for expenses") or rows.get("net operating income")
    if not stop:
        raise ValueError(f"{ws.title}: no 'Total for Expenses' row found")

    out = {}
    for col, name in props.items():
        income, expenses = OrderedDict(), OrderedDict()
        for r in range(hdr + 1, stop):
            lab = norm(ws.cell(r, 1).value)
            if not lab:
                continue
            v = ws.cell(r, col).value
            if not isinstance(v, (int, float)):
                continue
            label = str(ws.cell(r, 1).value).strip()
            if lab in INCOME_ITEMS:
                income[label] = income.get(label, 0.0) + v
            elif lab in STRUCTURE:
                continue
            else:
                expenses[label] = expenses.get(label, 0.0) + v

        def at(label):
            r = rows.get(label)
            return num(ws.cell(r, col).value) if r else 0.0

        rep_income = at("total for income") or sum(income.values())
        rep_exp = at("total for expenses")
        rep_noi = at("net operating income")
        debt = at("debt payments") or at("debt payments")
        if not debt:
            for k in ("debt payments", "debt payment"):
                if k in rows:
                    debt = num(ws.cell(rows[k], col).value)
                    break
        capex = at("capital improvements")

        calc_exp = sum(expenses.values())
        out[name] = {
            "income": {k: round(v, 2) for k, v in income.items()},
            "expenses": {k: round(v, 2) for k, v in expenses.items()},
            "totalIncome": round(rep_income, 2),
            "totalExpenses": round(rep_exp, 2),
            "totalExpensesRecorded": round(calc_exp, 2),
            "noi": round(rep_noi, 2),
            "debt": round(debt, 2),
            "capex": round(capex, 2),
            "cashflow": round(rep_noi - debt, 2),
            "variance": round(calc_exp - rep_exp, 2),
        }
    return out, entities


def parse_consolidated(wb, months):
    """
    Read the workbook's own portfolio roll-up sheet, so the site can flag any
    drift between it and the sum of the property columns.
    """
    sheet = next((n for n in wb.sheetnames if "consolidated" in n.lower()), None)
    if not sheet:
        return None
    ws = wb[sheet]
    hdr = None
    for r in range(1, 12):
        vals = [norm(ws.cell(r, c).value) for c in range(2, 16)]
        if any(v.startswith("jan") for v in vals):
            hdr = r
            break
    if hdr is None:
        return None
    cols = {}
    for c in range(2, ws.max_column + 1):
        v = norm(ws.cell(hdr, c).value)
        for m in months:
            if v.startswith(m.lower()):
                cols[m] = c
    rows = {}
    for r in range(hdr + 1, ws.max_row + 1):
        lab = norm(ws.cell(r, 1).value)
        if lab and lab not in rows:
            rows[lab] = r
    def series(label):
        r = rows.get(label)
        if not r:
            return [0.0] * len(months)
        return [round(num(ws.cell(r, cols[m]).value), 2) if m in cols else 0.0
                for m in months]
    return {
        "totalIncome": series("total for income"),
        "totalExpenses": series("total for expenses"),
        "noi": series("net operating income"),
        "debt": series("debt payments"),
        "capex": series("capital improvements"),
    }


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    month_sheets = []
    for name in wb.sheetnames:
        m = MONTH_RE.match(name.strip())
        if m and m.group(2).capitalize() in MONTHS:
            month_sheets.append((int(m.group(1)), m.group(2).capitalize(), name))
    month_sheets.sort()
    if not month_sheets:
        raise ValueError("no monthly sheets found")

    n = len(month_sheets)
    months = [ABBR[m] for _, m, _ in month_sheets]
    per_month, entities = [], OrderedDict()
    for _, month, sheet in month_sheets:
        parsed, ents = parse_month(wb[sheet])
        per_month.append(parsed)
        for e, members in ents.items():
            entities.setdefault(e, [])
            for m in members:
                if m not in entities[e]:
                    entities[e].append(m)

    # Every property gets one slot per month, in month order. A property absent
    # from a given month gets a zero slot IN THAT POSITION - never appended at
    # the end, which would silently shift the whole series.
    def blank():
        return {"income": {}, "expenses": {}, "totalIncome": 0.0,
                "totalExpenses": 0.0, "totalExpensesRecorded": 0.0, "noi": 0.0,
                "debt": 0.0, "capex": 0.0, "cashflow": 0.0, "variance": 0.0,
                "missing": True}
    all_props = []
    for parsed in per_month:
        for p in parsed:
            if p not in all_props:
                all_props.append(p)
    data = {p: [per_month[i].get(p) or blank() for i in range(n)] for p in all_props}

    # Year label from the report subtitle, e.g. "January 1-31, 2026"
    yr = None
    for row in (3, 2, 4):
        v = wb[month_sheets[0][2]].cell(row, 1).value
        if isinstance(v, str):
            m = re.search(r"(20\d{2})", v)
            if m:
                yr = m.group(1)
                break

    entities = OrderedDict((e, m) for e, m in entities.items() if m)
    consolidated = parse_consolidated(wb, months)

    return {
        "year": yr or "",
        "months": months,
        "throughMonth": months[-1],
        "entities": entities,
        "groups": {g: sorted(v) for g, v in GROUPS.items()},
        "properties": data,
        "consolidated": consolidated,
    }


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "latest.xlsx"
    payload = parse(src)
    out = sys.argv[2] if len(sys.argv) > 2 else "payload.json"
    with open(out, "w") as f:
        json.dump(payload, f, separators=(",", ":"))
    print(f"{len(payload['properties'])} properties · {len(payload['months'])} months "
          f"· {payload['year']}")
    for e, members in payload["entities"].items():
        print(f"  {e}: {', '.join(members)}")
