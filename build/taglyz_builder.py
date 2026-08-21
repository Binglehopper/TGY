#!/usr/bin/env python3
"""
TAGLYZ portfolio report - self-contained builder.

  pip install openpyxl cryptography
  python3 taglyz_builder.py <workbook.xlsx> "<source file name>" "<passphrase>" <outdir>

Writes <outdir>/index.html: the full report with its data encrypted under the
passphrase. Upload that file to github.com/Binglehopper/TGY to publish it.

Generated file - bundles parse.py, build.py and template.html so the whole
pipeline travels as one artifact. Regenerate with bundle.py after editing those.
"""
#!/usr/bin/env python3
"""
Parse a TAGLYZ Consolidated Profit and Loss workbook into a normalised JSON payload.

Written to be resilient to the ways the CPA's file shifts month to month:
  - property columns are located by header text, not fixed column letters
  - line items are located by row label, not fixed row numbers
  - section subtotals are read from the workbook's own subtotal rows AND
    recomputed from the line items, so discrepancies are surfaced, not hidden
"""
import json
import re
import sys
from collections import OrderedDict

import openpyxl

MONTH_RE = re.compile(r"^(\d{2})\.\s*(\w+)", re.I)
MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]
ABBR = {m: m[:3] for m in MONTHS}

# Row labels that are structure, not data.
STRUCTURE = {
    "income", "revenue", "total for revenue", "total for income",
    "cost of goods sold", "gross profit", "expenses",
    "administrative expenses", "total for administrative expenses",
    "rental expenses", "total for rental expenses",
    "utilities", "total for utilities", "total for expenses",
    "net operating income", "other income", "other expenses",
    "net other income", "net income", "capital improvements",
    "debt payments", "cash flow", "net cash flow",
}
INCOME_ITEMS = {"rental income", "laundry", "interest income", "other income"}

# Canonical grouping for the expense-mix view.
GROUPS = {
    "Taxes & insurance": {"property taxes", "insurance"},
    "Repairs & handyman": {"repairs & maintenance", "handyman expense",
                           "commission expense"},
    "Utilities": {"electric", "gas", "water", "trash", "sewage/stormwater",
                  "landscaping", "security", "cable & internet"},
    "Management & admin": {"property manager", "accounting", "legal",
                           "bank fees", "advertising", "investor interest"},
}


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def find_header_row(ws):
    """The header row is the one carrying the property/entity column names."""
    for r in range(1, 15):
        filled = sum(1 for c in range(2, ws.max_column + 1)
                     if isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip())
        if filled >= 5:
            return r
    return 5


def property_columns(ws, hdr):
    """
    Map column index -> property name, and capture entity structure.
    Skips entity-name columns (which head a group) and 'Total for ...' columns.
    """
    props, entities, current = OrderedDict(), OrderedDict(), None
    for c in range(2, ws.max_column + 1):
        raw = ws.cell(hdr, c).value
        if not isinstance(raw, str) or not raw.strip():
            continue
        name = raw.strip()
        low = name.lower()
        if low.startswith("total"):   # 'Total for <entity>' and the grand 'Total'
            continue
        # An entity header column is empty in the data rows; it labels the group.
        has_data = any(isinstance(ws.cell(r, c).value, (int, float))
                       for r in range(hdr + 1, min(hdr + 45, ws.max_row + 1)))
        if not has_data and re.match(r"^(taglyz|.*\bllc\b)", low):
            current = name
            entities.setdefault(current, [])
            continue
        if re.match(r"^(taglyz\b|.*\bllc$)", low):
            current = name
            entities.setdefault(current, [])
            continue
        props[c] = name
        if current:
            entities.setdefault(current, []).append(name)
    return props, entities


def parse_month(ws):
    hdr = find_header_row(ws)
    props, entities = property_columns(ws, hdr)

    # Locate the boundary rows we care about.
    rows = {}
    for r in range(hdr + 1, ws.max_row + 1):
        lab = norm(ws.cell(r, 1).value)
        if lab and lab not in rows:
            rows[lab] = r
    stop = rows.get("total for expenses") or rows.get("net operating income")
    if not stop:
        raise ValueError(f"{ws.title}: no 'Total for Expenses' row found")

    out = {}
    for col, name in props.items():
        income, expenses = OrderedDict(), OrderedDict()
        for r in range(hdr + 1, stop):
            lab = norm(ws.cell(r, 1).value)
            if not lab:
                continue
            v = ws.cell(r, col).value
            if not isinstance(v, (int, float)):
                continue
            label = str(ws.cell(r, 1).value).strip()
            if lab in INCOME_ITEMS:
                income[label] = income.get(label, 0.0) + v
            elif lab in STRUCTURE:
                continue
            else:
                expenses[label] = expenses.get(label, 0.0) + v

        def at(label):
            r = rows.get(label)
            return num(ws.cell(r, col).value) if r else 0.0

        rep_income = at("total for income") or sum(income.values())
        rep_exp = at("total for expenses")
        rep_noi = at("net operating income")
        debt = at("debt payments") or at("debt payments")
        if not debt:
            for k in ("debt payments", "debt payment"):
                if k in rows:
                    debt = num(ws.cell(rows[k], col).value)
                    break
        capex = at("capital improvements")

        calc_exp = sum(expenses.values())
        out[name] = {
            "income": {k: round(v, 2) for k, v in income.items()},
            "expenses": {k: round(v, 2) for k, v in expenses.items()},
            "totalIncome": round(rep_income, 2),
            "totalExpenses": round(rep_exp, 2),
            "totalExpensesRecorded": round(calc_exp, 2),
            "noi": round(rep_noi, 2),
            "debt": round(debt, 2),
            "capex": round(capex, 2),
            "cashflow": round(rep_noi - debt, 2),
            "variance": round(calc_exp - rep_exp, 2),
        }
    return out, entities


def parse_consolidated(wb, months):
    """
    Read the workbook's own portfolio roll-up sheet, so the site can flag any
    drift between it and the sum of the property columns.
    """
    sheet = next((n for n in wb.sheetnames if "consolidated" in n.lower()), None)
    if not sheet:
        return None
    ws = wb[sheet]
    hdr = None
    for r in range(1, 12):
        vals = [norm(ws.cell(r, c).value) for c in range(2, 16)]
        if any(v.startswith("jan") for v in vals):
            hdr = r
            break
    if hdr is None:
        return None
    cols = {}
    for c in range(2, ws.max_column + 1):
        v = norm(ws.cell(hdr, c).value)
        for m in months:
            if v.startswith(m.lower()):
                cols[m] = c
    rows = {}
    for r in range(hdr + 1, ws.max_row + 1):
        lab = norm(ws.cell(r, 1).value)
        if lab and lab not in rows:
            rows[lab] = r
    def series(label):
        r = rows.get(label)
        if not r:
            return [0.0] * len(months)
        return [round(num(ws.cell(r, cols[m]).value), 2) if m in cols else 0.0
                for m in months]
    return {
        "totalIncome": series("total for income"),
        "totalExpenses": series("total for expenses"),
        "noi": series("net operating income"),
        "debt": series("debt payments"),
        "capex": series("capital improvements"),
    }


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    month_sheets = []
    for name in wb.sheetnames:
        m = MONTH_RE.match(name.strip())
        if m and m.group(2).capitalize() in MONTHS:
            month_sheets.append((int(m.group(1)), m.group(2).capitalize(), name))
    month_sheets.sort()
    if not month_sheets:
        raise ValueError("no monthly sheets found")

    n = len(month_sheets)
    months = [ABBR[m] for _, m, _ in month_sheets]
    per_month, entities = [], OrderedDict()
    for _, month, sheet in month_sheets:
        parsed, ents = parse_month(wb[sheet])
        per_month.append(parsed)
        for e, members in ents.items():
            entities.setdefault(e, [])
            for m in members:
                if m not in entities[e]:
                    entities[e].append(m)

    # Every property gets one slot per month, in month order. A property absent
    # from a given month gets a zero slot IN THAT POSITION - never appended at
    # the end, which would silently shift the whole series.
    def blank():
        return {"income": {}, "expenses": {}, "totalIncome": 0.0,
                "totalExpenses": 0.0, "totalExpensesRecorded": 0.0, "noi": 0.0,
                "debt": 0.0, "capex": 0.0, "cashflow": 0.0, "variance": 0.0,
                "missing": True}
    all_props = []
    for parsed in per_month:
        for p in parsed:
            if p not in all_props:
                all_props.append(p)
    data = {p: [per_month[i].get(p) or blank() for i in range(n)] for p in all_props}

    # Year label from the report subtitle, e.g. "January 1-31, 2026"
    yr = None
    for row in (3, 2, 4):
        v = wb[month_sheets[0][2]].cell(row, 1).value
        if isinstance(v, str):
            m = re.search(r"(20\d{2})", v)
            if m:
                yr = m.group(1)
                break

    entities = OrderedDict((e, m) for e, m in entities.items() if m)
    consolidated = parse_consolidated(wb, months)

    return {
        "year": yr or "",
        "months": months,
        "throughMonth": months[-1],
        "entities": entities,
        "groups": {g: sorted(v) for g, v in GROUPS.items()},
        "properties": data,
        "consolidated": consolidated,
    }



#!/usr/bin/env python3
"""
Build the encrypted TAGLYZ portfolio site.

  python3 build.py <workbook.xlsx> <source-name> <passphrase> [outdir]

The data payload is encrypted with AES-256-GCM under a PBKDF2-SHA256 key derived
from the passphrase, then base64'd into the page. The published HTML contains
ciphertext only - without the passphrase there is nothing readable in the source.
"""
import base64
import datetime
import json
import os
import sys

from cryptography.hazmat.primitives.ciphers.aead import AESGCM


ITERATIONS = 600_000


def encrypt(plaintext: bytes, passphrase: str):
    import hashlib
    salt = os.urandom(16)
    iv = os.urandom(12)
    key = hashlib.pbkdf2_hmac("sha256", passphrase.encode(), salt, ITERATIONS, 32)
    body = AESGCM(key).encrypt(iv, plaintext, None)
    return base64.b64encode(salt + iv + body).decode()


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    workbook, source, passphrase = sys.argv[1], sys.argv[2], sys.argv[3]
    outdir = sys.argv[4] if len(sys.argv) > 4 else "dist"

    payload = parse_workbook(workbook)
    raw = json.dumps(payload, separators=(",", ":")).encode()
    blob = encrypt(raw, passphrase)

    meta = {
        "iterations": ITERATIONS,
        "source": source,
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%d %b %Y"),
    }

    html = base64.b64decode(TEMPLATE_B64).decode()
    html = html.replace("__PAYLOAD__", blob).replace("__META__", json.dumps(meta))

    os.makedirs(outdir, exist_ok=True)
    out = os.path.join(outdir, "index.html")
    with open(out, "w") as f:
        f.write(html)
    # Keep GitHub Pages from running the content through Jekyll.
    open(os.path.join(outdir, ".nojekyll"), "w").close()

    props = len(payload["properties"])
    print(f"built {out}  ({len(html):,} bytes)")
    print(f"  {props} properties · {len(payload['months'])} months · through {payload['throughMonth']} {payload['year']}")
    print(f"  plaintext {len(raw):,} B → ciphertext {len(blob):,} B (base64)")
    assert "__PAYLOAD__" not in html and "__META__" not in html, "placeholder left unreplaced"
    # Leak check: nothing may appear in the built page that was not already in the
    # template source. Counting rather than membership avoids flagging labels the
    # template legitimately hardcodes (e.g. "Rental Income" in the shaping code).
    tmpl = base64.b64decode(TEMPLATE_B64).decode()
    # The TAGLYZ name itself appears by design (page title, source filename in the
    # footer). What must not appear are property names, line labels and figures.
    probes = list(payload["properties"].keys()) + [
        "Repairs & Maintenance", "Property Taxes", "Handyman Expense"]
    leaked = [p for p in probes if html.count(p) > tmpl.count(p)]
    assert not leaked, f"plaintext leak in output: {leaked}"
    # And no recognisable figure from the workbook.
    fig = f"{payload['properties']['2727 Broadway'][0]['noi']:.2f}"
    assert fig not in html, f"numeric leak: {fig}"
    print(f"  leak check passed: {len(probes)} names + sample figures absent from the HTML")




TEMPLATE_B64 = "".join([
    "PCFET0NUWVBFIGh0bWw+CjxodG1sIGxhbmc9ImVuIiBkYXRhLXRoZW1lPSJsaWdodCI+CjxoZWFkPgo8bWV0YSBjaGFyc2V0PSJ1"
    "dGYtOCI+CjxtZXRhIG5hbWU9InZpZXdwb3J0IiBjb250ZW50PSJ3aWR0aD1kZXZpY2Utd2lkdGgsIGluaXRpYWwtc2NhbGU9MSI+"
    "CjxtZXRhIG5hbWU9InJvYm90cyIgY29udGVudD0ibm9pbmRleCwgbm9mb2xsb3csIG5vYXJjaGl2ZSI+CjxtZXRhIG5hbWU9InJl"
    "ZmVycmVyIiBjb250ZW50PSJuby1yZWZlcnJlciI+Cjx0aXRsZT5UQUdMWVogUG9ydGZvbGlvPC90aXRsZT4KPHN0eWxlPgogIDpy"
    "b290IHsKICAgIGNvbG9yLXNjaGVtZTogbGlnaHQ7CiAgICAtLXBhZ2U6I2Y5ZjlmNzsgLS1zdXJmYWNlLTE6I2ZjZmNmYjsKICAg"
    "IC0tdGV4dC1wcmltYXJ5OiMwYjBiMGI7IC0tdGV4dC1zZWNvbmRhcnk6IzUyNTE0ZTsgLS10ZXh0LW11dGVkOiM4OTg3ODE7CiAg"
    "ICAtLWdyaWQ6I2UxZTBkOTsgLS1heGlzOiNjM2MyYjc7IC0tYm9yZGVyOnJnYmEoMTEsMTEsMTEsMC4xMCk7CiAgICAtLXNlcmll"
    "cy0xOiMyYTc4ZDY7IC0tc2VyaWVzLTI6I2ViNjgzNDsgLS1zZXJpZXMtMzojMWJhZjdhOyAtLXNlcmllcy00OiNlZGExMDA7CiAg"
    "ICAtLXBvczojMmE3OGQ2OyAtLW5lZ2I6I2QwM2IzYjsKICAgIC0tZ29vZDojMDA2MzAwOyAtLWNyaXRpY2FsOiNkMDNiM2I7IC0t"
    "d2FybmluZzojZmFiMjE5OwogICAgLS1ob3ZlcjpyZ2JhKDExLDExLDExLDAuMDQpOwogIH0KICA6cm9vdFtkYXRhLXRoZW1lPSJk"
    "YXJrIl0gewogICAgY29sb3Itc2NoZW1lOiBkYXJrOwogICAgLS1wYWdlOiMwZDBkMGQ7IC0tc3VyZmFjZS0xOiMxYTFhMTk7CiAg"
    "ICAtLXRleHQtcHJpbWFyeTojZmZmZmZmOyAtLXRleHQtc2Vjb25kYXJ5OiNjM2MyYjc7IC0tdGV4dC1tdXRlZDojODk4NzgxOwog"
    "ICAgLS1ncmlkOiMyYzJjMmE7IC0tYXhpczojMzgzODM1OyAtLWJvcmRlcjpyZ2JhKDI1NSwyNTUsMjU1LDAuMTApOwogICAgLS1z"
    "ZXJpZXMtMTojMzk4N2U1OyAtLXNlcmllcy0yOiNkOTU5MjY7IC0tc2VyaWVzLTM6IzE5OWU3MDsgLS1zZXJpZXMtNDojYzk4NTAw"
    "OwogICAgLS1wb3M6IzM5ODdlNTsgLS1uZWdiOiNkMDNiM2I7CiAgICAtLWdvb2Q6IzBjYTMwYzsgLS1jcml0aWNhbDojZDAzYjNi"
    "OyAtLXdhcm5pbmc6I2ZhYjIxOTsKICAgIC0taG92ZXI6cmdiYSgyNTUsMjU1LDI1NSwwLjA2KTsKICB9CiAgKiB7IGJveC1zaXpp"
    "bmc6Ym9yZGVyLWJveDsgfQogIGJvZHkgeyBtYXJnaW46MDsgYmFja2dyb3VuZDp2YXIoLS1wYWdlKTsgY29sb3I6dmFyKC0tdGV4"
    "dC1wcmltYXJ5KTsKICAgIGZvbnQtZmFtaWx5OnN5c3RlbS11aSwtYXBwbGUtc3lzdGVtLCJTZWdvZSBVSSIsc2Fucy1zZXJpZjsg"
    "Zm9udC1zaXplOjE0cHg7IGxpbmUtaGVpZ2h0OjEuNTsKICAgIC13ZWJraXQtZm9udC1zbW9vdGhpbmc6YW50aWFsaWFzZWQ7IH0K"
    "ICAud3JhcCB7IG1heC13aWR0aDoxMTgwcHg7IG1hcmdpbjowIGF1dG87IHBhZGRpbmc6MjhweCAyNHB4IDY0cHg7IH0KICBbaGlk"
    "ZGVuXSB7IGRpc3BsYXk6bm9uZSAhaW1wb3J0YW50OyB9CgogIC8qIC0tLS0tLS0tLS0gbG9jayBzY3JlZW4gLS0tLS0tLS0tLSAq"
    "LwogICNsb2NrIHsgbWluLWhlaWdodDoxMDB2aDsgZGlzcGxheTpmbGV4OyBhbGlnbi1pdGVtczpjZW50ZXI7IGp1c3RpZnktY29u"
    "dGVudDpjZW50ZXI7IHBhZGRpbmc6MjRweDsgfQogIC5sb2NrY2FyZCB7IGJhY2tncm91bmQ6dmFyKC0tc3VyZmFjZS0xKTsgYm9y"
    "ZGVyOjFweCBzb2xpZCB2YXIoLS1ib3JkZXIpOyBib3JkZXItcmFkaXVzOjE2cHg7CiAgICBwYWRkaW5nOjM0cHggMzJweDsgd2lk"
    "dGg6MTAwJTsgbWF4LXdpZHRoOjQyMHB4OyB9CiAgLmxvY2tjYXJkIGgxIHsgZm9udC1zaXplOjE5cHg7IG1hcmdpbjowIDAgNnB4"
    "OyBmb250LXdlaWdodDo2NDA7IGxldHRlci1zcGFjaW5nOi0wLjAxZW07IH0KICAubG9ja2NhcmQgcCB7IGNvbG9yOnZhcigtLXRl"
    "eHQtc2Vjb25kYXJ5KTsgZm9udC1zaXplOjEzcHg7IG1hcmdpbjowIDAgMjBweDsgfQogIC5sb2NrY2FyZCBsYWJlbCB7IGRpc3Bs"
    "YXk6YmxvY2s7IGZvbnQtc2l6ZToxMi41cHg7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgbWFyZ2luLWJvdHRvbTo2cHg7"
    "IH0KICAubG9ja2NhcmQgaW5wdXQgeyB3aWR0aDoxMDAlOyBwYWRkaW5nOjExcHggMTNweDsgZm9udDppbmhlcml0OyBib3JkZXIt"
    "cmFkaXVzOjEwcHg7CiAgICBib3JkZXI6MXB4IHNvbGlkIHZhcigtLWF4aXMpOyBiYWNrZ3JvdW5kOnZhcigtLXBhZ2UpOyBjb2xv"
    "cjp2YXIoLS10ZXh0LXByaW1hcnkpOyB9CiAgLmxvY2tjYXJkIGlucHV0OmZvY3VzIHsgb3V0bGluZToycHggc29saWQgdmFyKC0t"
    "c2VyaWVzLTEpOyBvdXRsaW5lLW9mZnNldDoxcHg7IGJvcmRlci1jb2xvcjp0cmFuc3BhcmVudDsgfQogIC5sb2NrY2FyZCBidXR0"
    "b24geyBtYXJnaW4tdG9wOjE0cHg7IHdpZHRoOjEwMCU7IHBhZGRpbmc6MTFweDsgZm9udDppbmhlcml0OyBmb250LXdlaWdodDo2"
    "MDA7CiAgICBib3JkZXI6MDsgYm9yZGVyLXJhZGl1czoxMHB4OyBiYWNrZ3JvdW5kOnZhcigtLXNlcmllcy0xKTsgY29sb3I6I2Zm"
    "ZjsgY3Vyc29yOnBvaW50ZXI7IH0KICAubG9ja2NhcmQgYnV0dG9uOmRpc2FibGVkIHsgb3BhY2l0eTouNTU7IGN1cnNvcjpkZWZh"
    "dWx0OyB9CiAgLmVyciB7IGNvbG9yOnZhcigtLWNyaXRpY2FsKTsgZm9udC1zaXplOjEyLjVweDsgbWFyZ2luLXRvcDoxMnB4OyBt"
    "aW4taGVpZ2h0OjE4cHg7IH0KCiAgLyogLS0tLS0tLS0tLSBjaHJvbWUgLS0tLS0tLS0tLSAqLwogIGhlYWRlci50b3AgeyBkaXNw"
    "bGF5OmZsZXg7IGFsaWduLWl0ZW1zOmZsZXgtc3RhcnQ7IGp1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuOyBnYXA6MjBweDsK"
    "ICAgIGZsZXgtd3JhcDp3cmFwOyBtYXJnaW4tYm90dG9tOjIycHg7IH0KICBoMS50aXRsZSB7IGZvbnQtc2l6ZToyMnB4OyBmb250"
    "LXdlaWdodDo2NTA7IG1hcmdpbjowIDAgNHB4OyBsZXR0ZXItc3BhY2luZzotMC4wMWVtOyB9CiAgLnN1YiB7IGNvbG9yOnZhcigt"
    "LXRleHQtc2Vjb25kYXJ5KTsgZm9udC1zaXplOjEzcHg7IG1hcmdpbjowOyB9CiAgLmNvbnRyb2xzIHsgZGlzcGxheTpmbGV4OyBn"
    "YXA6OHB4OyBhbGlnbi1pdGVtczpjZW50ZXI7IGZsZXgtd3JhcDp3cmFwOyB9CiAgc2VsZWN0LCAudG9nZ2xlIHsgYmFja2dyb3Vu"
    "ZDp2YXIoLS1zdXJmYWNlLTEpOyBib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJvcmRlcik7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25k"
    "YXJ5KTsKICAgIGJvcmRlci1yYWRpdXM6OTk5cHg7IHBhZGRpbmc6OHB4IDE0cHg7IGZvbnQ6aW5oZXJpdDsgZm9udC1zaXplOjEy"
    "LjVweDsgY3Vyc29yOnBvaW50ZXI7IH0KICBzZWxlY3QgeyBib3JkZXItcmFkaXVzOjEwcHg7IH0KICAudG9nZ2xlOmhvdmVyLCBz"
    "ZWxlY3Q6aG92ZXIgeyBjb2xvcjp2YXIoLS10ZXh0LXByaW1hcnkpOyB9CgogIC8qIC0tLS0tLS0tLS0gbW9udGggY2hpcHMgLS0t"
    "LS0tLS0tLSAqLwogIC5tb250aHJvdyB7IGRpc3BsYXk6ZmxleDsgZ2FwOjZweDsgZmxleC13cmFwOndyYXA7IGFsaWduLWl0ZW1z"
    "OmNlbnRlcjsgbWFyZ2luOjAgMCAxMHB4OyB9CiAgLm1jaGlwIHsgZm9udDppbmhlcml0OyBmb250LXNpemU6MTIuNXB4OyBiYWNr"
    "Z3JvdW5kOnZhcigtLXN1cmZhY2UtMSk7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsKICAgIGJvcmRlcjoxcHggc29saWQg"
    "dmFyKC0tYm9yZGVyKTsgYm9yZGVyLXJhZGl1czo5OTlweDsgcGFkZGluZzo2cHggMTRweDsgY3Vyc29yOnBvaW50ZXI7IH0KICAu"
    "bWNoaXA6aG92ZXIgeyBjb2xvcjp2YXIoLS10ZXh0LXByaW1hcnkpOyBiYWNrZ3JvdW5kOnZhcigtLWhvdmVyKTsgfQogIC5tY2hp"
    "cFthcmlhLXByZXNzZWQ9InRydWUiXSB7IGJhY2tncm91bmQ6dmFyKC0tc2VyaWVzLTEpOyBib3JkZXItY29sb3I6dmFyKC0tc2Vy"
    "aWVzLTEpOyBjb2xvcjojZmZmOyB9CiAgLm1jaGlwLmlucmFuZ2UgeyBib3JkZXItY29sb3I6dmFyKC0tc2VyaWVzLTEpOyBjb2xv"
    "cjp2YXIoLS10ZXh0LXByaW1hcnkpOyB9CiAgLm1vbnRocm93IC5zZXAgeyB3aWR0aDoxcHg7IGhlaWdodDoyMHB4OyBiYWNrZ3Jv"
    "dW5kOnZhcigtLWdyaWQpOyBtYXJnaW46MCA1cHg7IH0KICAubW9udGhyb3cgLmhpbnQgeyBmb250LXNpemU6MTJweDsgY29sb3I6"
    "dmFyKC0tdGV4dC1tdXRlZCk7IG1hcmdpbi1sZWZ0OjRweDsgfQoKICAvKiAtLS0tLS0tLS0tIHByb3BlcnR5IGZpbHRlciAtLS0t"
    "LS0tLS0tICovCiAgLmZpbHRlcnJvdyB7IGRpc3BsYXk6ZmxleDsgYWxpZ24taXRlbXM6ZmxleC1zdGFydDsgZ2FwOjEwcHg7IGZs"
    "ZXgtd3JhcDp3cmFwOyBtYXJnaW46MCAwIDE2cHg7IH0KICAuZmlsdGVyd3JhcCB7IHBvc2l0aW9uOnJlbGF0aXZlOyB9CiAgLmZp"
    "bHRlcmJ0biB7IGJhY2tncm91bmQ6dmFyKC0tc3VyZmFjZS0xKTsgYm9yZGVyOjFweCBzb2xpZCB2YXIoLS1ib3JkZXIpOyBjb2xv"
    "cjp2YXIoLS10ZXh0LXNlY29uZGFyeSk7CiAgICBib3JkZXItcmFkaXVzOjEwcHg7IHBhZGRpbmc6OHB4IDE0cHg7IGZvbnQ6aW5o"
    "ZXJpdDsgZm9udC1zaXplOjEyLjVweDsgY3Vyc29yOnBvaW50ZXI7CiAgICBkaXNwbGF5OmlubGluZS1mbGV4OyBhbGlnbi1pdGVt"
    "czpjZW50ZXI7IGdhcDo4cHg7IH0KICAuZmlsdGVyYnRuOmhvdmVyIHsgY29sb3I6dmFyKC0tdGV4dC1wcmltYXJ5KTsgfQogIC5m"
    "aWx0ZXJidG4uYWN0aXZlIHsgYm9yZGVyLWNvbG9yOnZhcigtLXNlcmllcy0yKTsgY29sb3I6dmFyKC0tdGV4dC1wcmltYXJ5KTsg"
    "fQogIC5maWx0ZXJidG4gLmNhcmV0IHsgZm9udC1zaXplOjEwcHg7IGNvbG9yOnZhcigtLXRleHQtbXV0ZWQpOyB9CiAgLmZpbHRl"
    "cnBhbmVsIHsgcG9zaXRpb246YWJzb2x1dGU7IHotaW5kZXg6MzA7IHRvcDpjYWxjKDEwMCUgKyA2cHgpOyBsZWZ0OjA7IHdpZHRo"
    "OjI5MHB4OwogICAgbWF4LWhlaWdodDo2MHZoOyBvdmVyZmxvdy15OmF1dG87IGJhY2tncm91bmQ6dmFyKC0tc3VyZmFjZS0xKTsg"
    "Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1ib3JkZXIpOwogICAgYm9yZGVyLXJhZGl1czoxMnB4OyBwYWRkaW5nOjEwcHg7IGJveC1z"
    "aGFkb3c6MCAxMHB4IDMwcHggcmdiYSgwLDAsMCwuMTgpOyB9CiAgLmZpbHRlcnBhbmVsIC5maGVhZCB7IGRpc3BsYXk6ZmxleDsg"
    "Z2FwOjhweDsgcGFkZGluZzo0cHggNnB4IDEwcHg7IGJvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWdyaWQpOyBtYXJnaW4t"
    "Ym90dG9tOjZweDsgfQogIC5maWx0ZXJwYW5lbCAuZmhlYWQgYnV0dG9uIHsgZmxleDoxOyBiYWNrZ3JvdW5kOnRyYW5zcGFyZW50"
    "OyBib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJvcmRlcik7CiAgICBib3JkZXItcmFkaXVzOjhweDsgcGFkZGluZzo2cHg7IGZvbnQ6"
    "aW5oZXJpdDsgZm9udC1zaXplOjEycHg7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgY3Vyc29yOnBvaW50ZXI7IH0KICAu"
    "ZmlsdGVycGFuZWwgLmZoZWFkIGJ1dHRvbjpob3ZlciB7IGNvbG9yOnZhcigtLXRleHQtcHJpbWFyeSk7IGJhY2tncm91bmQ6dmFy"
    "KC0taG92ZXIpOyB9CiAgLmZncm91cCB7IGZvbnQtc2l6ZToxMS41cHg7IGNvbG9yOnZhcigtLXRleHQtbXV0ZWQpOyB0ZXh0LXRy"
    "YW5zZm9ybTp1cHBlcmNhc2U7IGxldHRlci1zcGFjaW5nOi4wNGVtOwogICAgcGFkZGluZzoxMHB4IDZweCA0cHg7IGRpc3BsYXk6"
    "ZmxleDsganVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47IGFsaWduLWl0ZW1zOmNlbnRlcjsgfQogIC5mZ3JvdXAgYnV0dG9u"
    "IHsgYmFja2dyb3VuZDp0cmFuc3BhcmVudDsgYm9yZGVyOjA7IGNvbG9yOnZhcigtLXRleHQtbXV0ZWQpOyBmb250OmluaGVyaXQ7"
    "CiAgICBmb250LXNpemU6MTFweDsgY3Vyc29yOnBvaW50ZXI7IHRleHQtZGVjb3JhdGlvbjp1bmRlcmxpbmU7IHRleHQtdW5kZXJs"
    "aW5lLW9mZnNldDoycHg7IH0KICAuZmdyb3VwIGJ1dHRvbjpob3ZlciB7IGNvbG9yOnZhcigtLXRleHQtcHJpbWFyeSk7IH0KICAu"
    "Zml0ZW0geyBkaXNwbGF5OmZsZXg7IGFsaWduLWl0ZW1zOmNlbnRlcjsgZ2FwOjlweDsgcGFkZGluZzo2cHggNnB4OyBib3JkZXIt"
    "cmFkaXVzOjhweDsKICAgIGZvbnQtc2l6ZToxM3B4OyBjdXJzb3I6cG9pbnRlcjsgfQogIC5maXRlbTpob3ZlciB7IGJhY2tncm91"
    "bmQ6dmFyKC0taG92ZXIpOyB9CiAgLmZpdGVtIGlucHV0IHsgYWNjZW50LWNvbG9yOnZhcigtLXNlcmllcy0xKTsgd2lkdGg6MTVw"
    "eDsgaGVpZ2h0OjE1cHg7IGN1cnNvcjpwb2ludGVyOyBmbGV4Om5vbmU7IH0KICAuZml0ZW0ub2ZmIHsgY29sb3I6dmFyKC0tdGV4"
    "dC1tdXRlZCk7IH0KICAuZml0ZW0gaW5wdXQ6ZGlzYWJsZWQgeyBjdXJzb3I6ZGVmYXVsdDsgb3BhY2l0eTouNTsgfQogIC5jaGlw"
    "cyB7IGRpc3BsYXk6ZmxleDsgZ2FwOjZweDsgZmxleC13cmFwOndyYXA7IGFsaWduLWl0ZW1zOmNlbnRlcjsgfQogIC5jaGlwcyAu"
    "bGJsIHsgZm9udC1zaXplOjEyLjVweDsgY29sb3I6dmFyKC0tdGV4dC1zZWNvbmRhcnkpOyB9CiAgLmNoaXAgeyBkaXNwbGF5Omlu"
    "bGluZS1mbGV4OyBhbGlnbi1pdGVtczpjZW50ZXI7IGdhcDo2cHg7IGZvbnQtc2l6ZToxMnB4OwogICAgYmFja2dyb3VuZDp2YXIo"
    "LS1zdXJmYWNlLTEpOyBib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJvcmRlcik7IGJvcmRlci1sZWZ0OjNweCBzb2xpZCB2YXIoLS1z"
    "ZXJpZXMtMSk7CiAgICBib3JkZXItcmFkaXVzOjk5OXB4OyBwYWRkaW5nOjRweCA2cHggNHB4IDEwcHg7IGNvbG9yOnZhcigtLXRl"
    "eHQtc2Vjb25kYXJ5KTsgfQogIC5jaGlwLm91dCB7IGJvcmRlci1sZWZ0LWNvbG9yOnZhcigtLXNlcmllcy0yKTsgfQogIC5tb3Jl"
    "YnRuIHsgYmFja2dyb3VuZDp0cmFuc3BhcmVudDsgYm9yZGVyOjA7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgZm9udDpp"
    "bmhlcml0OwogICAgZm9udC1zaXplOjEycHg7IGN1cnNvcjpwb2ludGVyOyB0ZXh0LWRlY29yYXRpb246dW5kZXJsaW5lOyB0ZXh0"
    "LXVuZGVybGluZS1vZmZzZXQ6M3B4OwogICAgcGFkZGluZzo0cHggNnB4OyBib3JkZXItcmFkaXVzOjhweDsgfQogIC5tb3JlYnRu"
    "OmhvdmVyIHsgY29sb3I6dmFyKC0tdGV4dC1wcmltYXJ5KTsgYmFja2dyb3VuZDp2YXIoLS1ob3Zlcik7IH0KICAuY2hpcCBidXR0"
    "b24geyBiYWNrZ3JvdW5kOnRyYW5zcGFyZW50OyBib3JkZXI6MDsgY29sb3I6dmFyKC0tdGV4dC1tdXRlZCk7IGN1cnNvcjpwb2lu"
    "dGVyOwogICAgZm9udDppbmhlcml0OyBmb250LXNpemU6MTRweDsgbGluZS1oZWlnaHQ6MTsgcGFkZGluZzowIDNweDsgYm9yZGVy"
    "LXJhZGl1czo1MCU7IH0KICAuY2hpcCBidXR0b246aG92ZXIgeyBjb2xvcjp2YXIoLS10ZXh0LXByaW1hcnkpOyBiYWNrZ3JvdW5k"
    "OnZhcigtLWhvdmVyKTsgfQogIC5lbXB0eXN0YXRlIHsgYmFja2dyb3VuZDp2YXIoLS1zdXJmYWNlLTEpOyBib3JkZXI6MXB4IHNv"
    "bGlkIHZhcigtLWJvcmRlcik7IGJvcmRlci1yYWRpdXM6MTRweDsKICAgIHBhZGRpbmc6NDBweCAyNHB4OyB0ZXh0LWFsaWduOmNl"
    "bnRlcjsgY29sb3I6dmFyKC0tdGV4dC1zZWNvbmRhcnkpOyB9CgogIC5oZXJvIHsgYmFja2dyb3VuZDp2YXIoLS1zdXJmYWNlLTEp"
    "OyBib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJvcmRlcik7IGJvcmRlci1yYWRpdXM6MTRweDsKICAgIHBhZGRpbmc6MjRweCAyNnB4"
    "OyBtYXJnaW4tYm90dG9tOjE2cHg7IGRpc3BsYXk6ZmxleDsgYWxpZ24taXRlbXM6ZmxleC1lbmQ7IGdhcDo0MHB4OyBmbGV4LXdy"
    "YXA6d3JhcDsgfQogIC5oZXJvIC5sYWJlbCB7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgZm9udC1zaXplOjEzcHg7IH0K"
    "ICAuaGVybyAudmFsdWUgeyBmb250LXNpemU6NTJweDsgZm9udC13ZWlnaHQ6NjQwOyBsZXR0ZXItc3BhY2luZzotMC4wMjVlbTsg"
    "bGluZS1oZWlnaHQ6MS4wNTsgbWFyZ2luLXRvcDoycHg7IH0KICAuaGVybyAuaGVyb25vdGUgeyBjb2xvcjp2YXIoLS10ZXh0LW11"
    "dGVkKTsgZm9udC1zaXplOjEyLjVweDsgbWFyZ2luLXRvcDo2cHg7IH0KICAuaGVyby1zaWRlIHsgZGlzcGxheTpmbGV4OyBnYXA6"
    "MzRweDsgZmxleC13cmFwOndyYXA7IHBhZGRpbmctYm90dG9tOjZweDsgfQogIC5oZXJvLXNpZGUgLmwgeyBjb2xvcjp2YXIoLS10"
    "ZXh0LXNlY29uZGFyeSk7IGZvbnQtc2l6ZToxMi41cHg7IH0KICAuaGVyby1zaWRlIC52IHsgZm9udC1zaXplOjIwcHg7IGZvbnQt"
    "d2VpZ2h0OjYwMDsgbGV0dGVyLXNwYWNpbmc6LTAuMDFlbTsgbWFyZ2luLXRvcDoycHg7IH0KCiAgLnRpbGVzIHsgZGlzcGxheTpn"
    "cmlkOyBncmlkLXRlbXBsYXRlLWNvbHVtbnM6cmVwZWF0KGF1dG8tZml0LG1pbm1heCgxNzBweCwxZnIpKTsgZ2FwOjEycHg7IG1h"
    "cmdpbi1ib3R0b206MjJweDsgfQogIC50aWxlIHsgYmFja2dyb3VuZDp2YXIoLS1zdXJmYWNlLTEpOyBib3JkZXI6MXB4IHNvbGlk"
    "IHZhcigtLWJvcmRlcik7IGJvcmRlci1yYWRpdXM6MTJweDsgcGFkZGluZzoxNnB4IDE4cHg7IH0KICAudGlsZSAubCB7IGNvbG9y"
    "OnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgZm9udC1zaXplOjEyLjVweDsgfQogIC50aWxlIC52IHsgZm9udC1zaXplOjI1cHg7IGZv"
    "bnQtd2VpZ2h0OjYyMDsgbGV0dGVyLXNwYWNpbmc6LTAuMDJlbTsgbWFyZ2luLXRvcDozcHg7IH0KICAudGlsZSAuZCB7IGZvbnQt"
    "c2l6ZToxMnB4OyBjb2xvcjp2YXIoLS10ZXh0LW11dGVkKTsgbWFyZ2luLXRvcDozcHg7IH0KICAucG9zIHsgY29sb3I6dmFyKC0t"
    "Z29vZCk7IH0gLm5lZyB7IGNvbG9yOnZhcigtLWNyaXRpY2FsKTsgfQoKICAuY2FyZCB7IGJhY2tncm91bmQ6dmFyKC0tc3VyZmFj"
    "ZS0xKTsgYm9yZGVyOjFweCBzb2xpZCB2YXIoLS1ib3JkZXIpOyBib3JkZXItcmFkaXVzOjE0cHg7CiAgICBwYWRkaW5nOjIycHgg"
    "MjRweCAxOHB4OyBtYXJnaW4tYm90dG9tOjE2cHg7IH0KICAuY2FyZCBoMiB7IGZvbnQtc2l6ZToxNXB4OyBmb250LXdlaWdodDo2"
    "MjA7IG1hcmdpbjowIDAgM3B4OyB9CiAgLmNhcmQgLmNhcCB7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgZm9udC1zaXpl"
    "OjEyLjVweDsgbWFyZ2luOjAgMCAxNnB4OyB9CiAgLmdyaWQyIHsgZGlzcGxheTpncmlkOyBncmlkLXRlbXBsYXRlLWNvbHVtbnM6"
    "MWZyIDFmcjsgZ2FwOjE2cHg7IH0KICBAbWVkaWEgKG1heC13aWR0aDo4ODBweCl7IC5ncmlkMntncmlkLXRlbXBsYXRlLWNvbHVt"
    "bnM6MWZyO30gLmhlcm8gLnZhbHVle2ZvbnQtc2l6ZTo0MnB4O30gfQoKICAubGVnZW5kIHsgZGlzcGxheTpmbGV4OyBnYXA6MThw"
    "eDsgZmxleC13cmFwOndyYXA7IG1hcmdpbjowIDAgMTBweDsgfQogIC5sZWdlbmQgc3BhbiB7IGRpc3BsYXk6aW5saW5lLWZsZXg7"
    "IGFsaWduLWl0ZW1zOmNlbnRlcjsgZ2FwOjdweDsgY29sb3I6dmFyKC0tdGV4dC1zZWNvbmRhcnkpOwogICAgZm9udC1zaXplOjEy"
    "LjVweDsgd2hpdGUtc3BhY2U6bm93cmFwOyB9CiAgLmtleSB7IHdpZHRoOjExcHg7IGhlaWdodDoxMXB4OyBib3JkZXItcmFkaXVz"
    "OjNweDsgZGlzcGxheTppbmxpbmUtYmxvY2s7IGZsZXg6bm9uZTsgfQogIC5rZXkubGluZSB7IGhlaWdodDozcHg7IHdpZHRoOjE1"
    "cHg7IGJvcmRlci1yYWRpdXM6MnB4OyB9CgogIHN2ZyB7IGRpc3BsYXk6YmxvY2s7IHdpZHRoOjEwMCU7IG92ZXJmbG93OnZpc2li"
    "bGU7IH0KICAudGljayB7IGZpbGw6dmFyKC0tdGV4dC1tdXRlZCk7IGZvbnQtc2l6ZToxMXB4OyBmb250LXZhcmlhbnQtbnVtZXJp"
    "Yzp0YWJ1bGFyLW51bXM7IH0KICAueGxhYiB7IGZpbGw6dmFyKC0tdGV4dC1zZWNvbmRhcnkpOyBmb250LXNpemU6MTEuNXB4OyB9"
    "CiAgLmRsYWIgeyBmaWxsOnZhcigtLXRleHQtcHJpbWFyeSk7IGZvbnQtc2l6ZToxMS41cHg7IGZvbnQtd2VpZ2h0OjYwMDsgfQog"
    "IC5kaW0geyBvcGFjaXR5OjAuMjsgfQogIDpyb290W2RhdGEtdGhlbWU9ImRhcmsiXSAuZGltIHsgb3BhY2l0eTowLjI2OyB9CiAg"
    "LnhsYWIub24geyBmaWxsOnZhcigtLXRleHQtcHJpbWFyeSk7IGZvbnQtd2VpZ2h0OjY1MDsgfQogIC5ncmlkbGluZSB7IHN0cm9r"
    "ZTp2YXIoLS1ncmlkKTsgc3Ryb2tlLXdpZHRoOjE7IH0KICAuYmFzZWxpbmUgeyBzdHJva2U6dmFyKC0tYXhpcyk7IHN0cm9rZS13"
    "aWR0aDoxOyB9CiAgLnJvd2hpdCB7IGZpbGw6dHJhbnNwYXJlbnQ7IGN1cnNvcjpwb2ludGVyOyB9CiAgLnJvd2hpdDpob3ZlciB7"
    "IGZpbGw6dmFyKC0taG92ZXIpOyB9CgogIC50aXAgeyBwb3NpdGlvbjpmaXhlZDsgcG9pbnRlci1ldmVudHM6bm9uZTsgei1pbmRl"
    "eDo0MDsgb3BhY2l0eTowOyB0cmFuc2l0aW9uOm9wYWNpdHkgLjFzOwogICAgYmFja2dyb3VuZDp2YXIoLS1zdXJmYWNlLTEpOyBi"
    "b3JkZXI6MXB4IHNvbGlkIHZhcigtLWJvcmRlcik7IGJvcmRlci1yYWRpdXM6MTBweDsgcGFkZGluZzo5cHggMTFweDsKICAgIGZv"
    "bnQtc2l6ZToxMi41cHg7IGJveC1zaGFkb3c6MCA2cHggMjBweCByZ2JhKDAsMCwwLC4xNCk7IG1pbi13aWR0aDoxNThweDsgfQog"
    "IC50aXAgLnQgeyBmb250LXdlaWdodDo2MjA7IG1hcmdpbi1ib3R0b206NXB4OyB9CiAgLnRpcCAuciB7IGRpc3BsYXk6ZmxleDsg"
    "YWxpZ24taXRlbXM6Y2VudGVyOyBnYXA6MTBweDsganVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47IGNvbG9yOnZhcigtLXRl"
    "eHQtc2Vjb25kYXJ5KTsgfQogIC50aXAgLnIgYiB7IGNvbG9yOnZhcigtLXRleHQtcHJpbWFyeSk7IGZvbnQtd2VpZ2h0OjYwMDsg"
    "Zm9udC12YXJpYW50LW51bWVyaWM6dGFidWxhci1udW1zOyB9CiAgLnRpcCAuciAubm0geyBkaXNwbGF5OmlubGluZS1mbGV4OyBh"
    "bGlnbi1pdGVtczpjZW50ZXI7IGdhcDo2cHg7IH0KCiAgZGV0YWlscy50YWJsZXdyYXAgeyBiYWNrZ3JvdW5kOnZhcigtLXN1cmZh"
    "Y2UtMSk7IGJvcmRlcjoxcHggc29saWQgdmFyKC0tYm9yZGVyKTsKICAgIGJvcmRlci1yYWRpdXM6MTRweDsgcGFkZGluZzoxOHB4"
    "IDI0cHg7IG1hcmdpbi1ib3R0b206MTZweDsgfQogIGRldGFpbHMudGFibGV3cmFwIHN1bW1hcnkgeyBjdXJzb3I6cG9pbnRlcjsg"
    "Zm9udC13ZWlnaHQ6NjAwOyBmb250LXNpemU6MTRweDsgfQogIC5zY3JvbGxlciB7IG92ZXJmbG93LXg6YXV0bzsgbWFyZ2luLXRv"
    "cDoxNHB4OyB9CiAgdGFibGUgeyBib3JkZXItY29sbGFwc2U6Y29sbGFwc2U7IHdpZHRoOjEwMCU7IGZvbnQtc2l6ZToxMi41cHg7"
    "IGZvbnQtdmFyaWFudC1udW1lcmljOnRhYnVsYXItbnVtczsgfQogIHRoLHRkIHsgcGFkZGluZzo3cHggMTBweDsgdGV4dC1hbGln"
    "bjpyaWdodDsgd2hpdGUtc3BhY2U6bm93cmFwOyBib3JkZXItYm90dG9tOjFweCBzb2xpZCB2YXIoLS1ncmlkKTsgfQogIHRoOmZp"
    "cnN0LWNoaWxkLCB0ZDpmaXJzdC1jaGlsZCB7IHRleHQtYWxpZ246bGVmdDsgZm9udC12YXJpYW50LW51bWVyaWM6bm9ybWFsOwog"
    "ICAgcG9zaXRpb246c3RpY2t5OyBsZWZ0OjA7IGJhY2tncm91bmQ6dmFyKC0tc3VyZmFjZS0xKTsgfQogIHRoZWFkIHRoIHsgY29s"
    "b3I6dmFyKC0tdGV4dC1zZWNvbmRhcnkpOyBmb250LXdlaWdodDo2MDA7IGJvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWF4"
    "aXMpOyB9CiAgdHIuc2VjdGlvbiB0ZCB7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgZm9udC13ZWlnaHQ6NjAwOyBwYWRk"
    "aW5nLXRvcDoxNHB4OyB9CiAgdHIudG90YWwgdGQgeyBmb250LXdlaWdodDo2NDA7IGJvcmRlci10b3A6MXB4IHNvbGlkIHZhcigt"
    "LWF4aXMpOyB9CiAgdGQuaW5kZW50IHsgcGFkZGluZy1sZWZ0OjI0cHg7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsgfQog"
    "IHRib2R5IHRyLmNsaWNrYWJsZSB7IGN1cnNvcjpwb2ludGVyOyB9CiAgdGJvZHkgdHIuY2xpY2thYmxlOmhvdmVyIHRkIHsgYmFj"
    "a2dyb3VuZDp2YXIoLS1ob3Zlcik7IH0KCiAgLnRhYmxldG9vbHMgeyBkaXNwbGF5OmZsZXg7IGp1c3RpZnktY29udGVudDpmbGV4"
    "LWVuZDsgZ2FwOjhweDsgbWFyZ2luLXRvcDoxMnB4OyB9CiAgLmRsYnRuIHsgYmFja2dyb3VuZDp2YXIoLS1zdXJmYWNlLTEpOyBi"
    "b3JkZXI6MXB4IHNvbGlkIHZhcigtLWJvcmRlcik7IGNvbG9yOnZhcigtLXRleHQtc2Vjb25kYXJ5KTsKICAgIGJvcmRlci1yYWRp"
    "dXM6OXB4OyBwYWRkaW5nOjdweCAxM3B4OyBmb250OmluaGVyaXQ7IGZvbnQtc2l6ZToxMi41cHg7IGN1cnNvcjpwb2ludGVyOwog"
    "ICAgZGlzcGxheTppbmxpbmUtZmxleDsgYWxpZ24taXRlbXM6Y2VudGVyOyBnYXA6N3B4OyB9CiAgLmRsYnRuOmhvdmVyIHsgY29s"
    "b3I6dmFyKC0tdGV4dC1wcmltYXJ5KTsgYmFja2dyb3VuZDp2YXIoLS1ob3Zlcik7IH0KICAuZGxidG4uZG9uZSB7IGNvbG9yOnZh"
    "cigtLWdvb2QpOyBib3JkZXItY29sb3I6dmFyKC0tZ29vZCk7IH0KCiAgLm5vdGUgeyBiYWNrZ3JvdW5kOnZhcigtLXN1cmZhY2Ut"
    "MSk7IGJvcmRlcjoxcHggc29saWQgdmFyKC0tYm9yZGVyKTsKICAgIGJvcmRlci1sZWZ0OjNweCBzb2xpZCB2YXIoLS13YXJuaW5n"
    "KTsgYm9yZGVyLXJhZGl1czoxMHB4OyBwYWRkaW5nOjE0cHggMThweDsKICAgIGZvbnQtc2l6ZToxMi41cHg7IGNvbG9yOnZhcigt"
    "LXRleHQtc2Vjb25kYXJ5KTsgbWFyZ2luLWJvdHRvbToxNnB4OyB9CiAgLm5vdGUgYiB7IGNvbG9yOnZhcigtLXRleHQtcHJpbWFy"
    "eSk7IH0KICBmb290ZXIgeyBjb2xvcjp2YXIoLS10ZXh0LW11dGVkKTsgZm9udC1zaXplOjEycHg7IG1hcmdpbi10b3A6MjJweDsg"
    "fQo8L3N0eWxlPgo8L2hlYWQ+Cjxib2R5PgoKPGRpdiBpZD0ibG9jayI+CiAgPGZvcm0gY2xhc3M9ImxvY2tjYXJkIiBpZD0ibG9j"
    "a2Zvcm0iPgogICAgPGgxPlRBR0xZWiBQb3J0Zm9saW88L2gxPgogICAgPHA+VGhpcyByZXBvcnQgaXMgZW5jcnlwdGVkLiBFbnRl"
    "ciB0aGUgcGFzc3BocmFzZSB0byB2aWV3IGl0LjwvcD4KICAgIDxsYWJlbCBmb3I9InB3Ij5QYXNzcGhyYXNlPC9sYWJlbD4KICAg"
    "IDxpbnB1dCB0eXBlPSJwYXNzd29yZCIgaWQ9InB3IiBhdXRvY29tcGxldGU9ImN1cnJlbnQtcGFzc3dvcmQiIGF1dG9mb2N1cz4K"
    "ICAgIDxidXR0b24gdHlwZT0ic3VibWl0IiBpZD0idW5sb2NrIj5VbmxvY2s8L2J1dHRvbj4KICAgIDxkaXYgY2xhc3M9ImVyciIg"
    "aWQ9ImVyciI+PC9kaXY+CiAgPC9mb3JtPgo8L2Rpdj4KCjxkaXYgY2xhc3M9IndyYXAiIGlkPSJhcHAiIGhpZGRlbj4KICA8aGVh"
    "ZGVyIGNsYXNzPSJ0b3AiPgogICAgPGRpdj4KICAgICAgPGgxIGNsYXNzPSJ0aXRsZSIgaWQ9InZpZXdUaXRsZSI+UG9ydGZvbGlv"
    "PC9oMT4KICAgICAgPHAgY2xhc3M9InN1YiIgaWQ9InZpZXdTdWIiPjwvcD4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0iY29u"
    "dHJvbHMiPgogICAgICA8c2VsZWN0IGlkPSJwcm9wU2VsIiBhcmlhLWxhYmVsPSJDaG9vc2UgYSB2aWV3Ij48L3NlbGVjdD4KICAg"
    "ICAgPGJ1dHRvbiBjbGFzcz0idG9nZ2xlIiBpZD0idGhlbWVCdG4iIHR5cGU9ImJ1dHRvbiI+RGFyayBtb2RlPC9idXR0b24+CiAg"
    "ICA8L2Rpdj4KICA8L2hlYWRlcj4KICA8ZGl2IGNsYXNzPSJtb250aHJvdyIgaWQ9Im1vbnRoUm93IiByb2xlPSJncm91cCIgYXJp"
    "YS1sYWJlbD0iQ2hvb3NlIG1vbnRocyI+PC9kaXY+CiAgPGRpdiBjbGFzcz0iZmlsdGVycm93IiBpZD0iZmlsdGVyUm93IiBoaWRk"
    "ZW4+CiAgICA8ZGl2IGNsYXNzPSJmaWx0ZXJ3cmFwIj4KICAgICAgPGJ1dHRvbiBjbGFzcz0iZmlsdGVyYnRuIiBpZD0iZmlsdGVy"
    "QnRuIiB0eXBlPSJidXR0b24iIGFyaWEtZXhwYW5kZWQ9ImZhbHNlIiBhcmlhLWNvbnRyb2xzPSJmaWx0ZXJQYW5lbCI+CiAgICAg"
    "ICAgPHNwYW4gaWQ9ImZpbHRlckxhYmVsIj5BbGwgcHJvcGVydGllczwvc3Bhbj48c3BhbiBjbGFzcz0iY2FyZXQiPiYjOTY2Mjs8"
    "L3NwYW4+CiAgICAgIDwvYnV0dG9uPgogICAgICA8ZGl2IGNsYXNzPSJmaWx0ZXJwYW5lbCIgaWQ9ImZpbHRlclBhbmVsIiBoaWRk"
    "ZW4gcm9sZT0iZ3JvdXAiIGFyaWEtbGFiZWw9IkNob29zZSBwcm9wZXJ0aWVzIHRvIGluY2x1ZGUiPjwvZGl2PgogICAgPC9kaXY+"
    "CiAgICA8ZGl2IGNsYXNzPSJjaGlwcyIgaWQ9ImZpbHRlckNoaXBzIj48L2Rpdj4KICA8L2Rpdj4KICA8ZGl2IGlkPSJib2R5Ij48"
    "L2Rpdj4KICA8Zm9vdGVyIGlkPSJmb290Ij48L2Zvb3Rlcj4KPC9kaXY+Cgo8ZGl2IGNsYXNzPSJ0aXAiIGlkPSJ0aXAiPjwvZGl2"
    "PgoKPHNjcmlwdD4KInVzZSBzdHJpY3QiOwpjb25zdCBCTE9CID0gIl9fUEFZTE9BRF9fIjsKY29uc3QgTUVUQSA9IF9fTUVUQV9f"
    "OwoKLyogPT09PT09PT09PT09PT09PT09PT09PT09PT09PSBkZWNyeXB0aW9uID09PT09PT09PT09PT09PT09PT09PT09PT09PT0g"
    "Ki8KY29uc3QgYjY0ID0gcyA9PiBVaW50OEFycmF5LmZyb20oYXRvYihzKSwgYyA9PiBjLmNoYXJDb2RlQXQoMCkpOwoKYXN5bmMg"
    "ZnVuY3Rpb24gZGVjcnlwdChwYXNzKSB7CiAgY29uc3QgcmF3ID0gYjY0KEJMT0IpOwogIGNvbnN0IHNhbHQgPSByYXcuc2xpY2Uo"
    "MCwgMTYpLCBpdiA9IHJhdy5zbGljZSgxNiwgMjgpLCBib2R5ID0gcmF3LnNsaWNlKDI4KTsKICBjb25zdCBiYXNlID0gYXdhaXQg"
    "Y3J5cHRvLnN1YnRsZS5pbXBvcnRLZXkoInJhdyIsIG5ldyBUZXh0RW5jb2RlcigpLmVuY29kZShwYXNzKSwKICAgICJQQktERjIi"
    "LCBmYWxzZSwgWyJkZXJpdmVLZXkiXSk7CiAgY29uc3Qga2V5ID0gYXdhaXQgY3J5cHRvLnN1YnRsZS5kZXJpdmVLZXkoCiAgICB7"
    "IG5hbWU6ICJQQktERjIiLCBzYWx0LCBpdGVyYXRpb25zOiBNRVRBLml0ZXJhdGlvbnMsIGhhc2g6ICJTSEEtMjU2IiB9LAogICAg"
    "YmFzZSwgeyBuYW1lOiAiQUVTLUdDTSIsIGxlbmd0aDogMjU2IH0sIGZhbHNlLCBbImRlY3J5cHQiXSk7CiAgY29uc3QgcGxhaW4g"
    "PSBhd2FpdCBjcnlwdG8uc3VidGxlLmRlY3J5cHQoeyBuYW1lOiAiQUVTLUdDTSIsIGl2IH0sIGtleSwgYm9keSk7CiAgcmV0dXJu"
    "IEpTT04ucGFyc2UobmV3IFRleHREZWNvZGVyKCkuZGVjb2RlKHBsYWluKSk7Cn0KCmNvbnN0IGZvcm0gPSBkb2N1bWVudC5nZXRF"
    "bGVtZW50QnlJZCgibG9ja2Zvcm0iKTsKZm9ybS5hZGRFdmVudExpc3RlbmVyKCJzdWJtaXQiLCBhc3luYyBlID0+IHsKICBlLnBy"
    "ZXZlbnREZWZhdWx0KCk7CiAgY29uc3QgYnRuID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoInVubG9jayIpLCBlcnIgPSBkb2N1"
    "bWVudC5nZXRFbGVtZW50QnlJZCgiZXJyIik7CiAgYnRuLmRpc2FibGVkID0gdHJ1ZTsgYnRuLnRleHRDb250ZW50ID0gIkRlY3J5"
    "cHRpbmfigKYiOyBlcnIudGV4dENvbnRlbnQgPSAiIjsKICB0cnkgewogICAgUCA9IGF3YWl0IGRlY3J5cHQoZG9jdW1lbnQuZ2V0"
    "RWxlbWVudEJ5SWQoInB3IikudmFsdWUpOwogICAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoImxvY2siKS5oaWRkZW4gPSB0cnVl"
    "OwogICAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoImFwcCIpLmhpZGRlbiA9IGZhbHNlOwogICAgYm9vdCgpOwogIH0gY2F0Y2gg"
    "KF8pIHsKICAgIGVyci50ZXh0Q29udGVudCA9ICJUaGF0IHBhc3NwaHJhc2UgZGlkbid0IHdvcmsuIjsKICAgIGJ0bi5kaXNhYmxl"
    "ZCA9IGZhbHNlOyBidG4udGV4dENvbnRlbnQgPSAiVW5sb2NrIjsKICAgIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCJwdyIpLnNl"
    "bGVjdCgpOwogIH0KfSk7CgovKiA9PT09PT09PT09PT09PT09PT09PT09PT09PT09IGhlbHBlcnMgPT09PT09PT09PT09PT09PT09"
    "PT09PT09PT09PSAqLwpsZXQgUCA9IG51bGw7CmNvbnN0IE5TID0gImh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIjsKY29uc3Qg"
    "c3VtID0gYSA9PiBhLnJlZHVjZSgoeCwgeSkgPT4geCArIHksIDApOwpjb25zdCBtb25leSA9IHYgPT4gKHYgPCAwID8gIi0kIiA6"
    "ICIkIikgKyBNYXRoLmFicyh2KS50b0xvY2FsZVN0cmluZygiZW4tVVMiLCB7IG1heGltdW1GcmFjdGlvbkRpZ2l0czogMCB9KTsK"
    "Y29uc3QgbW9uZXkyID0gdiA9PiAodiA8IDAgPyAiLSQiIDogIiQiKSArIE1hdGguYWJzKHYpLnRvTG9jYWxlU3RyaW5nKCJlbi1V"
    "UyIsIHsgbWluaW11bUZyYWN0aW9uRGlnaXRzOiAyLCBtYXhpbXVtRnJhY3Rpb25EaWdpdHM6IDIgfSk7CmNvbnN0IGNvbXBhY3Qg"
    "PSB2ID0+IHsKICBpZiAoTWF0aC5hYnModikgPCAxMDAwKSByZXR1cm4gbW9uZXkodik7CiAgY29uc3QgayA9IChNYXRoLmFicyh2"
    "KSAvIDEwMDApLnRvRml4ZWQoTWF0aC5hYnModikgPCAxMDAwMCA/IDEgOiAwKS5yZXBsYWNlKC9cLjAkLywgIiIpOwogIHJldHVy"
    "biAodiA8IDAgPyAiLSQiIDogIiQiKSArIGsgKyAiayI7Cn07CmNvbnN0IGNzc3YgPSBuID0+IGdldENvbXB1dGVkU3R5bGUoZG9j"
    "dW1lbnQuZG9jdW1lbnRFbGVtZW50KS5nZXRQcm9wZXJ0eVZhbHVlKG4pLnRyaW0oKTsKY29uc3QgZXNjID0gcyA9PiBTdHJpbmco"
    "cykucmVwbGFjZSgvWyY8PiJdL2csIGMgPT4gKHsgIiYiOiAiJmFtcDsiLCAiPCI6ICImbHQ7IiwgIj4iOiAiJmd0OyIsICciJzog"
    "IiZxdW90OyIgfVtjXSkpOwoKZnVuY3Rpb24gZWwodGFnLCBhdHRycywgcGFyZW50KSB7CiAgY29uc3QgZSA9IGRvY3VtZW50LmNy"
    "ZWF0ZUVsZW1lbnROUyhOUywgdGFnKTsKICBmb3IgKGNvbnN0IGsgaW4gYXR0cnMpIGUuc2V0QXR0cmlidXRlKGssIGF0dHJzW2td"
    "KTsKICBpZiAocGFyZW50KSBwYXJlbnQuYXBwZW5kQ2hpbGQoZSk7CiAgcmV0dXJuIGU7Cn0KZnVuY3Rpb24gbmljZVRpY2tzKG1h"
    "eCwgbWF4VGlja3MpIHsKICBpZiAobWF4IDw9IDApIHJldHVybiBbMCwgMV07CiAgY29uc3QgbWFnID0gTWF0aC5wb3coMTAsIE1h"
    "dGguZmxvb3IoTWF0aC5sb2cxMChtYXgpKSAtIDEpOwogIGxldCBzdGVwID0gbWFnOwogIGZvciAoY29uc3QgcyBvZiBbMSwgMiwg"
    "Mi41LCA1LCAxMCwgMjAsIDI1LCA1MCwgMTAwLCAyMDAsIDI1MCwgNTAwXSkgewogICAgc3RlcCA9IHMgKiBtYWc7CiAgICBpZiAo"
    "TWF0aC5jZWlsKG1heCAvIHN0ZXApICsgMSA8PSBtYXhUaWNrcykgYnJlYWs7CiAgfQogIGNvbnN0IG91dCA9IFtdOyBsZXQgdiA9"
    "IDA7CiAgd2hpbGUgKHYgPCBtYXggLSAxZS05KSB7IG91dC5wdXNoKHYpOyB2ICs9IHN0ZXA7IH0KICBvdXQucHVzaCh2KTsKICBy"
    "ZXR1cm4gb3V0Owp9CmZ1bmN0aW9uIGNvbFBhdGgoeCwgeSwgdywgaCwgcikgewogIHIgPSBNYXRoLm1pbihyLCB3IC8gMiwgTWF0"
    "aC5tYXgoaCwgMCkpOwogIGlmIChoIDw9IDAuNSkgcmV0dXJuIGBNJHt4fSAke3kgKyBofSBoJHt3fWA7CiAgcmV0dXJuIGBNJHt4"
    "fSAke3kgKyBofSBWJHt5ICsgcn0gYSR7cn0gJHtyfSAwIDAgMSAke3J9ICR7LXJ9IGgke3cgLSAyICogcn0gYSR7cn0gJHtyfSAw"
    "IDAgMSAke3J9ICR7cn0gViR7eSArIGh9IFpgOwp9Cgpjb25zdCB0aXAgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgidGlwIik7"
    "CmZ1bmN0aW9uIHNob3dUaXAoaHRtbCwgZXZ0KSB7CiAgdGlwLmlubmVySFRNTCA9IGh0bWw7IHRpcC5zdHlsZS5vcGFjaXR5ID0g"
    "MTsKICBjb25zdCBwYWQgPSAxNCwgciA9IHRpcC5nZXRCb3VuZGluZ0NsaWVudFJlY3QoKTsKICBsZXQgeCA9IGV2dC5jbGllbnRY"
    "ICsgcGFkLCB5ID0gZXZ0LmNsaWVudFkgKyBwYWQ7CiAgaWYgKHggKyByLndpZHRoID4gaW5uZXJXaWR0aCAtIDgpIHggPSBldnQu"
    "Y2xpZW50WCAtIHIud2lkdGggLSBwYWQ7CiAgaWYgKHkgKyByLmhlaWdodCA+IGlubmVySGVpZ2h0IC0gOCkgeSA9IGV2dC5jbGll"
    "bnRZIC0gci5oZWlnaHQgLSBwYWQ7CiAgdGlwLnN0eWxlLmxlZnQgPSB4ICsgInB4IjsgdGlwLnN0eWxlLnRvcCA9IE1hdGgubWF4"
    "KDgsIHkpICsgInB4IjsKfQpjb25zdCBoaWRlVGlwID0gKCkgPT4geyB0aXAuc3R5bGUub3BhY2l0eSA9IDA7IH07CmNvbnN0IHRp"
    "cFJvdyA9IChjLCBuLCB2KSA9PiBgPGRpdiBjbGFzcz0iciI+PHNwYW4gY2xhc3M9Im5tIj48aSBjbGFzcz0ia2V5IiBzdHlsZT0i"
    "YmFja2dyb3VuZDoke2N9Ij48L2k+JHtufTwvc3Bhbj48Yj4ke3Z9PC9iPjwvZGl2PmA7CmZ1bmN0aW9uIGF0dGFjaFRpcChub2Rl"
    "LCBidWlsZCkgewogIG5vZGUuYWRkRXZlbnRMaXN0ZW5lcigibW91c2Vtb3ZlIiwgZSA9PiBzaG93VGlwKGJ1aWxkKCksIGUpKTsK"
    "ICBub2RlLmFkZEV2ZW50TGlzdGVuZXIoIm1vdXNlbGVhdmUiLCBoaWRlVGlwKTsKICBub2RlLmFkZEV2ZW50TGlzdGVuZXIoImZv"
    "Y3VzIiwgKCkgPT4gewogICAgY29uc3QgYiA9IG5vZGUuZ2V0Qm91bmRpbmdDbGllbnRSZWN0KCk7CiAgICBzaG93VGlwKGJ1aWxk"
    "KCksIHsgY2xpZW50WDogYi5sZWZ0ICsgYi53aWR0aCAvIDIsIGNsaWVudFk6IGIudG9wIH0pOwogIH0pOwogIG5vZGUuYWRkRXZl"
    "bnRMaXN0ZW5lcigiYmx1ciIsIGhpZGVUaXApOwogIG5vZGUuc2V0QXR0cmlidXRlKCJ0YWJpbmRleCIsICIwIik7Cn0KCi8qID09"
    "PT09PT09PT09PT09PT09PT09PT09PT09PT0gZGF0YSBzaGFwaW5nID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gKi8KY29u"
    "c3QgR1JPVVBfT0YgPSB7fTsKZnVuY3Rpb24gaW5pdEdyb3VwcygpIHsKICBmb3IgKGNvbnN0IGcgaW4gUC5ncm91cHMpIGZvciAo"
    "Y29uc3QgbGFiIG9mIFAuZ3JvdXBzW2ddKSBHUk9VUF9PRltsYWJdID0gZzsKfQpjb25zdCBHUk9VUF9OQU1FUyA9IFsiVGF4ZXMg"
    "JiBpbnN1cmFuY2UiLCAiUmVwYWlycyAmIGhhbmR5bWFuIiwgIlV0aWxpdGllcyIsICJNYW5hZ2VtZW50ICYgYWRtaW4iXTsKY29u"
    "c3QgR1JPVVBfVkFSID0gWyItLXNlcmllcy0xIiwgIi0tc2VyaWVzLTIiLCAiLS1zZXJpZXMtMyIsICItLXNlcmllcy00Il07Cgov"
    "KiAtLS0tIHZpZXcgbW9kZWwgLS0tLQogICBBIHBvcnRmb2xpbyB2aWV3IGlzIGEgU0VUIG9mIHByb3BlcnRpZXMuICJBbGwgcHJv"
    "cGVydGllcyIgaXMganVzdCB0aGUgc2V0IG9mCiAgIGV2ZXJ5dGhpbmc7IGFuIGVudGl0eSBwcmVzZXQgaXMgdGhhdCBlbnRpdHkn"
    "cyBtZW1iZXJzOyBhIHNhdmVkIGdyb3VwIGlzIGEgc2V0CiAgIHRoZSB2aWV3ZXIgbmFtZWQuIFNpbmdsZS1wcm9wZXJ0eSB2aWV3"
    "cyBpZ25vcmUgdGhlIHNldCBlbnRpcmVseSAtIGV4Y2x1ZGluZyBhCiAgIHByb3BlcnR5IGZyb20gdGhlIHJvbGwtdXAgZG9lcyBu"
    "b3QgaGlkZSBpdHMgb3duIHBhZ2UuICovCmNvbnN0IFZJRVcgPSB7IHR5cGU6ICJwb3J0Zm9saW8iLCBsYWJlbDogIlBvcnRmb2xp"
    "byIsIGluY2x1ZGU6IG51bGwgfTsgIC8vIGluY2x1ZGU6bnVsbCA9IGFsbApjb25zdCBhbGxQcm9wcyA9ICgpID0+IE9iamVjdC5r"
    "ZXlzKFAucHJvcGVydGllcyk7CmNvbnN0IGluY2x1ZGVkUHJvcHMgPSAoKSA9PgogIFZJRVcuaW5jbHVkZSA/IGFsbFByb3BzKCku"
    "ZmlsdGVyKHAgPT4gVklFVy5pbmNsdWRlLmhhcyhwKSkgOiBhbGxQcm9wcygpOwpjb25zdCBpc0ZpbHRlcmVkID0gKCkgPT4gISFW"
    "SUVXLmluY2x1ZGUgJiYgVklFVy5pbmNsdWRlLnNpemUgPCBhbGxQcm9wcygpLmxlbmd0aDsKCi8qIFNhdmVkIGdyb3VwcyBsaXZl"
    "IGluIHRoaXMgdmlld2VyJ3MgYnJvd3NlciBvbmx5LiBUaGV5IHN1cnZpdmUgdGhlIG1vbnRobHkKICAgcmVidWlsZCAobm90aGlu"
    "ZyBhYm91dCB0aGVtIGlzIGJha2VkIGludG8gdGhlIHBhZ2UpIGJ1dCB0aGV5IGRvIG5vdCBmb2xsb3cgdGhlCiAgIHZpZXdlciB0"
    "byBhbm90aGVyIGRldmljZSwgYW5kIGFub3RoZXIgdmlld2VyIHNlZXMgdGhlaXIgb3duLiBTaGFyaW5nIGEgZ3JvdXAKICAgbWVh"
    "bnMgc2hhcmluZyB0aGUgVVJMLCB3aGljaCBlbmNvZGVzIHRoZSBzZWxlY3Rpb24uICovCi8qIC0tLS0gbW9udGggc2VsZWN0aW9u"
    "IC0tLS0KICAgW3N0YXJ0LCBlbmRdIGluY2x1c2l2ZSBpbmRpY2VzIGludG8gUC5tb250aHM7IHN0YXJ0PT09ZW5kIGlzIGEgc2lu"
    "Z2xlIG1vbnRoLgogICBDaGFydHMga2VlcCBldmVyeSBtb250aCBhbmQgZGltIHdoYXQgaXMgb3V0c2lkZSB0aGlzIHdpbmRvdzsg"
    "ZXZlcnkgTlVNQkVSIC0KICAgaGVhZGxpbmUsIHRpbGVzLCByYW5raW5ncywgdGFibGUsIENTViAtIGNvdmVycyBvbmx5IHRoZSBz"
    "ZWxlY3Rpb24uICovCmxldCBNU0VMID0gbnVsbDsgICAgICAgICAgICAgICAgICAgICAgIC8vIHNldCBvbmNlIFAgaXMga25vd24K"
    "Y29uc3QgbUFsbCA9ICgpID0+IFswLCBQLm1vbnRocy5sZW5ndGggLSAxXTsKY29uc3QgbUZpbHRlcmVkID0gKCkgPT4gTVNFTCAm"
    "JiAoTVNFTFswXSAhPT0gMCB8fCBNU0VMWzFdICE9PSBQLm1vbnRocy5sZW5ndGggLSAxKTsKY29uc3QgbUxhYmVsID0gKCkgPT4g"
    "TVNFTFswXSA9PT0gTVNFTFsxXQogID8gYCR7UC5tb250aHNbTVNFTFswXV19ICR7UC55ZWFyfWAKICA6IGAke1AubW9udGhzW01T"
    "RUxbMF1dfVx1MjAxMyR7UC5tb250aHNbTVNFTFsxXV19ICR7UC55ZWFyfWA7Cgpjb25zdCBHS0VZID0gInRhZ2x5ei5ncm91cHMu"
    "djEiOwpmdW5jdGlvbiBsb2FkR3JvdXBzKCkgewogIHRyeSB7IHJldHVybiBKU09OLnBhcnNlKGxvY2FsU3RvcmFnZS5nZXRJdGVt"
    "KEdLRVkpKSB8fCBbXTsgfSBjYXRjaCAoXykgeyByZXR1cm4gW107IH0KfQpmdW5jdGlvbiBzYXZlR3JvdXBzKGcpIHsKICB0cnkg"
    "eyBsb2NhbFN0b3JhZ2Uuc2V0SXRlbShHS0VZLCBKU09OLnN0cmluZ2lmeShnKSk7IHJldHVybiB0cnVlOyB9IGNhdGNoIChfKSB7"
    "IHJldHVybiBmYWxzZTsgfQp9CmxldCBHUk9VUFMgPSBbXTsKCi8qIEEgInNlcmllcyIgaXMgdGhlIHNoYXBlIGV2ZXJ5IGNoYXJ0"
    "IGNvbnN1bWVzLCBmb3Igb25lIHByb3BlcnR5IG9yIHRoZSB3aG9sZSBwb3J0Zm9saW8uICovCmZ1bmN0aW9uIHNlcmllc0Zvcihu"
    "YW1lKSB7CiAgY29uc3QgbiA9IFAubW9udGhzLmxlbmd0aDsKICBjb25zdCB6ZXJvcyA9ICgpID0+IG5ldyBBcnJheShuKS5maWxs"
    "KDApOwogIGNvbnN0IHMgPSB7CiAgICBuYW1lLCBtb250aHM6IFAubW9udGhzLAogICAgaW5jb21lOiB6ZXJvcygpLCBleHBlbnNl"
    "czogemVyb3MoKSwgZXhwZW5zZXNSZWNvcmRlZDogemVyb3MoKSwKICAgIG5vaTogemVyb3MoKSwgZGVidDogemVyb3MoKSwgY2Fw"
    "ZXg6IHplcm9zKCksIGNhc2hmbG93OiB6ZXJvcygpLCB2YXJpYW5jZTogemVyb3MoKSwKICAgIHJlbnQ6IHplcm9zKCksIGxpbmVz"
    "OiB7fSwgZ3JvdXBzOiB7fQogIH07CiAgR1JPVVBfTkFNRVMuZm9yRWFjaChnID0+IHMuZ3JvdXBzW2ddID0gemVyb3MoKSk7CiAg"
    "Y29uc3QgbGlzdCA9IG5hbWUgPT09ICJfX0FMTF9fIiA/IGluY2x1ZGVkUHJvcHMoKSA6IFtuYW1lXTsKICBmb3IgKGNvbnN0IHBy"
    "b3Agb2YgbGlzdCkgewogICAgY29uc3QgbW9udGhzID0gUC5wcm9wZXJ0aWVzW3Byb3BdOwogICAgZm9yIChsZXQgaSA9IDA7IGkg"
    "PCBuOyBpKyspIHsKICAgICAgY29uc3QgbSA9IG1vbnRoc1tpXTsKICAgICAgcy5pbmNvbWVbaV0gKz0gbS50b3RhbEluY29tZTsg"
    "cy5leHBlbnNlc1tpXSArPSBtLnRvdGFsRXhwZW5zZXM7CiAgICAgIHMuZXhwZW5zZXNSZWNvcmRlZFtpXSArPSBtLnRvdGFsRXhw"
    "ZW5zZXNSZWNvcmRlZDsKICAgICAgcy5ub2lbaV0gKz0gbS5ub2k7IHMuZGVidFtpXSArPSBtLmRlYnQ7IHMuY2FwZXhbaV0gKz0g"
    "bS5jYXBleDsKICAgICAgcy5jYXNoZmxvd1tpXSArPSBtLmNhc2hmbG93OyBzLnZhcmlhbmNlW2ldICs9IG0udmFyaWFuY2U7CiAg"
    "ICAgIHMucmVudFtpXSArPSAobS5pbmNvbWVbIlJlbnRhbCBJbmNvbWUiXSB8fCAwKTsKICAgICAgZm9yIChjb25zdCBsYWIgaW4g"
    "bS5leHBlbnNlcykgewogICAgICAgIChzLmxpbmVzW2xhYl0gPSBzLmxpbmVzW2xhYl0gfHwgemVyb3MoKSlbaV0gKz0gbS5leHBl"
    "bnNlc1tsYWJdOwogICAgICAgIGNvbnN0IGcgPSBHUk9VUF9PRltsYWIudG9Mb3dlckNhc2UoKV07CiAgICAgICAgaWYgKGcpIHMu"
    "Z3JvdXBzW2ddW2ldICs9IG0uZXhwZW5zZXNbbGFiXTsKICAgICAgfQogICAgICBmb3IgKGNvbnN0IGxhYiBpbiBtLmluY29tZSkg"
    "KHMubGluZXNbIisiICsgbGFiXSA9IHMubGluZXNbIisiICsgbGFiXSB8fCB6ZXJvcygpKVtpXSArPSBtLmluY29tZVtsYWJdOwog"
    "ICAgfQogIH0KICAvLyBGdWxsLWxlbmd0aCBhcnJheXMgc3RheSBpbnRhY3QgZm9yIHRoZSBjaGFydHM7IGV2ZXJ5IHRvdGFsIHJl"
    "c3BlY3RzIE1TRUwuCiAgY29uc3QgW2EsIGJdID0gTVNFTCB8fCBbMCwgbiAtIDFdOwogIHMucmFuZ2UgPSAoKSA9PiBbYSwgYl07"
    "CiAgcy5uTW9udGhzID0gKCkgPT4gYiAtIGEgKyAxOwogIHMuc2VsTW9udGhzID0gKCkgPT4gcy5tb250aHMuc2xpY2UoYSwgYiAr"
    "IDEpOwogIHMuc2xpY2UgPSBrID0+IHNba10uc2xpY2UoYSwgYiArIDEpOwogIHMueXRkID0gayA9PiBNYXRoLnJvdW5kKHN1bShz"
    "LnNsaWNlKGspKSAqIDEwMCkgLyAxMDA7CiAgcmV0dXJuIHM7Cn0KCi8qID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gY2hh"
    "cnRzID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gKi8KZnVuY3Rpb24gY2hhcnRJbmNvbWVFeHBlbnNlKHN2Zywgcykgewog"
    "IHN2Zy5pbm5lckhUTUwgPSAiIjsKICBjb25zdCBXID0gNTIwLCBIID0gMzAwLCBtID0geyB0OiAxNCwgcjogMTQsIGI6IDM0LCBs"
    "OiA1NiB9OwogIGNvbnN0IHB3ID0gVyAtIG0ubCAtIG0uciwgcGggPSBIIC0gbS50IC0gbS5iOwogIGNvbnN0IHRpY2tzID0gbmlj"
    "ZVRpY2tzKE1hdGgubWF4KC4uLnMuaW5jb21lLCAuLi5zLmV4cGVuc2VzKSwgOCksIHRvcCA9IHRpY2tzW3RpY2tzLmxlbmd0aCAt"
    "IDFdOwogIGNvbnN0IHkgPSB2ID0+IG0udCArIHBoIC0gKHYgLyB0b3ApICogcGg7CiAgdGlja3MuZm9yRWFjaCh0ID0+IHsKICAg"
    "IGVsKCJsaW5lIiwgeyB4MTogbS5sLCB4MjogbS5sICsgcHcsIHkxOiB5KHQpLCB5MjogeSh0KSwgY2xhc3M6IHQgPT09IDAgPyAi"
    "YmFzZWxpbmUiIDogImdyaWRsaW5lIiB9LCBzdmcpOwogICAgZWwoInRleHQiLCB7IHg6IG0ubCAtIDksIHk6IHkodCkgKyA0LCBj"
    "bGFzczogInRpY2siLCAidGV4dC1hbmNob3IiOiAiZW5kIiB9LCBzdmcpLnRleHRDb250ZW50ID0gY29tcGFjdCh0KTsKICB9KTsK"
    "ICBjb25zdCBbcmEsIHJiXSA9IHMucmFuZ2UoKTsKICBjb25zdCBiYW5kID0gcHcgLyBzLm1vbnRocy5sZW5ndGgsIGdhcCA9IDIs"
    "IGJ3ID0gTWF0aC5taW4oMjQsIChiYW5kIC0gMTYgLSBnYXApIC8gMik7CiAgcy5tb250aHMuZm9yRWFjaCgobW8sIGkpID0+IHsK"
    "ICAgIGNvbnN0IGN4ID0gbS5sICsgYmFuZCAqIGkgKyBiYW5kIC8gMiwgeDAgPSBjeCAtIGJ3IC0gZ2FwIC8gMjsKICAgIGNvbnN0"
    "IG9uID0gaSA+PSByYSAmJiBpIDw9IHJiOwogICAgY29uc3QgZyA9IGVsKCJnIiwgeyBjbGFzczogb24gPyAiIiA6ICJkaW0iIH0s"
    "IHN2Zyk7CiAgICBbWyJpbmNvbWUiLCAiLS1zZXJpZXMtMSIsIHgwXSwgWyJleHBlbnNlcyIsICItLXNlcmllcy0yIiwgeDAgKyBi"
    "dyArIGdhcF1dLmZvckVhY2goKFtrLCBjdiwgeF0pID0+IHsKICAgICAgY29uc3QgdiA9IHNba11baV07CiAgICAgIGVsKCJwYXRo"
    "IiwgeyBkOiBjb2xQYXRoKHgsIHkodiksIGJ3LCAodiAvIHRvcCkgKiBwaCwgNCksIGZpbGw6IGNzc3YoY3YpIH0sIGcpOwogICAg"
    "fSk7CiAgICBlbCgidGV4dCIsIHsgeDogY3gsIHk6IEggLSBtLmIgKyAxOCwgY2xhc3M6ICJ4bGFiIiArIChvbiAmJiBtRmlsdGVy"
    "ZWQoKSA/ICIgb24iIDogIiIpIH0sIHN2ZykudGV4dENvbnRlbnQgPSBtbzsKICAgIC8vIENsaWNrIGEgY29sdW1uIHRvIGZvY3Vz"
    "IHRoYXQgbW9udGggKGFuZCBhZ2FpbiB0byBjbGVhcikgLSB0aGUgY2hhcnQgZG91YmxlcwogICAgLy8gYXMgYSBzaG9ydGN1dCBm"
    "b3IgdGhlIGNoaXAgcm93IGFib3ZlIGl0LgogICAgY29uc3QgaGl0ID0gZWwoInJlY3QiLCB7IHg6IG0ubCArIGJhbmQgKiBpLCB5"
    "OiBtLnQsIHdpZHRoOiBiYW5kLCBoZWlnaHQ6IHBoLAogICAgICBmaWxsOiAidHJhbnNwYXJlbnQiLCBzdHlsZTogImN1cnNvcjpw"
    "b2ludGVyIiB9LCBzdmcpOwogICAgYXR0YWNoVGlwKGhpdCwgKCkgPT4gYDxkaXYgY2xhc3M9InQiPiR7bW99PC9kaXY+YAogICAg"
    "ICArIHRpcFJvdyhjc3N2KCItLXNlcmllcy0xIiksICJJbmNvbWUiLCBtb25leTIocy5pbmNvbWVbaV0pKQogICAgICArIHRpcFJv"
    "dyhjc3N2KCItLXNlcmllcy0yIiksICJFeHBlbnNlcyIsIG1vbmV5MihzLmV4cGVuc2VzW2ldKSkKICAgICAgKyB0aXBSb3coInRy"
    "YW5zcGFyZW50IiwgIk5PSSIsIG1vbmV5MihzLm5vaVtpXSkpCiAgICAgICsgYDxkaXYgY2xhc3M9InIiIHN0eWxlPSJtYXJnaW4t"
    "dG9wOjVweCI+PHNwYW4gY2xhc3M9Im5tIj5DbGljayB0byBmb2N1cyAke21vfTwvc3Bhbj48L2Rpdj5gKTsKICAgIGhpdC5hZGRF"
    "dmVudExpc3RlbmVyKCJjbGljayIsICgpID0+IHNldE1vbnRocyhyYSA9PT0gaSAmJiByYiA9PT0gaSA/IG1BbGwoKSA6IFtpLCBp"
    "XSkpOwogIH0pOwp9CgpmdW5jdGlvbiBjaGFydE5vaUNhc2goc3ZnLCBzKSB7CiAgc3ZnLmlubmVySFRNTCA9ICIiOwogIGNvbnN0"
    "IFcgPSA1MjAsIEggPSAzMDAsIG0gPSB7IHQ6IDIwLCByOiA1OCwgYjogMzQsIGw6IDU4IH07CiAgY29uc3QgcHcgPSBXIC0gbS5s"
    "IC0gbS5yLCBwaCA9IEggLSBtLnQgLSBtLmI7CiAgY29uc3QgYWxsID0gcy5ub2kuY29uY2F0KHMuY2FzaGZsb3cpOwogIGNvbnN0"
    "IGxvID0gTWF0aC5taW4oMCwgLi4uYWxsKSwgaGkgPSBNYXRoLm1heCgwLCAuLi5hbGwpOwogIGNvbnN0IHNwYW4gPSBoaSAtIGxv"
    "IHx8IDE7CiAgY29uc3Qgc3RlcCA9IG5pY2VUaWNrcyhzcGFuLCA2KVsxXSB8fCAxOwogIGNvbnN0IHRsbyA9IE1hdGguZmxvb3Io"
    "bG8gLyBzdGVwKSAqIHN0ZXAsIHRoaSA9IE1hdGguY2VpbChoaSAvIHN0ZXApICogc3RlcDsKICBjb25zdCB0aWNrcyA9IFtdOyBm"
    "b3IgKGxldCB2ID0gdGxvOyB2IDw9IHRoaSArIHN0ZXAgKiAwLjAwMTsgdiArPSBzdGVwKSB0aWNrcy5wdXNoKHYpOwogIGNvbnN0"
    "IHkgPSB2ID0+IG0udCArIHBoIC0gKCh2IC0gdGxvKSAvICh0aGkgLSB0bG8pKSAqIHBoOwogIGNvbnN0IHggPSBpID0+IG0ubCAr"
    "IChwdyAvIChzLm1vbnRocy5sZW5ndGggLSAxKSkgKiBpOwogIHRpY2tzLmZvckVhY2godCA9PiB7CiAgICBlbCgibGluZSIsIHsg"
    "eDE6IG0ubCwgeDI6IG0ubCArIHB3LCB5MTogeSh0KSwgeTI6IHkodCksIGNsYXNzOiBNYXRoLmFicyh0KSA8IDFlLTkgPyAiYmFz"
    "ZWxpbmUiIDogImdyaWRsaW5lIiB9LCBzdmcpOwogICAgZWwoInRleHQiLCB7IHg6IG0ubCAtIDksIHk6IHkodCkgKyA0LCBjbGFz"
    "czogInRpY2siLCAidGV4dC1hbmNob3IiOiAiZW5kIiB9LCBzdmcpLnRleHRDb250ZW50ID0gY29tcGFjdCh0KTsKICB9KTsKICBz"
    "Lm1vbnRocy5mb3JFYWNoKChtbywgaSkgPT4gZWwoInRleHQiLCB7IHg6IHgoaSksIHk6IEggLSBtLmIgKyAxOCwgY2xhc3M6ICJ4"
    "bGFiIiwgInRleHQtYW5jaG9yIjogIm1pZGRsZSIgfSwgc3ZnKS50ZXh0Q29udGVudCA9IG1vKTsKICBjb25zdCBbcmEsIHJiXSA9"
    "IHMucmFuZ2UoKTsKICBjb25zdCBzZXIgPSBbWyJub2kiLCAiLS1zZXJpZXMtMSJdLCBbImNhc2hmbG93IiwgIi0tc2VyaWVzLTIi"
    "XV07CiAgY29uc3QgcGF0aCA9IChrLCBmcm9tLCB0bykgPT4gc1trXS5zbGljZShmcm9tLCB0byArIDEpCiAgICAubWFwKCh2LCBq"
    "KSA9PiAoaiA/ICJMIiA6ICJNIikgKyB4KGZyb20gKyBqKSArICIgIiArIHkodikpLmpvaW4oIiAiKTsKICBzZXIuZm9yRWFjaCgo"
    "W2ssIGN2XSkgPT4gewogICAgaWYgKG1GaWx0ZXJlZCgpKSB7CiAgICAgIC8vIHRoZSB3aG9sZSB5ZWFyIHN0YXlzIG9uIHNjcmVl"
    "biwgcmVjZXNzaXZlOyB0aGUgd2luZG93IGlzIGRyYXduIG92ZXIgaXQKICAgICAgZWwoInBhdGgiLCB7IGQ6IHBhdGgoaywgMCwg"
    "cy5tb250aHMubGVuZ3RoIC0gMSksIGZpbGw6ICJub25lIiwgc3Ryb2tlOiBjc3N2KGN2KSwKICAgICAgICAic3Ryb2tlLXdpZHRo"
    "IjogMiwgInN0cm9rZS1saW5lam9pbiI6ICJyb3VuZCIsICJzdHJva2UtbGluZWNhcCI6ICJyb3VuZCIsIGNsYXNzOiAiZGltIiB9"
    "LCBzdmcpOwogICAgfQogICAgZWwoInBhdGgiLCB7IGQ6IHBhdGgoaywgcmEsIHJiKSwgZmlsbDogIm5vbmUiLCBzdHJva2U6IGNz"
    "c3YoY3YpLCAic3Ryb2tlLXdpZHRoIjogMiwKICAgICAgInN0cm9rZS1saW5lam9pbiI6ICJyb3VuZCIsICJzdHJva2UtbGluZWNh"
    "cCI6ICJyb3VuZCIgfSwgc3ZnKTsKICB9KTsKICBzZXIuZm9yRWFjaCgoW2ssIGN2XSkgPT4gc1trXS5mb3JFYWNoKCh2LCBpKSA9"
    "PiB7CiAgICBjb25zdCBvbiA9IGkgPj0gcmEgJiYgaSA8PSByYjsKICAgIGVsKCJjaXJjbGUiLCB7IGN4OiB4KGkpLCBjeTogeSh2"
    "KSwgcjogb24gPyA0IDogMywgZmlsbDogY3NzdihjdiksCiAgICAgIHN0cm9rZTogY3NzdigiLS1zdXJmYWNlLTEiKSwgInN0cm9r"
    "ZS13aWR0aCI6IDIsIGNsYXNzOiBvbiA/ICIiIDogImRpbSIgfSwgc3ZnKTsKICB9KSk7CiAgZWwoInRleHQiLCB7IHg6IHgocmIp"
    "ICsgMTAsIHk6IHkocy5ub2lbcmJdKSArIDQsIGNsYXNzOiAiZGxhYiIgfSwgc3ZnKS50ZXh0Q29udGVudCA9IGNvbXBhY3Qocy5u"
    "b2lbcmJdKTsKICBlbCgidGV4dCIsIHsgeDogeChyYikgKyAxMCwgeTogeShzLmNhc2hmbG93W3JiXSkgKyA0LCBjbGFzczogImRs"
    "YWIiIH0sIHN2ZykudGV4dENvbnRlbnQgPSBjb21wYWN0KHMuY2FzaGZsb3dbcmJdKTsKICBzLm1vbnRocy5mb3JFYWNoKChtbywg"
    "aSkgPT4gewogICAgY29uc3QgYncgPSBwdyAvIChzLm1vbnRocy5sZW5ndGggLSAxKTsKICAgIGNvbnN0IGhpdCA9IGVsKCJyZWN0"
    "IiwgeyB4OiBNYXRoLm1heChtLmwgLSA0LCB4KGkpIC0gYncgLyAyKSwgeTogbS50IC0gMTAsIHdpZHRoOiBidywgaGVpZ2h0OiBw"
    "aCArIDIwLCBmaWxsOiAidHJhbnNwYXJlbnQiLCBzdHlsZTogImN1cnNvcjpjcm9zc2hhaXIiIH0sIHN2Zyk7CiAgICBsZXQgbGlu"
    "ZSA9IG51bGw7CiAgICBoaXQuYWRkRXZlbnRMaXN0ZW5lcigibW91c2VlbnRlciIsICgpID0+IHsgbGluZSA9IGVsKCJsaW5lIiwg"
    "eyB4MTogeChpKSwgeDI6IHgoaSksIHkxOiBtLnQgLSA2LCB5MjogbS50ICsgcGgsIHN0cm9rZTogY3NzdigiLS1heGlzIiksICJz"
    "dHJva2Utd2lkdGgiOiAxIH0sIHN2Zyk7IH0pOwogICAgaGl0LmFkZEV2ZW50TGlzdGVuZXIoIm1vdXNlbGVhdmUiLCAoKSA9PiB7"
    "IGlmIChsaW5lKSB7IGxpbmUucmVtb3ZlKCk7IGxpbmUgPSBudWxsOyB9IH0pOwogICAgYXR0YWNoVGlwKGhpdCwgKCkgPT4gYDxk"
    "aXYgY2xhc3M9InQiPiR7bW99PC9kaXY+YAogICAgICArIHRpcFJvdyhjc3N2KCItLXNlcmllcy0xIiksICJOT0kiLCBtb25leTIo"
    "cy5ub2lbaV0pKQogICAgICArIHRpcFJvdyhjc3N2KCItLXNlcmllcy0yIiksICJDYXNoIGZsb3ciLCBtb25leTIocy5jYXNoZmxv"
    "d1tpXSkpCiAgICAgICsgdGlwUm93KCJ0cmFuc3BhcmVudCIsICJEZWJ0IHNlcnZpY2UiLCBtb25leTIocy5kZWJ0W2ldKSkpOwog"
    "IH0pOwp9CgpmdW5jdGlvbiBjaGFydEV4cGVuc2VCYXJzKHN2ZywgcykgewogIHN2Zy5pbm5lckhUTUwgPSAiIjsKICBjb25zdCBj"
    "YXRzID0gT2JqZWN0LmtleXMocy5saW5lcykuZmlsdGVyKGsgPT4ga1swXSAhPT0gIisiKQogICAgLm1hcChrID0+IFtrLCBzdW0o"
    "cy5saW5lc1trXSldKS5maWx0ZXIoZCA9PiBkWzFdID4gMCkuc29ydCgoYSwgYikgPT4gYlsxXSAtIGFbMV0pOwogIGNvbnN0IFcg"
    "PSAxMDQwLCBtID0geyB0OiA4LCByOiAxMjAsIGI6IDgsIGw6IDIwMCB9LCByb3dIID0gMjg7CiAgY29uc3QgSCA9IG0udCArIGNh"
    "dHMubGVuZ3RoICogcm93SCArIG0uYjsKICBzdmcuc2V0QXR0cmlidXRlKCJ2aWV3Qm94IiwgYDAgMCAke1d9ICR7SH1gKTsKICBj"
    "b25zdCBwdyA9IFcgLSBtLmwgLSBtLnIsIG1heCA9IGNhdHNbMF0gPyBjYXRzWzBdWzFdIDogMSwgYyA9IGNzc3YoIi0tc2VyaWVz"
    "LTEiKTsKICBjb25zdCB0b3RhbCA9IHN1bShjYXRzLm1hcChkID0+IGRbMV0pKTsKICBlbCgibGluZSIsIHsgeDE6IG0ubCwgeDI6"
    "IG0ubCwgeTE6IG0udCwgeTI6IEggLSBtLmIsIGNsYXNzOiAiYmFzZWxpbmUiIH0sIHN2Zyk7CiAgY2F0cy5mb3JFYWNoKChbbmFt"
    "ZSwgdmFsXSwgaSkgPT4gewogICAgY29uc3QgeVRvcCA9IG0udCArIGkgKiByb3dILCBiaCA9IDE1LCBieSA9IHlUb3AgKyAocm93"
    "SCAtIGJoKSAvIDIsIHcgPSAodmFsIC8gbWF4KSAqIHB3OwogICAgZWwoInRleHQiLCB7IHg6IG0ubCAtIDEyLCB5OiBieSArIGJo"
    "IC8gMiArIDQsIGNsYXNzOiAieGxhYiIsICJ0ZXh0LWFuY2hvciI6ICJlbmQiIH0sIHN2ZykudGV4dENvbnRlbnQgPSBuYW1lOwog"
    "ICAgY29uc3QgciA9IE1hdGgubWluKDQsIHcgLyAyKTsKICAgIGNvbnN0IGQgPSB3IDw9IDAuNSA/IGBNJHttLmx9ICR7Ynl9IHYk"
    "e2JofWAgOgogICAgICBgTSR7bS5sfSAke2J5fSBIJHttLmwgKyB3IC0gcn0gYSR7cn0gJHtyfSAwIDAgMSAke3J9ICR7cn0gdiR7"
    "YmggLSAyICogcn0gYSR7cn0gJHtyfSAwIDAgMSAkey1yfSAke3J9IEgke20ubH0gWmA7CiAgICBlbCgicGF0aCIsIHsgZCwgZmls"
    "bDogYyB9LCBzdmcpOwogICAgZWwoInRleHQiLCB7IHg6IG0ubCArIHcgKyAxMCwgeTogYnkgKyBiaCAvIDIgKyA0LCBjbGFzczog"
    "ImRsYWIiIH0sIHN2ZykudGV4dENvbnRlbnQgPSBtb25leSh2YWwpOwogICAgY29uc3QgaGl0ID0gZWwoInJlY3QiLCB7IHg6IG0u"
    "bCwgeTogeVRvcCwgd2lkdGg6IHB3ICsgbS5yLCBoZWlnaHQ6IHJvd0gsIGZpbGw6ICJ0cmFuc3BhcmVudCIgfSwgc3ZnKTsKICAg"
    "IGF0dGFjaFRpcChoaXQsICgpID0+IGA8ZGl2IGNsYXNzPSJ0Ij4ke25hbWV9PC9kaXY+YAogICAgICArIHRpcFJvdyhjLCAiWVRE"
    "IHRvdGFsIiwgbW9uZXkyKHZhbCkpCiAgICAgICsgdGlwUm93KCJ0cmFuc3BhcmVudCIsICJTaGFyZSIsICh2YWwgLyB0b3RhbCAq"
    "IDEwMCkudG9GaXhlZCgxKSArICIlIikKICAgICAgKyB0aXBSb3coInRyYW5zcGFyZW50IiwgIlBlciBtb250aCIsIG1vbmV5Mih2"
    "YWwgLyBzLm1vbnRocy5sZW5ndGgpKSk7CiAgfSk7Cn0KCmZ1bmN0aW9uIGNoYXJ0RXhwZW5zZU1peChzdmcsIHMpIHsKICBzdmcu"
    "aW5uZXJIVE1MID0gIiI7CiAgY29uc3QgVyA9IDEwNDAsIEggPSAzMjAsIG0gPSB7IHQ6IDE2LCByOiAxNiwgYjogMzYsIGw6IDcy"
    "IH07CiAgY29uc3QgcHcgPSBXIC0gbS5sIC0gbS5yLCBwaCA9IEggLSBtLnQgLSBtLmI7CiAgY29uc3QgdG90YWxzID0gcy5tb250"
    "aHMubWFwKChfLCBpKSA9PiBzdW0oR1JPVVBfTkFNRVMubWFwKGcgPT4gcy5ncm91cHNbZ11baV0pKSk7CiAgY29uc3QgdGlja3Mg"
    "PSBuaWNlVGlja3MoTWF0aC5tYXgoLi4udG90YWxzKSwgOCksIHRvcCA9IHRpY2tzW3RpY2tzLmxlbmd0aCAtIDFdOwogIGNvbnN0"
    "IHkgPSB2ID0+IG0udCArIHBoIC0gKHYgLyB0b3ApICogcGg7CiAgdGlja3MuZm9yRWFjaCh0ID0+IHsKICAgIGVsKCJsaW5lIiwg"
    "eyB4MTogbS5sLCB4MjogbS5sICsgcHcsIHkxOiB5KHQpLCB5MjogeSh0KSwgY2xhc3M6IHQgPT09IDAgPyAiYmFzZWxpbmUiIDog"
    "ImdyaWRsaW5lIiB9LCBzdmcpOwogICAgZWwoInRleHQiLCB7IHg6IG0ubCAtIDEwLCB5OiB5KHQpICsgNCwgY2xhc3M6ICJ0aWNr"
    "IiwgInRleHQtYW5jaG9yIjogImVuZCIgfSwgc3ZnKS50ZXh0Q29udGVudCA9IGNvbXBhY3QodCk7CiAgfSk7CiAgY29uc3QgW3Jh"
    "LCByYl0gPSBzLnJhbmdlKCk7CiAgY29uc3QgYmFuZCA9IHB3IC8gcy5tb250aHMubGVuZ3RoLCBidyA9IE1hdGgubWluKDI0LCBi"
    "YW5kICogMC4zNCk7CiAgcy5tb250aHMuZm9yRWFjaCgobW8sIGkpID0+IHsKICAgIGNvbnN0IGN4ID0gbS5sICsgYmFuZCAqIGkg"
    "KyBiYW5kIC8gMiwgeCA9IGN4IC0gYncgLyAyOwogICAgY29uc3Qgb24gPSBpID49IHJhICYmIGkgPD0gcmI7CiAgICBjb25zdCBj"
    "b2wgPSBlbCgiZyIsIHsgY2xhc3M6IG9uID8gIiIgOiAiZGltIiB9LCBzdmcpOwogICAgbGV0IGFjYyA9IDA7CiAgICBHUk9VUF9O"
    "QU1FUy5mb3JFYWNoKChnLCBnaSkgPT4gewogICAgICBjb25zdCB2ID0gcy5ncm91cHNbZ11baV07CiAgICAgIGlmICh2IDw9IDAp"
    "IHJldHVybjsKICAgICAgY29uc3QgeVRvcCA9IHkoYWNjICsgdiksIHlCb3QgPSB5KGFjYyk7CiAgICAgIGNvbnN0IGggPSBNYXRo"
    "Lm1heCgwLCB5Qm90IC0geVRvcCAtIChhY2MgPiAwID8gMiA6IDApKTsKICAgICAgY29uc3QgZCA9IGdpID09PSBHUk9VUF9OQU1F"
    "Uy5sZW5ndGggLSAxID8gY29sUGF0aCh4LCB5VG9wLCBidywgaCwgNCkgOiBgTSR7eH0gJHt5VG9wfSBoJHtid30gdiR7aH0gaCR7"
    "LWJ3fSBaYDsKICAgICAgZWwoInBhdGgiLCB7IGQsIGZpbGw6IGNzc3YoR1JPVVBfVkFSW2dpXSkgfSwgY29sKTsKICAgICAgYWNj"
    "ICs9IHY7CiAgICB9KTsKICAgIGVsKCJ0ZXh0IiwgeyB4OiBjeCwgeTogSCAtIG0uYiArIDE4LCBjbGFzczogInhsYWIiICsgKG9u"
    "ICYmIG1GaWx0ZXJlZCgpID8gIiBvbiIgOiAiIiksICJ0ZXh0LWFuY2hvciI6ICJtaWRkbGUiIH0sIHN2ZykudGV4dENvbnRlbnQg"
    "PSBtbzsKICAgIGNvbnN0IGhpdCA9IGVsKCJyZWN0IiwgeyB4OiBtLmwgKyBiYW5kICogaSwgeTogbS50LCB3aWR0aDogYmFuZCwg"
    "aGVpZ2h0OiBwaCwgZmlsbDogInRyYW5zcGFyZW50IiB9LCBzdmcpOwogICAgYXR0YWNoVGlwKGhpdCwgKCkgPT4gYDxkaXYgY2xh"
    "c3M9InQiPiR7bW99PC9kaXY+YAogICAgICArIEdST1VQX05BTUVTLm1hcCgoZywgZ2kpID0+IHRpcFJvdyhjc3N2KEdST1VQX1ZB"
    "UltnaV0pLCBnLCBtb25leTIocy5ncm91cHNbZ11baV0pKSkuam9pbigiIikKICAgICAgKyB0aXBSb3coInRyYW5zcGFyZW50Iiwg"
    "IlRvdGFsIiwgbW9uZXkyKHRvdGFsc1tpXSkpKTsKICB9KTsKfQoKLyogRGl2ZXJnaW5nIGhvcml6b250YWwgYmFyczogWVREIGNh"
    "c2ggZmxvdyBhZnRlciBkZWJ0LCBieSBwcm9wZXJ0eS4gKi8KZnVuY3Rpb24gY2hhcnRSYW5raW5nKHN2Zywgcm93cywgb25QaWNr"
    "KSB7CiAgc3ZnLmlubmVySFRNTCA9ICIiOwogIGNvbnN0IFcgPSAxMDQwLCBtID0geyB0OiA4LCByOiAyNCwgYjogOCwgbDogMTc2"
    "IH0sIHJvd0ggPSAzMDsKICBjb25zdCBIID0gbS50ICsgcm93cy5sZW5ndGggKiByb3dIICsgbS5iOwogIHN2Zy5zZXRBdHRyaWJ1"
    "dGUoInZpZXdCb3giLCBgMCAwICR7V30gJHtIfWApOwogIC8vIFZhbHVlIGxhYmVscyBzaXQgb3V0c2lkZSB0aGUgYmFyIGVuZHMs"
    "IHNvIHJlc2VydmUgYSBndXR0ZXIgb24gZWFjaCBzaWRlIHdpZGUKICAvLyBlbm91Z2ggZm9yIHRoZSBsb25nZXN0IG9uZS4gV2l0"
    "aG91dCBpdCB0aGUgbGFyZ2VzdCBuZWdhdGl2ZSBiYXIgcnVucyBpdHMKICAvLyBsYWJlbCBzdHJhaWdodCBpbnRvIHRoZSBwcm9w"
    "ZXJ0eS1uYW1lIGNvbHVtbi4KICBjb25zdCBtYXhBYnMgPSBNYXRoLm1heCguLi5yb3dzLm1hcChyID0+IE1hdGguYWJzKHIudmFs"
    "dWUpKSkgfHwgMTsKICBjb25zdCBndXR0ZXIgPSBNYXRoLm1heCg1NiwgbW9uZXkoLW1heEFicykubGVuZ3RoICogNy4yKTsKICBj"
    "b25zdCBiYXJMZWZ0ID0gbS5sICsgZ3V0dGVyLCBiYXJSaWdodCA9IFcgLSBtLnIgLSBndXR0ZXI7CiAgY29uc3QgaGFsZiA9IChi"
    "YXJSaWdodCAtIGJhckxlZnQpIC8gMiwgemVybyA9IGJhckxlZnQgKyBoYWxmOwogIGNvbnN0IHBvc0MgPSBjc3N2KCItLXBvcyIp"
    "LCBuZWdDID0gY3NzdigiLS1uZWdiIik7CiAgcm93cy5mb3JFYWNoKChyLCBpKSA9PiB7CiAgICBjb25zdCB5VG9wID0gbS50ICsg"
    "aSAqIHJvd0gsIGJoID0gMTUsIGJ5ID0geVRvcCArIChyb3dIIC0gYmgpIC8gMjsKICAgIGNvbnN0IHcgPSBNYXRoLmFicyhyLnZh"
    "bHVlKSAvIG1heEFicyAqIGhhbGY7CiAgICBjb25zdCBwb3NpdGl2ZSA9IHIudmFsdWUgPj0gMDsKICAgIGNvbnN0IHggPSBwb3Np"
    "dGl2ZSA/IHplcm8gOiB6ZXJvIC0gdzsKICAgIGNvbnN0IHJhZCA9IE1hdGgubWluKDQsIHcgLyAyKTsKICAgIGNvbnN0IGhpdCA9"
    "IGVsKCJyZWN0IiwgeyB4OiAwLCB5OiB5VG9wLCB3aWR0aDogVywgaGVpZ2h0OiByb3dILCBjbGFzczogInJvd2hpdCIgfSwgc3Zn"
    "KTsKICAgIGVsKCJ0ZXh0IiwgeyB4OiBtLmwgLSAxNCwgeTogYnkgKyBiaCAvIDIgKyA0LCBjbGFzczogInhsYWIiLCAidGV4dC1h"
    "bmNob3IiOiAiZW5kIiB9LCBzdmcpLnRleHRDb250ZW50ID0gci5uYW1lOwogICAgaWYgKHcgPiAwLjUpIHsKICAgICAgY29uc3Qg"
    "ZCA9IHBvc2l0aXZlCiAgICAgICAgPyBgTSR7eH0gJHtieX0gSCR7eCArIHcgLSByYWR9IGEke3JhZH0gJHtyYWR9IDAgMCAxICR7"
    "cmFkfSAke3JhZH0gdiR7YmggLSAyICogcmFkfSBhJHtyYWR9ICR7cmFkfSAwIDAgMSAkey1yYWR9ICR7cmFkfSBIJHt4fSBaYAog"
    "ICAgICAgIDogYE0ke3ggKyB3fSAke2J5fSBIJHt4ICsgcmFkfSBhJHtyYWR9ICR7cmFkfSAwIDAgMCAkey1yYWR9ICR7cmFkfSB2"
    "JHtiaCAtIDIgKiByYWR9IGEke3JhZH0gJHtyYWR9IDAgMCAwICR7cmFkfSAke3JhZH0gSCR7eCArIHd9IFpgOwogICAgICBlbCgi"
    "cGF0aCIsIHsgZCwgZmlsbDogcG9zaXRpdmUgPyBwb3NDIDogbmVnQyB9LCBzdmcpOwogICAgfQogICAgZWwoInRleHQiLCB7CiAg"
    "ICAgIHg6IHBvc2l0aXZlID8geCArIHcgKyA5IDogeCAtIDksIHk6IGJ5ICsgYmggLyAyICsgNCwgY2xhc3M6ICJkbGFiIiwKICAg"
    "ICAgInRleHQtYW5jaG9yIjogcG9zaXRpdmUgPyAic3RhcnQiIDogImVuZCIKICAgIH0sIHN2ZykudGV4dENvbnRlbnQgPSBtb25l"
    "eShyLnZhbHVlKTsKICAgIGF0dGFjaFRpcChoaXQsICgpID0+IGA8ZGl2IGNsYXNzPSJ0Ij4ke3IubmFtZX08L2Rpdj5gCiAgICAg"
    "ICsgdGlwUm93KHBvc2l0aXZlID8gcG9zQyA6IG5lZ0MsIG1GaWx0ZXJlZCgpID8gIkNhc2ggZmxvdyIgOiAiQ2FzaCBmbG93IFlU"
    "RCIsIG1vbmV5MihyLnZhbHVlKSkKICAgICAgKyB0aXBSb3coInRyYW5zcGFyZW50IiwgIk5PSSIsIG1vbmV5MihyLm5vaSkpCiAg"
    "ICAgICsgdGlwUm93KCJ0cmFuc3BhcmVudCIsICJEZWJ0IHNlcnZpY2UiLCBtb25leTIoci5kZWJ0KSkKICAgICAgKyB0aXBSb3co"
    "InRyYW5zcGFyZW50IiwgIk5PSSBtYXJnaW4iLCByLm1hcmdpbikKICAgICAgKyBgPGRpdiBjbGFzcz0iciIgc3R5bGU9Im1hcmdp"
    "bi10b3A6NXB4Ij48c3BhbiBjbGFzcz0ibm0iPkNsaWNrIHRvIG9wZW48L3NwYW4+PC9kaXY+YCk7CiAgICBoaXQuYWRkRXZlbnRM"
    "aXN0ZW5lcigiY2xpY2siLCAoKSA9PiBvblBpY2soci5uYW1lKSk7CiAgICBoaXQuYWRkRXZlbnRMaXN0ZW5lcigia2V5ZG93biIs"
    "IGUgPT4geyBpZiAoZS5rZXkgPT09ICJFbnRlciIpIG9uUGljayhyLm5hbWUpOyB9KTsKICB9KTsKICBlbCgibGluZSIsIHsgeDE6"
    "IHplcm8sIHgyOiB6ZXJvLCB5MTogbS50LCB5MjogSCAtIG0uYiwgY2xhc3M6ICJiYXNlbGluZSIgfSwgc3ZnKTsKfQoKLyogPT09"
    "PT09PT09PT09PT09PT09PT09PT09PT09PSB0YWJsZSA9PT09PT09PT09PT09PT09PT09PT09PT09PT09ICovCmZ1bmN0aW9uIHRh"
    "YmxlUm93cyhzKSB7CiAgY29uc3QgaW5jTGFiZWxzID0gT2JqZWN0LmtleXMocy5saW5lcykuZmlsdGVyKGsgPT4ga1swXSA9PT0g"
    "IisiKTsKICBjb25zdCBbcmEsIHJiXSA9IHMucmFuZ2UoKTsKICBjb25zdCB3aW4gPSBhcnIgPT4gc3VtKGFyci5zbGljZShyYSwg"
    "cmIgKyAxKSk7CiAgY29uc3QgZXhwTGFiZWxzID0gT2JqZWN0LmtleXMocy5saW5lcykuZmlsdGVyKGsgPT4ga1swXSAhPT0gIisi"
    "KQogICAgLmZpbHRlcihrID0+IHdpbihzLmxpbmVzW2tdKSAhPT0gMCkuc29ydCgoYSwgYikgPT4gd2luKHMubGluZXNbYl0pIC0g"
    "d2luKHMubGluZXNbYV0pKTsKICBjb25zdCByb3dzID0gW1sic2VjdGlvbiIsICJJbmNvbWUiXV07CiAgaW5jTGFiZWxzLmZvckVh"
    "Y2goayA9PiByb3dzLnB1c2goWyJpdGVtIiwgay5zbGljZSgxKSwgcy5saW5lc1trXV0pKTsKICByb3dzLnB1c2goWyJ0b3RhbCIs"
    "ICJUb3RhbCBpbmNvbWUiLCBzLmluY29tZV0sIFsic2VjdGlvbiIsICJPcGVyYXRpbmcgZXhwZW5zZXMiXSk7CiAgZXhwTGFiZWxz"
    "LmZvckVhY2goayA9PiByb3dzLnB1c2goWyJpdGVtIiwgaywgcy5saW5lc1trXV0pKTsKICByb3dzLnB1c2goWyJ0b3RhbCIsICJU"
    "b3RhbCBvcGVyYXRpbmcgZXhwZW5zZXMgKGFzIHN1YnRvdGFsZWQpIiwgcy5leHBlbnNlc10sCiAgICBbInNlY3Rpb24iLCAiIl0s"
    "IFsidG90YWwiLCAiTmV0IG9wZXJhdGluZyBpbmNvbWUiLCBzLm5vaV0sCiAgICBbIml0ZW0iLCAiRGVidCBzZXJ2aWNlIiwgcy5k"
    "ZWJ0XSwgWyJ0b3RhbCIsICJDYXNoIGZsb3cgYWZ0ZXIgZGVidCIsIHMuY2FzaGZsb3ddKTsKICBpZiAod2luKHMuY2FwZXgpICE9"
    "PSAwKSByb3dzLnB1c2goWyJpdGVtIiwgIkNhcGl0YWwgaW1wcm92ZW1lbnRzIiwgcy5jYXBleF0pOwogIHJldHVybiByb3dzOwp9"
    "CgpmdW5jdGlvbiB0YWJsZUZvcihzKSB7CiAgY29uc3Qgcm93cyA9IHRhYmxlUm93cyhzKTsKICBjb25zdCBbYSwgYl0gPSBzLnJh"
    "bmdlKCk7CiAgY29uc3QgdG90YWxDb2wgPSBtRmlsdGVyZWQoKSA/IChzLm5Nb250aHMoKSA9PT0gMSA/ICJUb3RhbCIgOiAiUGVy"
    "aW9kIikgOiAiWVREIjsKICBsZXQgaCA9ICI8dGhlYWQ+PHRyPjx0aD5MaW5lPC90aD4iICsgcy5zZWxNb250aHMoKS5tYXAobSA9"
    "PiBgPHRoPiR7bX08L3RoPmApLmpvaW4oIiIpICsKICAgIGA8dGg+JHt0b3RhbENvbH08L3RoPjwvdHI+PC90aGVhZD48dGJvZHk+"
    "YDsKICByb3dzLmZvckVhY2gociA9PiB7CiAgICBpZiAoclswXSA9PT0gInNlY3Rpb24iKSB7IGggKz0gYDx0ciBjbGFzcz0ic2Vj"
    "dGlvbiI+PHRkIGNvbHNwYW49IiR7cy5uTW9udGhzKCkgKyAyfSI+JHtlc2MoclsxXSl9PC90ZD48L3RyPmA7IHJldHVybjsgfQog"
    "ICAgY29uc3QgdmFscyA9IHJbMl0uc2xpY2UoYSwgYiArIDEpOwogICAgaCArPSBgPHRyIGNsYXNzPSIke3JbMF0gPT09ICJ0b3Rh"
    "bCIgPyAidG90YWwiIDogIiJ9Ij48dGQgY2xhc3M9IiR7clswXSA9PT0gIml0ZW0iID8gImluZGVudCIgOiAiIn0iPiR7ZXNjKHJb"
    "MV0pfTwvdGQ+YAogICAgICArIHZhbHMubWFwKHYgPT4gYDx0ZD4ke21vbmV5Mih2KX08L3RkPmApLmpvaW4oIiIpICsgYDx0ZD4k"
    "e21vbmV5MihzdW0odmFscykpfTwvdGQ+PC90cj5gOwogIH0pOwogIHJldHVybiBgPGRpdiBjbGFzcz0idGFibGV0b29scyI+CiAg"
    "ICAgIDxidXR0b24gdHlwZT0iYnV0dG9uIiBjbGFzcz0iZGxidG4iIGRhdGEtZG93bmxvYWQ9ImNzdiI+JiM4NTk1OyBEb3dubG9h"
    "ZCBDU1Y8L2J1dHRvbj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0ic2Nyb2xsZXIiPjx0YWJsZT4ke2h9PC90Ym9keT48L3Rh"
    "YmxlPjwvZGl2PmA7Cn0KCi8qIC0tLS0tLS0tLS0gQ1NWIGV4cG9ydCAtLS0tLS0tLS0tCiAgIFZhbHVlcyBnbyBvdXQgYXMgcmF3"
    "IG51bWJlcnMsIG5vdCB0aGUgZm9ybWF0dGVkIHN0cmluZ3MgaW4gdGhlIHRhYmxlLCBzbyB0aGUKICAgZmlsZSBsYW5kcyBpbiBh"
    "IHNwcmVhZHNoZWV0IGFzIG51bWJlcnMgcmF0aGVyIHRoYW4gdGV4dC4gVGhlIGxlYWRpbmcgYmxvY2sKICAgcmVjb3JkcyB3aGlj"
    "aCBzbGljZSBvZiB0aGUgcG9ydGZvbGlvIHRoaXMgaXMgLSBhIENTViBvZiBhIGZpbHRlcmVkIHZpZXcgdGhhdAogICBkb2VzIG5v"
    "dCBzYXkgc28gaXMgZXhhY3RseSB0aGUgdGhpbmcgdGhlIG9uLXNjcmVlbiBjaHJvbWUgZ3VhcmRzIGFnYWluc3QuICovCmZ1bmN0"
    "aW9uIGNzdkVzY2FwZSh2KSB7CiAgY29uc3QgdCA9IFN0cmluZyh2KTsKICByZXR1cm4gL1siLFxuXS8udGVzdCh0KSA/ICciJyAr"
    "IHQucmVwbGFjZSgvIi9nLCAnIiInKSArICciJyA6IHQ7Cn0KZnVuY3Rpb24gYnVpbGRDU1YocywgdGl0bGUpIHsKICBjb25zdCBM"
    "ID0gW107CiAgY29uc3QgaW5jID0gaW5jbHVkZWRQcm9wcygpLCBleGNsdWRlZCA9IGFsbFByb3BzKCkuZmlsdGVyKHAgPT4gIWlu"
    "Yy5pbmNsdWRlcyhwKSk7CiAgTC5wdXNoKFsiVEFHTFlaIHBvcnRmb2xpbyByZXBvcnQiXSk7CiAgTC5wdXNoKFsiVmlldyIsIHRp"
    "dGxlXSk7CiAgaWYgKFZJRVcudHlwZSA9PT0gInBvcnRmb2xpbyIpIHsKICAgIEwucHVzaChbIlByb3BlcnRpZXMgaW5jbHVkZWQi"
    "LCBpc0ZpbHRlcmVkKCkgPyBgJHtpbmMubGVuZ3RofSBvZiAke2FsbFByb3BzKCkubGVuZ3RofWAgOiBgYWxsICR7YWxsUHJvcHMo"
    "KS5sZW5ndGh9YF0pOwogICAgTC5wdXNoKFsiSW5jbHVkZWQiLCBpbmMuam9pbigiOyAiKV0pOwogICAgaWYgKGV4Y2x1ZGVkLmxl"
    "bmd0aCkgTC5wdXNoKFsiRXhjbHVkZWQiLCBleGNsdWRlZC5qb2luKCI7ICIpXSk7CiAgfQogIEwucHVzaChbIlBlcmlvZCIsIHMu"
    "bk1vbnRocygpID09PSAxID8gYCR7cy5zZWxNb250aHMoKVswXX0gJHtQLnllYXJ9YAogICAgOiBgJHtzLnNlbE1vbnRocygpWzBd"
    "fS0ke3Muc2VsTW9udGhzKClbcy5uTW9udGhzKCkgLSAxXX0gJHtQLnllYXJ9YF0pOwogIGlmIChtRmlsdGVyZWQoKSkgTC5wdXNo"
    "KFsiU2NvcGUiLCBgQSBtb250aCBzZWxlY3Rpb24gaXMgYWN0aXZlLiBUaGlzIGV4cG9ydCBjb3ZlcnMgJHtzLm5Nb250aHMoKX0g"
    "b2YgJHtzLm1vbnRocy5sZW5ndGh9IG1vbnRocyBpbiB0aGUgc291cmNlIHdvcmtib29rLmBdKTsKICBMLnB1c2goWyJTb3VyY2Ui"
    "LCBNRVRBLnNvdXJjZV0pOwogIEwucHVzaChbIkJhc2lzIiwgIkFjY3J1YWwuIEZpZ3VyZXMgYXMgcmVwb3J0ZWQgaW4gdGhlIHNv"
    "dXJjZSB3b3JrYm9vay4iXSk7CiAgY29uc3QgdmFyaWFuY2UgPSBzLnl0ZCgidmFyaWFuY2UiKTsKICBpZiAoTWF0aC5hYnModmFy"
    "aWFuY2UpID49IDEpIHsKICAgIEwucHVzaChbIk5vdGUiLCBgVGhlIHNvdXJjZSB3b3JrYm9vaydzIGV4cGVuc2Ugc3VidG90YWxz"
    "IG9taXQgYW4gaW5zdXJhbmNlIGFtb3VudCByZWNvcmRlZCBpbiB0aGUgc2FtZSBjb2x1bW47IGAgKwogICAgICBgJHttb25leTIo"
    "dmFyaWFuY2UpfSBpbiB0b3RhbC4gQ291bnRpbmcgaXQsIG9wZXJhdGluZyBleHBlbnNlcyBhcmUgJHttb25leTIocy55dGQoImV4"
    "cGVuc2VzUmVjb3JkZWQiKSl9LmBdKTsKICB9CiAgTC5wdXNoKFtdKTsKICBjb25zdCBbY2EsIGNiXSA9IHMucmFuZ2UoKTsKICBM"
    "LnB1c2goWyJMaW5lIiwgLi4ucy5zZWxNb250aHMoKSwgbUZpbHRlcmVkKCkgPyAiVG90YWwiIDogIllURCJdKTsKICB0YWJsZVJv"
    "d3MocykuZm9yRWFjaChyID0+IHsKICAgIGlmIChyWzBdID09PSAic2VjdGlvbiIpIHsgaWYgKHJbMV0pIEwucHVzaChbXSksIEwu"
    "cHVzaChbclsxXV0pOyByZXR1cm47IH0KICAgIGNvbnN0IHZhbHMgPSByWzJdLnNsaWNlKGNhLCBjYiArIDEpOwogICAgTC5wdXNo"
    "KFtyWzFdLCAuLi52YWxzLm1hcCh2ID0+IHYudG9GaXhlZCgyKSksIHN1bSh2YWxzKS50b0ZpeGVkKDIpXSk7CiAgfSk7CiAgcmV0"
    "dXJuIEwubWFwKHJvdyA9PiByb3cubWFwKGNzdkVzY2FwZSkuam9pbigiLCIpKS5qb2luKCJcclxuIik7Cn0KCmZ1bmN0aW9uIGRv"
    "d25sb2FkQ1NWKGJ0bikgewogIGNvbnN0IHMgPSBWSUVXLnR5cGUgPT09ICJwb3J0Zm9saW8iID8gc2VyaWVzRm9yKCJfX0FMTF9f"
    "IikgOiBzZXJpZXNGb3IoVklFVy5sYWJlbCk7CiAgY29uc3QgdGl0bGUgPSAoVklFVy50eXBlID09PSAicG9ydGZvbGlvIgogICAg"
    "PyBWSUVXLmxhYmVsICsgKGlzRmlsdGVyZWQoKSA/IGAgKCR7aW5jbHVkZWRQcm9wcygpLmxlbmd0aH0gb2YgJHthbGxQcm9wcygp"
    "Lmxlbmd0aH0gcHJvcGVydGllcylgIDogIiIpCiAgICA6IFZJRVcubGFiZWwpICsgKG1GaWx0ZXJlZCgpID8gYCDigJQgJHttTGFi"
    "ZWwoKX1gIDogIiIpOwogIGNvbnN0IHNsdWcgPSAoVklFVy5sYWJlbCB8fCAicG9ydGZvbGlvIikucmVwbGFjZSgvW15BLVphLXow"
    "LTldKy9nLCAiLSIpLnJlcGxhY2UoL14tfC0kL2csICIiKTsKICBjb25zdCBzbSA9IHMuc2VsTW9udGhzKCk7CiAgY29uc3Qgc3Bh"
    "biA9IHNtLmxlbmd0aCA9PT0gMSA/IHNtWzBdIDogYCR7c21bMF19LSR7c21bc20ubGVuZ3RoIC0gMV19YDsKICBjb25zdCBuYW1l"
    "ID0gYFRBR0xZWi0ke3NsdWd9LSR7c3Bhbn0tJHtQLnllYXJ9LmNzdmA7CiAgLy8gXHVGRUZGOiB3aXRob3V0IHRoZSBCT00gRXhj"
    "ZWwgcmVhZHMgdGhlIGZpbGUgYXMgdGhlIGxvY2FsIGNvZGVwYWdlIGFuZCBtYW5nbGVzCiAgLy8gdGhlIG5vbi1BU0NJSSBjaGFy"
    "YWN0ZXJzIGluIHRoZSBub3RlIGxpbmVzLgogIGNvbnN0IGJsb2IgPSBuZXcgQmxvYihbIlx1RkVGRiIgKyBidWlsZENTVihzLCB0"
    "aXRsZSldLCB7IHR5cGU6ICJ0ZXh0L2NzdjtjaGFyc2V0PXV0Zi04IiB9KTsKICBjb25zdCB1cmwgPSBVUkwuY3JlYXRlT2JqZWN0"
    "VVJMKGJsb2IpOwogIGNvbnN0IGEgPSBkb2N1bWVudC5jcmVhdGVFbGVtZW50KCJhIik7CiAgYS5ocmVmID0gdXJsOyBhLmRvd25s"
    "b2FkID0gbmFtZTsKICBkb2N1bWVudC5ib2R5LmFwcGVuZENoaWxkKGEpOyBhLmNsaWNrKCk7IGEucmVtb3ZlKCk7CiAgc2V0VGlt"
    "ZW91dCgoKSA9PiBVUkwucmV2b2tlT2JqZWN0VVJMKHVybCksIDIwMDApOwogIGlmIChidG4pIHsKICAgIGNvbnN0IHdhcyA9IGJ0"
    "bi5pbm5lckhUTUw7CiAgICBidG4uaW5uZXJIVE1MID0gIiYjMTAwMDM7ICIgKyBuYW1lOwogICAgYnRuLmNsYXNzTGlzdC5hZGQo"
    "ImRvbmUiKTsKICAgIHNldFRpbWVvdXQoKCkgPT4geyBidG4uaW5uZXJIVE1MID0gd2FzOyBidG4uY2xhc3NMaXN0LnJlbW92ZSgi"
    "ZG9uZSIpOyB9LCAyNjAwKTsKICB9Cn0KCi8qID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gdmlld3MgPT09PT09PT09PT09"
    "PT09PT09PT09PT09PT09PSAqLwpmdW5jdGlvbiBrcGlUaWxlcyhzKSB7CiAgY29uc3QgaW5jID0gcy55dGQoImluY29tZSIpLCBl"
    "eHAgPSBzLnl0ZCgiZXhwZW5zZXMiKSwgbm9pID0gcy55dGQoIm5vaSIpOwogIGNvbnN0IGRlYnQgPSBzLnl0ZCgiZGVidCIpLCBj"
    "YXNoID0gcy55dGQoImNhc2hmbG93Iik7CiAgY29uc3QgZHNjciA9IGRlYnQgPyAobm9pIC8gZGVidCkgOiAwOwogIGNvbnN0IG5l"
    "Z01vbnRocyA9IHMuc2xpY2UoImNhc2hmbG93IikuZmlsdGVyKHYgPT4gdiA8IDApLmxlbmd0aDsKICBjb25zdCBuID0gcy5uTW9u"
    "dGhzKCk7CiAgcmV0dXJuIGA8c2VjdGlvbiBjbGFzcz0idGlsZXMiPgogICAgPGRpdiBjbGFzcz0idGlsZSI+PGRpdiBjbGFzcz0i"
    "bCI+VG90YWwgaW5jb21lPC9kaXY+PGRpdiBjbGFzcz0idiI+JHttb25leShpbmMpfTwvZGl2PjxkaXYgY2xhc3M9ImQiPiR7bW9u"
    "ZXkocy55dGQoInJlbnQiKSl9IG9mIGl0IHJlbnQ8L2Rpdj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InRpbGUiPjxkaXYgY2xhc3M9"
    "ImwiPk9wZXJhdGluZyBleHBlbnNlczwvZGl2PjxkaXYgY2xhc3M9InYiPiR7bW9uZXkoZXhwKX08L2Rpdj48ZGl2IGNsYXNzPSJk"
    "Ij4ke2luYyA/IChleHAgLyBpbmMgKiAxMDApLnRvRml4ZWQoMSkgOiAwfSUgb2YgaW5jb21lPC9kaXY+PC9kaXY+CiAgICA8ZGl2"
    "IGNsYXNzPSJ0aWxlIj48ZGl2IGNsYXNzPSJsIj5EZWJ0IHNlcnZpY2U8L2Rpdj48ZGl2IGNsYXNzPSJ2Ij4ke21vbmV5KGRlYnQp"
    "fTwvZGl2PjxkaXYgY2xhc3M9ImQiPiR7biA9PT0gMSA/ICJtb3J0Z2FnZSBwYXltZW50cyIgOiBtb25leShkZWJ0IC8gbikgKyAi"
    "IGEgbW9udGggYXZnIn08L2Rpdj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InRpbGUiPjxkaXYgY2xhc3M9ImwiPk5ldCBvcGVyYXRp"
    "bmcgaW5jb21lPC9kaXY+PGRpdiBjbGFzcz0idiAke25vaSA+PSAwID8gIiIgOiAibmVnIn0iPiR7bW9uZXkobm9pKX08L2Rpdj48"
    "ZGl2IGNsYXNzPSJkIj4ke2luYyA/IChub2kgLyBpbmMgKiAxMDApLnRvRml4ZWQoMSkgOiAwfSUgbWFyZ2luIG9uIGluY29tZTwv"
    "ZGl2PjwvZGl2PgogICAgPGRpdiBjbGFzcz0idGlsZSI+PGRpdiBjbGFzcz0ibCI+RGVidCBzZXJ2aWNlIGNvdmVyYWdlPC9kaXY+"
    "PGRpdiBjbGFzcz0idiI+JHtkc2NyLnRvRml4ZWQoMil9JnRpbWVzOzwvZGl2PjxkaXYgY2xhc3M9ImQiPk5PSSAmZGl2aWRlOyBk"
    "ZWJ0IHNlcnZpY2U8L2Rpdj48L2Rpdj4KICA8L3NlY3Rpb24+YDsKfQoKZnVuY3Rpb24gaGVyb0ZvcihzLCBsYWJlbCkgewogIGNv"
    "bnN0IG5vaSA9IHMueXRkKCJub2kiKSwgaW5jID0gcy55dGQoImluY29tZSIpOwogIGNvbnN0IGNhc2ggPSBzLnl0ZCgiY2FzaGZs"
    "b3ciKSwgZGVidCA9IHMueXRkKCJkZWJ0Iik7CiAgY29uc3QgbiA9IHMubk1vbnRocygpOwogIC8vIFRoZSBoZWFkbGluZSBpcyBj"
    "YXNoIGZsb3cgYWZ0ZXIgZGVidDsgdGhlIHN1cHBvcnRpbmcgbGluZSBjYXJyaWVzIE5PSSwgc28gdGhlCiAgLy8gdHdvIGZpZ3Vy"
    "ZXMgZWFjaCBhcHBlYXIgZXhhY3RseSBvbmNlIGluIHRoZSB0b3AgYm94LgogIGNvbnN0IHNlbENhc2ggPSBzLnNsaWNlKCJjYXNo"
    "ZmxvdyIpLCBzZWxNID0gcy5zZWxNb250aHMoKTsKICBjb25zdCBiZXN0ID0gc2VsQ2FzaC5pbmRleE9mKE1hdGgubWF4KC4uLnNl"
    "bENhc2gpKSwgd29yc3QgPSBzZWxDYXNoLmluZGV4T2YoTWF0aC5taW4oLi4uc2VsQ2FzaCkpOwogIGNvbnN0IHBlcmlvZCA9IG1G"
    "aWx0ZXJlZCgpID8gbUxhYmVsKCkgOiAieWVhciB0byBkYXRlIjsKICBjb25zdCBzaWRlID0gbiA9PT0gMQogICAgPyBgPGRpdj48"
    "ZGl2IGNsYXNzPSJsIj5JbmNvbWU8L2Rpdj48ZGl2IGNsYXNzPSJ2Ij4ke2NvbXBhY3QoaW5jKX08L2Rpdj48L2Rpdj4KICAgICAg"
    "IDxkaXY+PGRpdiBjbGFzcz0ibCI+T3BlcmF0aW5nIGV4cGVuc2VzPC9kaXY+PGRpdiBjbGFzcz0idiI+JHtjb21wYWN0KHMueXRk"
    "KCJleHBlbnNlcyIpKX08L2Rpdj48L2Rpdj4KICAgICAgIDxkaXY+PGRpdiBjbGFzcz0ibCI+TmV0IG9wZXJhdGluZyBpbmNvbWU8"
    "L2Rpdj48ZGl2IGNsYXNzPSJ2Ij4ke2NvbXBhY3Qobm9pKX08L2Rpdj48L2Rpdj5gCiAgICA6IGA8ZGl2PjxkaXYgY2xhc3M9Imwi"
    "PkJlc3QgbW9udGg8L2Rpdj48ZGl2IGNsYXNzPSJ2Ij4ke3NlbE1bYmVzdF19IMK3ICR7Y29tcGFjdChzZWxDYXNoW2Jlc3RdKX08"
    "L2Rpdj48L2Rpdj4KICAgICAgIDxkaXY+PGRpdiBjbGFzcz0ibCI+V2Vha2VzdCBtb250aDwvZGl2PjxkaXYgY2xhc3M9InYiPiR7"
    "c2VsTVt3b3JzdF19IMK3ICR7Y29tcGFjdChzZWxDYXNoW3dvcnN0XSl9PC9kaXY+PC9kaXY+CiAgICAgICA8ZGl2PjxkaXYgY2xh"
    "c3M9ImwiPk1vbnRobHkgYXZlcmFnZTwvZGl2PjxkaXYgY2xhc3M9InYiPiR7Y29tcGFjdChjYXNoIC8gbil9PC9kaXY+PC9kaXY+"
    "YDsKICByZXR1cm4gYDxzZWN0aW9uIGNsYXNzPSJoZXJvIj4KICAgIDxkaXY+CiAgICAgIDxkaXYgY2xhc3M9ImxhYmVsIj4ke2xh"
    "YmVsfSR7bUZpbHRlcmVkKCkgPyAiIMK3ICIgOiAiLCAifSR7cGVyaW9kfTwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJ2YWx1ZSAk"
    "e2Nhc2ggPj0gMCA/ICIiIDogIm5lZyJ9Ij4ke21vbmV5KGNhc2gpfTwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJoZXJvbm90ZSI+"
    "JHttb25leShub2kpfSBvZiBuZXQgb3BlcmF0aW5nIGluY29tZSR7aW5jID8gYCDigJQgYSAkeyhub2kgLyBpbmMgKiAxMDApLnRv"
    "Rml4ZWQoMSl9JSBtYXJnaW5gIDogIiJ9IOKAlCBsZXNzICR7bW9uZXkoZGVidCl9IG9mIGRlYnQgc2VydmljZSDCtyAke259IG1v"
    "bnRoJHtuID09PSAxID8gIiIgOiAicyJ9PC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgY2xhc3M9Imhlcm8tc2lkZSI+JHtzaWRl"
    "fTwvZGl2PgogIDwvc2VjdGlvbj5gOwp9Cgpjb25zdCBjaGFydENhcmQgPSAodGl0bGUsIGNhcCwgbGVnZW5kLCBpZCwgdmIpID0+"
    "IGA8c2VjdGlvbiBjbGFzcz0iY2FyZCI+CiAgPGgyPiR7dGl0bGV9PC9oMj48cCBjbGFzcz0iY2FwIj4ke2NhcH08L3A+JHtsZWdl"
    "bmR9CiAgPHN2ZyBpZD0iJHtpZH0iIHZpZXdCb3g9IiR7dmJ9IiByb2xlPSJpbWciIGFyaWEtbGFiZWw9IiR7ZXNjKHRpdGxlKX0i"
    "Pjwvc3ZnPgo8L3NlY3Rpb24+YDsKCmNvbnN0IExFR0VORF9JRSA9IGA8ZGl2IGNsYXNzPSJsZWdlbmQiPgogIDxzcGFuPjxpIGNs"
    "YXNzPSJrZXkiIHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLXNlcmllcy0xKSI+PC9pPkluY29tZTwvc3Bhbj4KICA8c3Bhbj48aSBj"
    "bGFzcz0ia2V5IiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS1zZXJpZXMtMikiPjwvaT5PcGVyYXRpbmcgZXhwZW5zZXM8L3NwYW4+"
    "PC9kaXY+YDsKY29uc3QgTEVHRU5EX05DID0gYDxkaXYgY2xhc3M9ImxlZ2VuZCI+CiAgPHNwYW4+PGkgY2xhc3M9ImtleSBsaW5l"
    "IiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS1zZXJpZXMtMSkiPjwvaT5OZXQgb3BlcmF0aW5nIGluY29tZTwvc3Bhbj4KICA8c3Bh"
    "bj48aSBjbGFzcz0ia2V5IGxpbmUiIHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLXNlcmllcy0yKSI+PC9pPkNhc2ggZmxvdyBhZnRl"
    "ciBkZWJ0PC9zcGFuPjwvZGl2PmA7CmNvbnN0IExFR0VORF9NSVggPSBgPGRpdiBjbGFzcz0ibGVnZW5kIj5gICsgR1JPVVBfTkFN"
    "RVMubWFwKChnLCBpKSA9PgogIGA8c3Bhbj48aSBjbGFzcz0ia2V5IiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS0ke0dST1VQX1ZB"
    "UltpXS5zbGljZSgyKX0pIj48L2k+JHtnfTwvc3Bhbj5gKS5qb2luKCIiKSArIGA8L2Rpdj5gOwoKZnVuY3Rpb24gdmFyaWFuY2VO"
    "b3RlKHMpIHsKICBjb25zdCB2ID0gcy55dGQoInZhcmlhbmNlIik7CiAgaWYgKE1hdGguYWJzKHYpIDwgMSkgcmV0dXJuICIiOwog"
    "IGNvbnN0IFtyYV0gPSBzLnJhbmdlKCk7CiAgY29uc3QgYmFkID0gcy5zZWxNb250aHMoKS5maWx0ZXIoKG0sIGkpID0+IE1hdGgu"
    "YWJzKHMudmFyaWFuY2VbcmEgKyBpXSkgPiAwLjAxKTsKICByZXR1cm4gYDxkaXYgY2xhc3M9Im5vdGUiPjxiPkEgc3VidG90YWwg"
    "ZGlzY3JlcGFuY3kgaW4gdGhlIHNvdXJjZSB3b3JrYm9vay48L2I+CiAgICBUaGUgZXhwZW5zZSBzdWJ0b3RhbCBleGNsdWRlcyBh"
    "biBpbnN1cmFuY2UgYW1vdW50IHRoYXQgaXMgcmVjb3JkZWQgaW4gdGhlIHNhbWUgY29sdW1uIGluCiAgICA8Yj4ke2JhZC5qb2lu"
    "KCIsICIpfTwvYj4g4oCUICR7bW9uZXkyKHYpfSBpbiB0b3RhbC4gRXZlcnkgZmlndXJlIGhlcmUgZm9sbG93cyB0aGUgd29ya2Jv"
    "b2sgYXMgcmVwb3J0ZWQuCiAgICBDb3VudGluZyB0aG9zZSBhbW91bnRzLCBvcGVyYXRpbmcgZXhwZW5zZXMgYXJlIDxiPiR7bW9u"
    "ZXkyKHMueXRkKCJleHBlbnNlc1JlY29yZGVkIikpfTwvYj4sCiAgICBOT0kgaXMgPGI+JHttb25leTIocy55dGQoImluY29tZSIp"
    "IC0gcy55dGQoImV4cGVuc2VzUmVjb3JkZWQiKSl9PC9iPiBhbmQgY2FzaCBmbG93IGFmdGVyIGRlYnQgaXMKICAgIDxiPiR7bW9u"
    "ZXkyKHMueXRkKCJpbmNvbWUiKSAtIHMueXRkKCJleHBlbnNlc1JlY29yZGVkIikgLSBzLnl0ZCgiZGVidCIpKX08L2I+LjwvZGl2"
    "PmA7Cn0KCmZ1bmN0aW9uIHNjb3BlTGFiZWwoKSB7CiAgY29uc3QgbiA9IGluY2x1ZGVkUHJvcHMoKS5sZW5ndGg7CiAgcmV0dXJu"
    "IGlzRmlsdGVyZWQoKSA/IGBhY3Jvc3MgdGhlICR7bn0gc2VsZWN0ZWQgcHJvcGVydCR7biA9PT0gMSA/ICJ5IiA6ICJpZXMifWAg"
    "OiAiYWNyb3NzIGFsbCBwcm9wZXJ0aWVzIjsKfQovKiBDaGFydHMga2VlcCBldmVyeSBtb250aCBvbiBzY3JlZW4sIHNvIHRoZWly"
    "IGNhcHRpb25zIHNheSBzbyBleHBsaWNpdGx5IC0gYSBkaW1tZWQKICAgY29sdW1uIGlzIGVhc3kgdG8gbWlzcmVhZCBhcyAibm8g"
    "ZGF0YSIgcmF0aGVyIHRoYW4gIm91dHNpZGUgeW91ciBzZWxlY3Rpb24iLiAqLwpjb25zdCBtb250aE5vdGUgPSAoKSA9PiBtRmls"
    "dGVyZWQoKQogID8gYCAke21MYWJlbCgpfSBoaWdobGlnaHRlZDsgdGhlIHJlc3Qgb2YgdGhlIHllYXIgc3RheXMgZm9yIGNvbnRl"
    "eHQuYCA6ICIiOwoKZnVuY3Rpb24gcmVuZGVyUG9ydGZvbGlvKCkgewogIGNvbnN0IHByb3BzID0gaW5jbHVkZWRQcm9wcygpOwog"
    "IGlmICghcHJvcHMubGVuZ3RoKSB7CiAgICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgidmlld1RpdGxlIikudGV4dENvbnRlbnQg"
    "PSBWSUVXLmxhYmVsOwogICAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoInZpZXdTdWIiKS50ZXh0Q29udGVudCA9ICJObyBwcm9w"
    "ZXJ0aWVzIHNlbGVjdGVkIjsKICAgIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCJib2R5IikuaW5uZXJIVE1MID0KICAgICAgYDxk"
    "aXYgY2xhc3M9ImVtcHR5c3RhdGUiPk5vIHByb3BlcnRpZXMgc2VsZWN0ZWQuIFBpY2sgYXQgbGVhc3Qgb25lIGZyb20gdGhlIGZp"
    "bHRlciBhYm92ZS48L2Rpdj5gOwogICAgcmV0dXJuOwogIH0KICBjb25zdCBzID0gc2VyaWVzRm9yKCJfX0FMTF9fIik7CiAgY29u"
    "c3Qgcm93cyA9IHByb3BzLm1hcChuYW1lID0+IHsKICAgIGNvbnN0IHAgPSBzZXJpZXNGb3IobmFtZSk7CiAgICBjb25zdCBub2kg"
    "PSBwLnl0ZCgibm9pIiksIGluYyA9IHAueXRkKCJpbmNvbWUiKTsKICAgIHJldHVybiB7IG5hbWUsIHZhbHVlOiBwLnl0ZCgiY2Fz"
    "aGZsb3ciKSwgbm9pLCBkZWJ0OiBwLnl0ZCgiZGVidCIpLAogICAgICBtYXJnaW46IChpbmMgPyAobm9pIC8gaW5jICogMTAwKS50"
    "b0ZpeGVkKDEpIDogIjAiKSArICIlIiB9OwogIH0pLnNvcnQoKGEsIGIpID0+IGIudmFsdWUgLSBhLnZhbHVlKTsKICBjb25zdCB3"
    "aW5uZXJzID0gcm93cy5maWx0ZXIociA9PiByLnZhbHVlID49IDApLmxlbmd0aDsKCiAgY29uc3QgdG90YWwgPSBhbGxQcm9wcygp"
    "Lmxlbmd0aDsKICBjb25zdCBleGNsdWRlZCA9IGFsbFByb3BzKCkuZmlsdGVyKHAgPT4gIXByb3BzLmluY2x1ZGVzKHApKTsKICBk"
    "b2N1bWVudC5nZXRFbGVtZW50QnlJZCgidmlld1RpdGxlIikudGV4dENvbnRlbnQgPSBWSUVXLmxhYmVsOwogIGRvY3VtZW50Lmdl"
    "dEVsZW1lbnRCeUlkKCJ2aWV3U3ViIikudGV4dENvbnRlbnQgPQogICAgKGlzRmlsdGVyZWQoKSA/IGAke3Byb3BzLmxlbmd0aH0g"
    "b2YgJHt0b3RhbH0gcHJvcGVydGllc2AgOiBgJHt0b3RhbH0gcHJvcGVydGllc2ApICsKICAgIGAgwrcgJHttRmlsdGVyZWQoKSA/"
    "IG1MYWJlbCgpIDogUC5tb250aHNbMF0gKyAiXHUyMDEzIiArIFAudGhyb3VnaE1vbnRoICsgIiAiICsgUC55ZWFyfSDCtyBhY2Ny"
    "dWFsIGJhc2lzYCArCiAgICAoZXhjbHVkZWQubGVuZ3RoICYmIGV4Y2x1ZGVkLmxlbmd0aCA8PSA0ID8gYCDCtyBleGNsdWRpbmcg"
    "JHtleGNsdWRlZC5qb2luKCIsICIpfWAgOiAiIik7CgogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCJib2R5IikuaW5uZXJIVE1M"
    "ID0KICAgIGhlcm9Gb3IocywgIlBvcnRmb2xpbyBjYXNoIGZsb3cgYWZ0ZXIgZGVidCIpICsga3BpVGlsZXMocykgKwogICAgYDxk"
    "ZXRhaWxzIGNsYXNzPSJ0YWJsZXdyYXAiPjxzdW1tYXJ5PlRhYmxlIHZpZXcg4oCUICR7aXNGaWx0ZXJlZCgpID8gZXNjKFZJRVcu"
    "bGFiZWwpIDogInBvcnRmb2xpbyJ9LCBmdWxsIG1vbnRobHkgZGV0YWlsPC9zdW1tYXJ5PiR7dGFibGVGb3Iocyl9PC9kZXRhaWxz"
    "PmAgKwogICAgYDxkaXYgY2xhc3M9ImdyaWQyIj4KICAgICAgJHtjaGFydENhcmQoIkluY29tZSB2cy4gb3BlcmF0aW5nIGV4cGVu"
    "c2VzIiwgYE1vbnRobHksICR7c2NvcGVMYWJlbCgpfS4ke21vbnRoTm90ZSgpfWAsIExFR0VORF9JRSwgImNJRSIsICIwIDAgNTIw"
    "IDMwMCIpfQogICAgICAke2NoYXJ0Q2FyZCgiTk9JIGFuZCBjYXNoIGZsb3cgYWZ0ZXIgZGVidCBzZXJ2aWNlIiwgYE1vbnRobHks"
    "ICR7c2NvcGVMYWJlbCgpfS4ke21vbnRoTm90ZSgpfWAsIExFR0VORF9OQywgImNOQyIsICIwIDAgNTIwIDMwMCIpfQogICAgPC9k"
    "aXY+YCArCiAgICBjaGFydENhcmQoIkNhc2ggZmxvdyBhZnRlciBkZWJ0IHNlcnZpY2UsIGJ5IHByb3BlcnR5IiwKICAgICAgYCR7"
    "bUZpbHRlcmVkKCkgPyBtTGFiZWwoKSA6ICJZZWFyIHRvIGRhdGUifS4gJHt3aW5uZXJzfSBvZiAke3Byb3BzLmxlbmd0aH0gJHtp"
    "c0ZpbHRlcmVkKCkgPyAic2VsZWN0ZWQgIiA6ICIifXByb3BlcnRpZXMgYXJlIGNhc2gtZmxvdyBwb3NpdGl2ZSBhZnRlciBkZWJ0"
    "LiBDbGljayBhbnkgcm93IHRvIG9wZW4gaXQuYCwKICAgICAgYDxkaXYgY2xhc3M9ImxlZ2VuZCI+PHNwYW4+PGkgY2xhc3M9Imtl"
    "eSIgc3R5bGU9ImJhY2tncm91bmQ6dmFyKC0tcG9zKSI+PC9pPlBvc2l0aXZlPC9zcGFuPjxzcGFuPjxpIGNsYXNzPSJrZXkiIHN0"
    "eWxlPSJiYWNrZ3JvdW5kOnZhcigtLW5lZ2IpIj48L2k+TmVnYXRpdmU8L3NwYW4+PC9kaXY+YCwKICAgICAgImNSYW5rIiwgIjAg"
    "MCAxMDQwIDQwMCIpICsKICAgIGNoYXJ0Q2FyZCgiRXhwZW5zZSBtaXgsIG1vbnRoIGJ5IG1vbnRoIiwgYFJlY29yZGVkIGxpbmUg"
    "aXRlbXMsIGdyb3VwZWQsICR7c2NvcGVMYWJlbCgpfS4ke21vbnRoTm90ZSgpfWAsIExFR0VORF9NSVgsICJjTWl4IiwgIjAgMCAx"
    "MDQwIDMyMCIpICsKICAgIGNoYXJ0Q2FyZCgiT3BlcmF0aW5nIHNwZW5kIGJ5IGNhdGVnb3J5IiwgYCR7bUZpbHRlcmVkKCkgPyBt"
    "TGFiZWwoKSA6ICJZZWFyLXRvLWRhdGUifSB0b3RhbHMgJHtzY29wZUxhYmVsKCl9LmAsICIiLCAiY0JhcnMiLCAiMCAwIDEwNDAg"
    "NDAwIikgKwogICAgdmFyaWFuY2VOb3RlKHMpOwoKICBjaGFydEluY29tZUV4cGVuc2UoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQo"
    "ImNJRSIpLCBzKTsKICBjaGFydE5vaUNhc2goZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoImNOQyIpLCBzKTsKICBjaGFydFJhbmtp"
    "bmcoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoImNSYW5rIiksIHJvd3MsIHBpY2spOwogIGNoYXJ0RXhwZW5zZU1peChkb2N1bWVu"
    "dC5nZXRFbGVtZW50QnlJZCgiY01peCIpLCBzKTsKICBjaGFydEV4cGVuc2VCYXJzKGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCJj"
    "QmFycyIpLCBzKTsKfQoKZnVuY3Rpb24gcmVuZGVyUHJvcGVydHkobmFtZSkgewogIGNvbnN0IHMgPSBzZXJpZXNGb3IobmFtZSk7"
    "CiAgY29uc3QgZW50aXR5ID0gT2JqZWN0LmtleXMoUC5lbnRpdGllcykuZmluZChlID0+IFAuZW50aXRpZXNbZV0uaW5jbHVkZXMo"
    "bmFtZSkpIHx8ICIiOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCJ2aWV3VGl0bGUiKS50ZXh0Q29udGVudCA9IG5hbWU7CiAg"
    "ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoInZpZXdTdWIiKS50ZXh0Q29udGVudCA9CiAgICBgJHtlbnRpdHkgPyBlbnRpdHkgKyAi"
    "IMK3ICIgOiAiIn0ke21GaWx0ZXJlZCgpID8gbUxhYmVsKCkgOiBQLm1vbnRoc1swXSArICJcdTIwMTMiICsgUC50aHJvdWdoTW9u"
    "dGggKyAiICIgKyBQLnllYXJ9IMK3IGFjY3J1YWwgYmFzaXNgOwoKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgiYm9keSIpLmlu"
    "bmVySFRNTCA9CiAgICBoZXJvRm9yKHMsICJDYXNoIGZsb3cgYWZ0ZXIgZGVidCIpICsga3BpVGlsZXMocykgKwogICAgYDxkZXRh"
    "aWxzIGNsYXNzPSJ0YWJsZXdyYXAiPjxzdW1tYXJ5PlRhYmxlIHZpZXcg4oCUIGZ1bGwgbW9udGhseSBkZXRhaWw8L3N1bW1hcnk+"
    "JHt0YWJsZUZvcihzKX08L2RldGFpbHM+YCArCiAgICBgPGRpdiBjbGFzcz0iZ3JpZDIiPgogICAgICAke2NoYXJ0Q2FyZCgiSW5j"
    "b21lIHZzLiBvcGVyYXRpbmcgZXhwZW5zZXMiLCBgTW9udGhseSwgaW4gZG9sbGFycy4ke21vbnRoTm90ZSgpfWAsIExFR0VORF9J"
    "RSwgImNJRSIsICIwIDAgNTIwIDMwMCIpfQogICAgICAke2NoYXJ0Q2FyZCgiTk9JIGFuZCBjYXNoIGZsb3cgYWZ0ZXIgZGVidCBz"
    "ZXJ2aWNlIiwgYE1vbnRobHkuIENhc2ggZmxvdyBpcyBOT0kgbGVzcyBtb3J0Z2FnZSBwYXltZW50cy4ke21vbnRoTm90ZSgpfWAs"
    "IExFR0VORF9OQywgImNOQyIsICIwIDAgNTIwIDMwMCIpfQogICAgPC9kaXY+YCArCiAgICBjaGFydENhcmQoIldoZXJlIHRoZSBv"
    "cGVyYXRpbmcgc3BlbmQgd2VudCIsIGAke21GaWx0ZXJlZCgpID8gbUxhYmVsKCkgOiAiWWVhci10by1kYXRlIn0gdG90YWwgYnkg"
    "Y2F0ZWdvcnksIGFzIHJlY29yZGVkIG9uIGVhY2ggbGluZS5gLCAiIiwgImNCYXJzIiwgIjAgMCAxMDQwIDQwMCIpICsKICAgIGNo"
    "YXJ0Q2FyZCgiRXhwZW5zZSBtaXgsIG1vbnRoIGJ5IG1vbnRoIiwgIkZpeGVkIGNvc3RzIGFnYWluc3QgdGhlIHZhcmlhYmxlIG9u"
    "ZXMuIiwgTEVHRU5EX01JWCwgImNNaXgiLCAiMCAwIDEwNDAgMzIwIikgKwogICAgdmFyaWFuY2VOb3RlKHMpOwoKICBjaGFydElu"
    "Y29tZUV4cGVuc2UoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoImNJRSIpLCBzKTsKICBjaGFydE5vaUNhc2goZG9jdW1lbnQuZ2V0"
    "RWxlbWVudEJ5SWQoImNOQyIpLCBzKTsKICBjaGFydEV4cGVuc2VCYXJzKGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCJjQmFycyIp"
    "LCBzKTsKICBjaGFydEV4cGVuc2VNaXgoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoImNNaXgiKSwgcyk7Cn0KCi8qID09PT09PT09"
    "PT09PT09PT09PT09PT09PT09PT0gZmlsdGVyIFVJID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gKi8KY29uc3QgJCA9IGlk"
    "ID0+IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKGlkKTsKCi8qIC0tLS0gbW9udGggY2hpcHMgLS0tLSAqLwpmdW5jdGlvbiBzZXRN"
    "b250aHMocmFuZ2UpIHsKICBNU0VMID0gcmFuZ2U7CiAgd3JpdGVIYXNoKCk7CiAgcmVuZGVyTW9udGhzKCk7CiAgZHJhdygpOwp9"
    "CmZ1bmN0aW9uIHJlbmRlck1vbnRocygpIHsKICBjb25zdCBbYSwgYl0gPSBNU0VMLCBsYXN0ID0gUC5tb250aHMubGVuZ3RoIC0g"
    "MTsKICBjb25zdCBhbGwgPSAhbUZpbHRlcmVkKCk7CiAgJCgibW9udGhSb3ciKS5pbm5lckhUTUwgPQogICAgYDxidXR0b24gdHlw"
    "ZT0iYnV0dG9uIiBjbGFzcz0ibWNoaXAiIGRhdGEtYWxsPSIxIiBhcmlhLXByZXNzZWQ9IiR7YWxsfSI+WWVhciB0byBkYXRlPC9i"
    "dXR0b24+CiAgICAgPHNwYW4gY2xhc3M9InNlcCI+PC9zcGFuPmAgKwogICAgUC5tb250aHMubWFwKChtbywgaSkgPT4gewogICAg"
    "ICBjb25zdCBzaW5nbGUgPSBhID09PSBiICYmIGkgPT09IGE7CiAgICAgIGNvbnN0IGluUmFuZ2UgPSAhYWxsICYmICFzaW5nbGUg"
    "JiYgaSA+PSBhICYmIGkgPD0gYjsKICAgICAgcmV0dXJuIGA8YnV0dG9uIHR5cGU9ImJ1dHRvbiIgY2xhc3M9Im1jaGlwJHtpblJh"
    "bmdlID8gIiBpbnJhbmdlIiA6ICIifSIgZGF0YS1pPSIke2l9IgogICAgICAgIGFyaWEtcHJlc3NlZD0iJHtzaW5nbGV9Ij4ke21v"
    "fTwvYnV0dG9uPmA7CiAgICB9KS5qb2luKCIiKSArCiAgICAoYWxsID8gYDxzcGFuIGNsYXNzPSJoaW50Ij5zaGlmdC1jbGljayBm"
    "b3IgYSByYW5nZTwvc3Bhbj5gIDogIiIpOwp9CiQoIm1vbnRoUm93IikuYWRkRXZlbnRMaXN0ZW5lcigiY2xpY2siLCBlID0+IHsK"
    "ICBjb25zdCBidG4gPSBlLnRhcmdldC5jbG9zZXN0KCJidXR0b24iKTsKICBpZiAoIWJ0bikgcmV0dXJuOwogIGlmIChidG4uZGF0"
    "YXNldC5hbGwpIHJldHVybiBzZXRNb250aHMobUFsbCgpKTsKICBjb25zdCBpID0gK2J0bi5kYXRhc2V0Lmk7CiAgaWYgKGUuc2hp"
    "ZnRLZXkpIHsKICAgIGNvbnN0IGFuY2hvciA9IE1TRUxbMF07CiAgICBzZXRNb250aHMoW01hdGgubWluKGFuY2hvciwgaSksIE1h"
    "dGgubWF4KGFuY2hvciwgaSldKTsKICB9IGVsc2UgewogICAgc2V0TW9udGhzKE1TRUxbMF0gPT09IGkgJiYgTVNFTFsxXSA9PT0g"
    "aSA/IG1BbGwoKSA6IFtpLCBpXSk7CiAgfQp9KTsKCmZ1bmN0aW9uIHNldEluY2x1ZGUoc2V0LCBsYWJlbCwgdHlwZSkgewogIFZJ"
    "RVcuaW5jbHVkZSA9IHNldCA/IG5ldyBTZXQoc2V0KSA6IG51bGw7CiAgVklFVy5sYWJlbCA9IGxhYmVsOwogIFZJRVcudHlwZSA9"
    "IHR5cGUgfHwgInBvcnRmb2xpbyI7Cn0KCmZ1bmN0aW9uIGVudGl0eU9mKHByb3ApIHsKICByZXR1cm4gT2JqZWN0LmtleXMoUC5l"
    "bnRpdGllcykuZmluZChlID0+IFAuZW50aXRpZXNbZV0uaW5jbHVkZXMocHJvcCkpIHx8ICJPdGhlciI7Cn0KZnVuY3Rpb24gcHJv"
    "cHNCeUVudGl0eSgpIHsKICBjb25zdCBvdXQgPSB7fTsKICBhbGxQcm9wcygpLmZvckVhY2gocCA9PiAob3V0W2VudGl0eU9mKHAp"
    "XSA9IG91dFtlbnRpdHlPZihwKV0gfHwgW10pLnB1c2gocCkpOwogIHJldHVybiBvdXQ7Cn0KCmZ1bmN0aW9uIGJ1aWxkRmlsdGVy"
    "UGFuZWwoKSB7CiAgY29uc3QgaW5jID0gbmV3IFNldChpbmNsdWRlZFByb3BzKCkpOwogIGNvbnN0IGJ5RW50ID0gcHJvcHNCeUVu"
    "dGl0eSgpOwogIGxldCBoID0gYDxkaXYgY2xhc3M9ImZoZWFkIj4KICAgICAgPGJ1dHRvbiB0eXBlPSJidXR0b24iIGRhdGEtYWxs"
    "PSIxIj5TZWxlY3QgYWxsPC9idXR0b24+CiAgICAgIDxidXR0b24gdHlwZT0iYnV0dG9uIiBkYXRhLW5vbmU9IjEiPkNsZWFyIGFs"
    "bDwvYnV0dG9uPgogICAgPC9kaXY+YDsKICBmb3IgKGNvbnN0IGVudCBpbiBieUVudCkgewogICAgaCArPSBgPGRpdiBjbGFzcz0i"
    "Zmdyb3VwIj48c3Bhbj4ke2VzYyhlbnQpfTwvc3Bhbj48YnV0dG9uIHR5cGU9ImJ1dHRvbiIgZGF0YS1lbnQ9IiR7ZXNjKGVudCl9"
    "Ij5vbmx5PC9idXR0b24+PC9kaXY+YDsKICAgIGggKz0gYnlFbnRbZW50XS5tYXAocCA9PiB7CiAgICAgIGNvbnN0IG9uID0gaW5j"
    "LmhhcyhwKTsKICAgICAgcmV0dXJuIGA8bGFiZWwgY2xhc3M9ImZpdGVtICR7b24gPyAiIiA6ICJvZmYifSI+CiAgICAgICAgPGlu"
    "cHV0IHR5cGU9ImNoZWNrYm94IiBkYXRhLXByb3A9IiR7ZXNjKHApfSIgJHtvbiA/ICJjaGVja2VkIiA6ICIifT4KICAgICAgICA8"
    "c3Bhbj4ke2VzYyhwKX08L3NwYW4+PC9sYWJlbD5gOwogICAgfSkuam9pbigiIik7CiAgfQogIGggKz0gYDxkaXYgY2xhc3M9ImZo"
    "ZWFkIiBzdHlsZT0iYm9yZGVyLWJvdHRvbTowO2JvcmRlci10b3A6MXB4IHNvbGlkIHZhcigtLWdyaWQpO21hcmdpbjo4cHggMCAw"
    "O3BhZGRpbmc6MTBweCA2cHggMnB4Ij4KICAgICAgPGJ1dHRvbiB0eXBlPSJidXR0b24iIGRhdGEtc2F2ZT0iMSI+U2F2ZSBzZWxl"
    "Y3Rpb24gYXMgYSBncm91cOKApjwvYnV0dG9uPgogICAgPC9kaXY+YDsKICBpZiAoR1JPVVBTLmxlbmd0aCkgewogICAgaCArPSBg"
    "PGRpdiBjbGFzcz0iZmdyb3VwIj48c3Bhbj5TYXZlZCBncm91cHM8L3NwYW4+PC9kaXY+YDsKICAgIGggKz0gR1JPVVBTLm1hcChn"
    "ID0+IGA8ZGl2IGNsYXNzPSJmaXRlbSIgc3R5bGU9Imp1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuIj4KICAgICAgICA8c3Bh"
    "biBzdHlsZT0iY3Vyc29yOnBvaW50ZXIiIGRhdGEtb3Blbj0iJHtlc2MoZy5pZCl9Ij4ke2VzYyhnLm5hbWUpfSA8c3BhbiBzdHls"
    "ZT0iY29sb3I6dmFyKC0tdGV4dC1tdXRlZCkiPigke2cucHJvcHMubGVuZ3RofSk8L3NwYW4+PC9zcGFuPgogICAgICAgIDxidXR0"
    "b24gdHlwZT0iYnV0dG9uIiBkYXRhLWRlbD0iJHtlc2MoZy5pZCl9IiBzdHlsZT0iYmFja2dyb3VuZDp0cmFuc3BhcmVudDtib3Jk"
    "ZXI6MDtjb2xvcjp2YXIoLS10ZXh0LW11dGVkKTtjdXJzb3I6cG9pbnRlcjtmb250LXNpemU6MTVweCI+JnRpbWVzOzwvYnV0dG9u"
    "PgogICAgICA8L2Rpdj5gKS5qb2luKCIiKTsKICB9CiAgJCgiZmlsdGVyUGFuZWwiKS5pbm5lckhUTUwgPSBoOwp9CgovKiBUaGUg"
    "Y2hpcCByb3cgbGlzdHMgd2hhdCBpcyBJTiB0aGUgY3VycmVudCB2aWV3IC0gdGhhdCBpcyB0aGUgdGhpbmcgYSByZWFkZXIKICAg"
    "bmVlZHMgdG8ga25vdywgYW5kIHdpdGggMiBvZiAxNyBzZWxlY3RlZCB0aGUgaW52ZXJzZSB3YXMgYSB3YWxsIG9mIDE1IGNoaXBz"
    "LgogICBQYXN0IENISVBfQ0FQIHRoZSBsaXN0IGNvbGxhcHNlcyBzbyBhIG5lYXItY29tcGxldGUgc2VsZWN0aW9uIHN0YXlzIHJl"
    "YWRhYmxlLiAqLwpjb25zdCBDSElQX0NBUCA9IDEwOwpsZXQgQ0hJUFNfRVhQQU5ERUQgPSBmYWxzZTsKCmZ1bmN0aW9uIHVwZGF0"
    "ZUZpbHRlclVJKCkgewogIGNvbnN0IGluYyA9IGluY2x1ZGVkUHJvcHMoKSwgdG90YWwgPSBhbGxQcm9wcygpLmxlbmd0aDsKICAk"
    "KCJmaWx0ZXJSb3ciKS5oaWRkZW4gPSBWSUVXLnR5cGUgPT09ICJwcm9wZXJ0eSI7CiAgJCgiZmlsdGVyTGFiZWwiKS50ZXh0Q29u"
    "dGVudCA9IGlzRmlsdGVyZWQoKSA/IGAke2luYy5sZW5ndGh9IG9mICR7dG90YWx9IHByb3BlcnRpZXNgIDogYEFsbCAke3RvdGFs"
    "fSBwcm9wZXJ0aWVzYDsKICAkKCJmaWx0ZXJCdG4iKS5jbGFzc0xpc3QudG9nZ2xlKCJhY3RpdmUiLCBpc0ZpbHRlcmVkKCkpOwog"
    "IGNvbnN0IGNoaXBzID0gJCgiZmlsdGVyQ2hpcHMiKTsKICBpZiAoIWlzRmlsdGVyZWQoKSkgeyBjaGlwcy5pbm5lckhUTUwgPSAi"
    "IjsgQ0hJUFNfRVhQQU5ERUQgPSBmYWxzZTsgcmV0dXJuOyB9CgogIGNvbnN0IHNob3duID0gQ0hJUFNfRVhQQU5ERUQgPyBpbmMg"
    "OiBpbmMuc2xpY2UoMCwgQ0hJUF9DQVApOwogIGNvbnN0IG9ubHkgPSBpbmMubGVuZ3RoID09PSAxOyAgIC8vIHRoZSB2aWV3IG11"
    "c3Qga2VlcCBhdCBsZWFzdCBvbmUgcHJvcGVydHkKICBsZXQgaCA9IGA8c3BhbiBjbGFzcz0ibGJsIj5TaG93aW5nOjwvc3Bhbj5g"
    "ICsgc2hvd24ubWFwKHAgPT4KICAgIGA8c3BhbiBjbGFzcz0iY2hpcCI+JHtlc2MocCl9YCArCiAgICAob25seSA/ICIiIDogYDxi"
    "dXR0b24gdHlwZT0iYnV0dG9uIiBkYXRhLXJlbW92ZT0iJHtlc2MocCl9IiBhcmlhLWxhYmVsPSJSZW1vdmUgJHtlc2MocCl9IGZy"
    "b20gdGhpcyB2aWV3IiB0aXRsZT0iUmVtb3ZlIGZyb20gdGhpcyB2aWV3Ij4mdGltZXM7PC9idXR0b24+YCkgKwogICAgYDwvc3Bh"
    "bj5gKS5qb2luKCIiKTsKICBpZiAoaW5jLmxlbmd0aCA+IENISVBfQ0FQKSB7CiAgICBoICs9IGA8YnV0dG9uIHR5cGU9ImJ1dHRv"
    "biIgY2xhc3M9Im1vcmVidG4iIGRhdGEtbW9yZT0iMSI+YCArCiAgICAgIChDSElQU19FWFBBTkRFRCA/ICJzaG93IGZld2VyIiA6"
    "IGArJHtpbmMubGVuZ3RoIC0gQ0hJUF9DQVB9IG1vcmVgKSArIGA8L2J1dHRvbj5gOwogIH0KICBjb25zdCBleE4gPSB0b3RhbCAt"
    "IGluYy5sZW5ndGg7CiAgaWYgKGV4TikgaCArPSBgPGJ1dHRvbiB0eXBlPSJidXR0b24iIGNsYXNzPSJtb3JlYnRuIiBkYXRhLWFk"
    "ZGJhY2s9IjEiIHRpdGxlPSJPcGVuIHRoZSBmaWx0ZXIgdG8gYWRkIHByb3BlcnRpZXMgYmFjayI+JHtleE59IGV4Y2x1ZGVkPC9i"
    "dXR0b24+YDsKICBjaGlwcy5pbm5lckhUTUwgPSBoOwp9CgpmdW5jdGlvbiBhcHBseVNlbGVjdGlvbihzZXQsIGxhYmVsKSB7CiAg"
    "Y29uc3QgYXJyID0gWy4uLnNldF07CiAgaWYgKGFyci5sZW5ndGggPD0gQ0hJUF9DQVApIENISVBTX0VYUEFOREVEID0gZmFsc2U7"
    "CiAgc2V0SW5jbHVkZShhcnIubGVuZ3RoID09PSBhbGxQcm9wcygpLmxlbmd0aCA/IG51bGwgOiBhcnIsIGxhYmVsIHx8ICJDdXN0"
    "b20gc2VsZWN0aW9uIik7CiAgd3JpdGVIYXNoKCk7CiAgc3luY1NlbGVjdG9yKCk7CiAgZHJhdygpOwp9CgokKCJmaWx0ZXJCdG4i"
    "KS5hZGRFdmVudExpc3RlbmVyKCJjbGljayIsICgpID0+IHsKICBjb25zdCBwID0gJCgiZmlsdGVyUGFuZWwiKSwgb3BlbiA9IHAu"
    "aGlkZGVuOwogIGlmIChvcGVuKSBidWlsZEZpbHRlclBhbmVsKCk7CiAgcC5oaWRkZW4gPSAhb3BlbjsKICAkKCJmaWx0ZXJCdG4i"
    "KS5zZXRBdHRyaWJ1dGUoImFyaWEtZXhwYW5kZWQiLCBTdHJpbmcob3BlbikpOwp9KTsKZG9jdW1lbnQuYWRkRXZlbnRMaXN0ZW5l"
    "cigiY2xpY2siLCBlID0+IHsKICBpZiAoJCgiZmlsdGVyUGFuZWwiKS5oaWRkZW4pIHJldHVybjsKICAvLyBjb21wb3NlZFBhdGgo"
    "KSBpcyBjYXB0dXJlZCBhdCBkaXNwYXRjaCwgc28gdGhpcyBzdGlsbCByZXBvcnRzIHRoZSB0cnVlIG9yaWdpbgogIC8vIGV2ZW4g"
    "d2hlbiBhIGhhbmRsZXIgdXBzdHJlYW0gaGFzIGFscmVhZHkgcmUtcmVuZGVyZWQgdGhlIHBhbmVsJ3MgY29udGVudHMKICAvLyAo"
    "ZS50YXJnZXQgd291bGQgYnkgdGhlbiBiZSBkZXRhY2hlZCwgYW5kIGNsb3Nlc3QoKSB3b3VsZCB3cm9uZ2x5IHNheSAib3V0c2lk"
    "ZSIpLgogIGlmIChlLmNvbXBvc2VkUGF0aCgpLmluY2x1ZGVzKCQoImZpbHRlclBhbmVsIikpIHx8IGUuY29tcG9zZWRQYXRoKCku"
    "aW5jbHVkZXMoJCgiZmlsdGVyQnRuIikpKSByZXR1cm47CiAgJCgiZmlsdGVyUGFuZWwiKS5oaWRkZW4gPSB0cnVlOwogICQoImZp"
    "bHRlckJ0biIpLnNldEF0dHJpYnV0ZSgiYXJpYS1leHBhbmRlZCIsICJmYWxzZSIpOwp9KTsKJCgiZmlsdGVyQ2hpcHMiKS5hZGRF"
    "dmVudExpc3RlbmVyKCJjbGljayIsIGUgPT4gewogIGNvbnN0IGJ0biA9IGUudGFyZ2V0LmNsb3Nlc3QoImJ1dHRvbiIpOwogIGlm"
    "ICghYnRuKSByZXR1cm47CiAgY29uc3QgZCA9IGJ0bi5kYXRhc2V0OwogIGlmIChkLm1vcmUpIHsgQ0hJUFNfRVhQQU5ERUQgPSAh"
    "Q0hJUFNfRVhQQU5ERUQ7IHVwZGF0ZUZpbHRlclVJKCk7IHJldHVybjsgfQogIGlmIChkLmFkZGJhY2spIHsKICAgIGJ1aWxkRmls"
    "dGVyUGFuZWwoKTsKICAgICQoImZpbHRlclBhbmVsIikuaGlkZGVuID0gZmFsc2U7CiAgICAkKCJmaWx0ZXJCdG4iKS5zZXRBdHRy"
    "aWJ1dGUoImFyaWEtZXhwYW5kZWQiLCAidHJ1ZSIpOwogICAgcmV0dXJuOwogIH0KICBpZiAoZC5yZW1vdmUpIHsKICAgIGNvbnN0"
    "IHNldCA9IG5ldyBTZXQoaW5jbHVkZWRQcm9wcygpKTsKICAgIHNldC5kZWxldGUoZC5yZW1vdmUpOwogICAgaWYgKCFzZXQuc2l6"
    "ZSkgcmV0dXJuOyAgICAgICAgICAgICAgICAgIC8vIG5ldmVyIGxlYXZlIHRoZSB2aWV3IGVtcHR5CiAgICBhcHBseVNlbGVjdGlv"
    "bihzZXQpOwogIH0KfSk7CiQoImZpbHRlclBhbmVsIikuYWRkRXZlbnRMaXN0ZW5lcigiY2hhbmdlIiwgZSA9PiB7CiAgY29uc3Qg"
    "cHJvcCA9IGUudGFyZ2V0LmRhdGFzZXQucHJvcDsKICBpZiAoIXByb3ApIHJldHVybjsKICBjb25zdCBzZXQgPSBuZXcgU2V0KGlu"
    "Y2x1ZGVkUHJvcHMoKSk7CiAgZS50YXJnZXQuY2hlY2tlZCA/IHNldC5hZGQocHJvcCkgOiBzZXQuZGVsZXRlKHByb3ApOwogIGlm"
    "ICghc2V0LnNpemUpIHsgZS50YXJnZXQuY2hlY2tlZCA9IHRydWU7IHJldHVybjsgfSAgIC8vIG5ldmVyIGxlYXZlIHRoZSB2aWV3"
    "IGVtcHR5CiAgYXBwbHlTZWxlY3Rpb24oc2V0KTsKICBidWlsZEZpbHRlclBhbmVsKCk7Cn0pOwokKCJmaWx0ZXJQYW5lbCIpLmFk"
    "ZEV2ZW50TGlzdGVuZXIoImNsaWNrIiwgZSA9PiB7CiAgY29uc3QgZCA9IGUudGFyZ2V0LmRhdGFzZXQ7CiAgaWYgKGQuYWxsKSB7"
    "IGFwcGx5U2VsZWN0aW9uKG5ldyBTZXQoYWxsUHJvcHMoKSksICJQb3J0Zm9saW8iKTsgYnVpbGRGaWx0ZXJQYW5lbCgpOyB9CiAg"
    "ZWxzZSBpZiAoZC5ub25lKSB7IC8qIGNsZWFyaW5nIGV2ZXJ5dGhpbmcgd291bGQgbGVhdmUgbm90aGluZyB0byBzaG93ICovCiAg"
    "ICBjb25zdCBmaXJzdCA9IGFsbFByb3BzKClbMF07CiAgICBhcHBseVNlbGVjdGlvbihuZXcgU2V0KFtmaXJzdF0pKTsgYnVpbGRG"
    "aWx0ZXJQYW5lbCgpOwogIH0KICBlbHNlIGlmIChkLmVudCkgeyBhcHBseVNlbGVjdGlvbihuZXcgU2V0KFAuZW50aXRpZXNbZC5l"
    "bnRdIHx8IFtdKSwgZC5lbnQpOyBidWlsZEZpbHRlclBhbmVsKCk7IH0KICBlbHNlIGlmIChkLnNhdmUpIHsgc2F2ZUN1cnJlbnRB"
    "c0dyb3VwKCk7IH0KICBlbHNlIGlmIChkLm9wZW4pIHsgY29uc3QgZyA9IEdST1VQUy5maW5kKHggPT4geC5pZCA9PT0gZC5vcGVu"
    "KTsgaWYgKGcpIG9wZW5Hcm91cChnKTsgfQogIGVsc2UgaWYgKGQuZGVsKSB7CiAgICBHUk9VUFMgPSBHUk9VUFMuZmlsdGVyKHgg"
    "PT4geC5pZCAhPT0gZC5kZWwpOyBzYXZlR3JvdXBzKEdST1VQUyk7CiAgICBidWlsZEZpbHRlclBhbmVsKCk7IHN5bmNTZWxlY3Rv"
    "cigpOwogIH0KfSk7CgpmdW5jdGlvbiBzYXZlQ3VycmVudEFzR3JvdXAoKSB7CiAgY29uc3QgcHJvcHMgPSBpbmNsdWRlZFByb3Bz"
    "KCk7CiAgY29uc3QgbmFtZSA9IChwcm9tcHQoIk5hbWUgdGhpcyBncm91cCIsIFZJRVcubGFiZWwgPT09ICJDdXN0b20gc2VsZWN0"
    "aW9uIiA/ICIiIDogVklFVy5sYWJlbCkgfHwgIiIpLnRyaW0oKTsKICBpZiAoIW5hbWUpIHJldHVybjsKICBjb25zdCBleGlzdGlu"
    "ZyA9IEdST1VQUy5maW5kKGcgPT4gZy5uYW1lLnRvTG93ZXJDYXNlKCkgPT09IG5hbWUudG9Mb3dlckNhc2UoKSk7CiAgaWYgKGV4"
    "aXN0aW5nKSBleGlzdGluZy5wcm9wcyA9IHByb3BzOwogIGVsc2UgR1JPVVBTLnB1c2goeyBpZDogImciICsgRGF0ZS5ub3coKS50"
    "b1N0cmluZygzNiksIG5hbWUsIHByb3BzIH0pOwogIGlmICghc2F2ZUdyb3VwcyhHUk9VUFMpKSB7CiAgICBhbGVydCgiVGhpcyBi"
    "cm93c2VyIHdvdWxkIG5vdCBsZXQgbWUgc2F2ZSB0aGUgZ3JvdXAsIHNvIGl0IHdpbGwgbm90IHBlcnNpc3QuIFRoZSBVUkwgc3Rp"
    "bGwgaG9sZHMgdGhlIHNlbGVjdGlvbiAtIGJvb2ttYXJrIGl0IGluc3RlYWQuIik7CiAgfQogIFZJRVcubGFiZWwgPSBuYW1lOwog"
    "IGJ1aWxkRmlsdGVyUGFuZWwoKTsgc3luY1NlbGVjdG9yKCk7IHdyaXRlSGFzaCgpOyBkcmF3KCk7Cn0KZnVuY3Rpb24gb3Blbkdy"
    "b3VwKGcpIHsKICBjb25zdCBwcm9wcyA9IGcucHJvcHMuZmlsdGVyKHAgPT4gUC5wcm9wZXJ0aWVzW3BdKTsgICAvLyBhIHByb3Bl"
    "cnR5IG1heSBoYXZlIGxlZnQgdGhlIHdvcmtib29rCiAgaWYgKCFwcm9wcy5sZW5ndGgpIHsgYWxlcnQoYE5vbmUgb2YgdGhlIHBy"
    "b3BlcnRpZXMgaW4gIiR7Zy5uYW1lfSIgYXJlIGluIHRoZSBjdXJyZW50IHJlcG9ydC5gKTsgcmV0dXJuOyB9CiAgc2V0SW5jbHVk"
    "ZShwcm9wcy5sZW5ndGggPT09IGFsbFByb3BzKCkubGVuZ3RoID8gbnVsbCA6IHByb3BzLCBnLm5hbWUpOwogIHdyaXRlSGFzaCgp"
    "OyBzeW5jU2VsZWN0b3IoKTsgZHJhdygpOwogICQoImZpbHRlclBhbmVsIikuaGlkZGVuID0gdHJ1ZTsKfQoKLyogPT09PT09PT09"
    "PT09PT09PT09PT09PT09PT09PSByb3V0aW5nID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gKi8KZnVuY3Rpb24gd3JpdGVI"
    "YXNoKCkgewogIGNvbnN0IHBhcnRzID0gW107CiAgaWYgKFZJRVcudHlwZSA9PT0gInByb3BlcnR5IikgcGFydHMucHVzaChlbmNv"
    "ZGVVUklDb21wb25lbnQoVklFVy5sYWJlbCkpOwogIGVsc2UgaWYgKGlzRmlsdGVyZWQoKSkgcGFydHMucHVzaCgiZz0iICsgaW5j"
    "bHVkZWRQcm9wcygpLm1hcChlbmNvZGVVUklDb21wb25lbnQpLmpvaW4oIiwiKSk7CiAgaWYgKG1GaWx0ZXJlZCgpKSBwYXJ0cy5w"
    "dXNoKCJtPSIgKyBNU0VMWzBdICsgIi0iICsgTVNFTFsxXSk7CiAgY29uc3QgaCA9IHBhcnRzLmpvaW4oIiYiKTsKICBoaXN0b3J5"
    "LnJlcGxhY2VTdGF0ZShudWxsLCAiIiwgaCA/ICIjIiArIGggOiBsb2NhdGlvbi5wYXRobmFtZSArIGxvY2F0aW9uLnNlYXJjaCk7"
    "Cn0KZnVuY3Rpb24gcmVhZEhhc2goKSB7CiAgY29uc3QgYWxsID0gbG9jYXRpb24uaGFzaC5zbGljZSgxKS5zcGxpdCgiJiIpOwog"
    "IGNvbnN0IGxhc3QgPSBQLm1vbnRocy5sZW5ndGggLSAxOwogIE1TRUwgPSBtQWxsKCk7CiAgY29uc3QgbVBhcnQgPSBhbGwuZmlu"
    "ZChwID0+IHAuc3RhcnRzV2l0aCgibT0iKSk7CiAgaWYgKG1QYXJ0KSB7CiAgICBjb25zdCBbYSwgYl0gPSBtUGFydC5zbGljZSgy"
    "KS5zcGxpdCgiLSIpLm1hcChOdW1iZXIpOwogICAgaWYgKE51bWJlci5pc0ludGVnZXIoYSkgJiYgTnVtYmVyLmlzSW50ZWdlcihi"
    "KSAmJiBhID49IDAgJiYgYiA8PSBsYXN0ICYmIGEgPD0gYikgTVNFTCA9IFthLCBiXTsKICB9CiAgY29uc3QgcmF3ID0gYWxsLmZp"
    "bmQocCA9PiAhcC5zdGFydHNXaXRoKCJtPSIpKSB8fCAiIjsKICBpZiAoIXJhdykgeyBzZXRJbmNsdWRlKG51bGwsICJQb3J0Zm9s"
    "aW8iKTsgcmV0dXJuOyB9CiAgaWYgKHJhdy5zdGFydHNXaXRoKCJnPSIpKSB7CiAgICBjb25zdCBwcm9wcyA9IHJhdy5zbGljZSgy"
    "KS5zcGxpdCgiLCIpLm1hcChkZWNvZGVVUklDb21wb25lbnQpLmZpbHRlcihwID0+IFAucHJvcGVydGllc1twXSk7CiAgICBpZiAo"
    "cHJvcHMubGVuZ3RoKSB7CiAgICAgIGNvbnN0IG5hbWVkID0gR1JPVVBTLmZpbmQoZyA9PiBnLnByb3BzLmxlbmd0aCA9PT0gcHJv"
    "cHMubGVuZ3RoICYmIGcucHJvcHMuZXZlcnkocCA9PiBwcm9wcy5pbmNsdWRlcyhwKSkpOwogICAgICBjb25zdCBlbnQgPSBPYmpl"
    "Y3Qua2V5cyhQLmVudGl0aWVzKS5maW5kKGUgPT4KICAgICAgICBQLmVudGl0aWVzW2VdLmxlbmd0aCA9PT0gcHJvcHMubGVuZ3Ro"
    "ICYmIFAuZW50aXRpZXNbZV0uZXZlcnkocCA9PiBwcm9wcy5pbmNsdWRlcyhwKSkpOwogICAgICBzZXRJbmNsdWRlKHByb3BzLCBu"
    "YW1lZCA/IG5hbWVkLm5hbWUgOiBlbnQgfHwgIkN1c3RvbSBzZWxlY3Rpb24iKTsKICAgICAgcmV0dXJuOwogICAgfQogIH0KICBj"
    "b25zdCBuYW1lID0gZGVjb2RlVVJJQ29tcG9uZW50KHJhdyk7CiAgaWYgKFAucHJvcGVydGllc1tuYW1lXSkgeyBWSUVXLnR5cGUg"
    "PSAicHJvcGVydHkiOyBWSUVXLmxhYmVsID0gbmFtZTsgVklFVy5pbmNsdWRlID0gbnVsbDsgcmV0dXJuOyB9CiAgc2V0SW5jbHVk"
    "ZShudWxsLCAiUG9ydGZvbGlvIik7Cn0KCmZ1bmN0aW9uIHN5bmNTZWxlY3RvcigpIHsKICBjb25zdCBzZWwgPSAkKCJwcm9wU2Vs"
    "Iik7CiAgbGV0IGh0bWwgPSBgPG9wdGlvbiB2YWx1ZT0iX19BTExfXyI+UG9ydGZvbGlvIOKAlCBhbGwgcHJvcGVydGllczwvb3B0"
    "aW9uPmA7CiAgaWYgKEdST1VQUy5sZW5ndGgpIHsKICAgIGh0bWwgKz0gYDxvcHRncm91cCBsYWJlbD0iU2F2ZWQgZ3JvdXBzIj5g"
    "ICsgR1JPVVBTLm1hcChnID0+CiAgICAgIGA8b3B0aW9uIHZhbHVlPSJfX2c6JHtlc2MoZy5pZCl9Ij4ke2VzYyhnLm5hbWUpfSAo"
    "JHtnLnByb3BzLmxlbmd0aH0pPC9vcHRpb24+YCkuam9pbigiIikgKyBgPC9vcHRncm91cD5gOwogIH0KICBjb25zdCBlbnRzID0g"
    "T2JqZWN0LmtleXMoUC5lbnRpdGllcykuZmlsdGVyKGUgPT4gUC5lbnRpdGllc1tlXS5sZW5ndGggPiAxKTsKICBpZiAoZW50cy5s"
    "ZW5ndGgpIHsKICAgIGh0bWwgKz0gYDxvcHRncm91cCBsYWJlbD0iRW50aXRpZXMiPmAgKyBlbnRzLm1hcChlID0+CiAgICAgIGA8"
    "b3B0aW9uIHZhbHVlPSJfX2U6JHtlc2MoZSl9Ij4ke2VzYyhlKX0gKCR7UC5lbnRpdGllc1tlXS5sZW5ndGh9KTwvb3B0aW9uPmAp"
    "LmpvaW4oIiIpICsgYDwvb3B0Z3JvdXA+YDsKICB9CiAgZm9yIChjb25zdCBlIGluIFAuZW50aXRpZXMpIHsKICAgIGh0bWwgKz0g"
    "YDxvcHRncm91cCBsYWJlbD0iJHtlc2MoZSl9Ij5gICsKICAgICAgUC5lbnRpdGllc1tlXS5tYXAocCA9PiBgPG9wdGlvbiB2YWx1"
    "ZT0iJHtlc2MocCl9Ij4ke2VzYyhwKX08L29wdGlvbj5gKS5qb2luKCIiKSArIGA8L29wdGdyb3VwPmA7CiAgfQogIGNvbnN0IGdy"
    "b3VwZWQgPSBuZXcgU2V0KE9iamVjdC52YWx1ZXMoUC5lbnRpdGllcykuZmxhdCgpKTsKICBjb25zdCBsb29zZSA9IGFsbFByb3Bz"
    "KCkuZmlsdGVyKHAgPT4gIWdyb3VwZWQuaGFzKHApKTsKICBpZiAobG9vc2UubGVuZ3RoKSBodG1sICs9IGA8b3B0Z3JvdXAgbGFi"
    "ZWw9Ik90aGVyIj5gICsKICAgIGxvb3NlLm1hcChwID0+IGA8b3B0aW9uIHZhbHVlPSIke2VzYyhwKX0iPiR7ZXNjKHApfTwvb3B0"
    "aW9uPmApLmpvaW4oIiIpICsgYDwvb3B0Z3JvdXA+YDsKICAvLyBSZWZsZWN0IHRoZSBjdXJyZW50IHZpZXcgaW4gdGhlIHNlbGVj"
    "dG9yIHdoZXJlIG9uZSBvZiBpdHMgb3B0aW9ucyBtYXRjaGVzIGl0LgogIGxldCB2YWwgPSAiX19BTExfXyI7CiAgaWYgKFZJRVcu"
    "dHlwZSA9PT0gInByb3BlcnR5IikgdmFsID0gVklFVy5sYWJlbDsKICBlbHNlIGlmIChpc0ZpbHRlcmVkKCkpIHsKICAgIGNvbnN0"
    "IGluYyA9IGluY2x1ZGVkUHJvcHMoKTsKICAgIGNvbnN0IGcgPSBHUk9VUFMuZmluZCh4ID0+IHgucHJvcHMubGVuZ3RoID09PSBp"
    "bmMubGVuZ3RoICYmIHgucHJvcHMuZXZlcnkocCA9PiBpbmMuaW5jbHVkZXMocCkpKTsKICAgIGNvbnN0IGUgPSBPYmplY3Qua2V5"
    "cyhQLmVudGl0aWVzKS5maW5kKHggPT4KICAgICAgUC5lbnRpdGllc1t4XS5sZW5ndGggPT09IGluYy5sZW5ndGggJiYgUC5lbnRp"
    "dGllc1t4XS5ldmVyeShwID0+IGluYy5pbmNsdWRlcyhwKSkpOwogICAgdmFsID0gZyA/ICJfX2c6IiArIGcuaWQgOiBlID8gIl9f"
    "ZToiICsgZSA6ICJfX0NVU1RPTV9fIjsKICAgIGlmICh2YWwgPT09ICJfX0NVU1RPTV9fIikgewogICAgICBodG1sID0gYDxvcHRp"
    "b24gdmFsdWU9Il9fQ1VTVE9NX18iPkN1c3RvbSBzZWxlY3Rpb24gKCR7aW5jLmxlbmd0aH0pPC9vcHRpb24+YCArIGh0bWw7CiAg"
    "ICB9CiAgfQogIHNlbC5pbm5lckhUTUwgPSBodG1sOwogIHNlbC52YWx1ZSA9IHZhbDsKfQoKZnVuY3Rpb24gcGljayh2KSB7CiAg"
    "aWYgKHYgPT09ICJfX0NVU1RPTV9fIikgcmV0dXJuOwogIGlmICh2ID09PSAiX19BTExfXyIpIHNldEluY2x1ZGUobnVsbCwgIlBv"
    "cnRmb2xpbyIpOwogIGVsc2UgaWYgKHYuc3RhcnRzV2l0aCgiX19nOiIpKSB7CiAgICBjb25zdCBnID0gR1JPVVBTLmZpbmQoeCA9"
    "PiB4LmlkID09PSB2LnNsaWNlKDQpKTsKICAgIGlmIChnKSByZXR1cm4gb3Blbkdyb3VwKGcpOwogIH0gZWxzZSBpZiAodi5zdGFy"
    "dHNXaXRoKCJfX2U6IikpIHsKICAgIGNvbnN0IGUgPSB2LnNsaWNlKDQpOwogICAgc2V0SW5jbHVkZShQLmVudGl0aWVzW2VdIHx8"
    "IFtdLCBlKTsKICB9IGVsc2UgaWYgKFAucHJvcGVydGllc1t2XSkgewogICAgVklFVy50eXBlID0gInByb3BlcnR5IjsgVklFVy5s"
    "YWJlbCA9IHY7IFZJRVcuaW5jbHVkZSA9IG51bGw7CiAgfQogIHdyaXRlSGFzaCgpOyBzeW5jU2VsZWN0b3IoKTsgZHJhdygpOwog"
    "IHNjcm9sbFRvKHsgdG9wOiAwLCBiZWhhdmlvcjogInNtb290aCIgfSk7Cn0KCmZ1bmN0aW9uIGRyYXcoKSB7CiAgaGlkZVRpcCgp"
    "OwogIHVwZGF0ZUZpbHRlclVJKCk7CiAgJCgibW9udGhSb3ciKS5oaWRkZW4gPSBmYWxzZTsKICBpZiAoVklFVy50eXBlID09PSAi"
    "cG9ydGZvbGlvIikgcmVuZGVyUG9ydGZvbGlvKCk7IGVsc2UgcmVuZGVyUHJvcGVydHkoVklFVy5sYWJlbCk7Cn0KCi8qID09PT09"
    "PT09PT09PT09PT09PT09PT09PT09PT0gYXBwID09PT09PT09PT09PT09PT09PT09PT09PT09PT0gKi8KZnVuY3Rpb24gYm9vdCgp"
    "IHsKICBpbml0R3JvdXBzKCk7CiAgTVNFTCA9IG1BbGwoKTsKICBHUk9VUFMgPSBsb2FkR3JvdXBzKCk7CiAgcmVhZEhhc2goKTsK"
    "ICByZW5kZXJNb250aHMoKTsKICBzeW5jU2VsZWN0b3IoKTsKICAkKCJwcm9wU2VsIikuYWRkRXZlbnRMaXN0ZW5lcigiY2hhbmdl"
    "IiwgZSA9PiBwaWNrKGUudGFyZ2V0LnZhbHVlKSk7CiAgYWRkRXZlbnRMaXN0ZW5lcigiaGFzaGNoYW5nZSIsICgpID0+IHsgcmVh"
    "ZEhhc2goKTsgc3luY1NlbGVjdG9yKCk7IGRyYXcoKTsgfSk7CgogICQoImZvb3QiKS5pbm5lckhUTUwgPQogICAgYEdlbmVyYXRl"
    "ZCAke2VzYyhNRVRBLmdlbmVyYXRlZCl9IGZyb20gPGI+JHtlc2MoTUVUQS5zb3VyY2UpfTwvYj4sIHRoZSBjb25zb2xpZGF0ZWQg"
    "cHJvZml0ICZhbXA7IGxvc3MgcHJlcGFyZWQgYnkgdGhlIHBvcnRmb2xpbyBhY2NvdW50YW50LiBgICsKICAgIGBGaWd1cmVzIGZv"
    "bGxvdyB0aGF0IHdvcmtib29rIGFzIHJlcG9ydGVkLiBSZWZyZXNoZWQgbW9udGhseS4gYCArCiAgICBgU2F2ZWQgZ3JvdXBzIGFy"
    "ZSBzdG9yZWQgaW4gdGhpcyBicm93c2VyOyBzaGFyZSBhIHZpZXcgYnkgY29weWluZyB0aGUgVVJMLmA7CgogIGNvbnN0IGJ0biA9"
    "ICQoInRoZW1lQnRuIik7CiAgY29uc3Qgc3luY0J0biA9ICgpID0+IGJ0bi50ZXh0Q29udGVudCA9IGRvY3VtZW50LmRvY3VtZW50"
    "RWxlbWVudC5kYXRhc2V0LnRoZW1lID09PSAiZGFyayIgPyAiTGlnaHQgbW9kZSIgOiAiRGFyayBtb2RlIjsKICBzeW5jQnRuKCk7"
    "CiAgYnRuLmFkZEV2ZW50TGlzdGVuZXIoImNsaWNrIiwgKCkgPT4gewogICAgZG9jdW1lbnQuZG9jdW1lbnRFbGVtZW50LmRhdGFz"
    "ZXQudGhlbWUgPSBkb2N1bWVudC5kb2N1bWVudEVsZW1lbnQuZGF0YXNldC50aGVtZSA9PT0gImRhcmsiID8gImxpZ2h0IiA6ICJk"
    "YXJrIjsKICAgIHN5bmNCdG4oKTsgZHJhdygpOwogIH0pOwogICQoImJvZHkiKS5hZGRFdmVudExpc3RlbmVyKCJjbGljayIsIGUg"
    "PT4gewogICAgY29uc3QgYnRuID0gZS50YXJnZXQuY2xvc2VzdCgiW2RhdGEtZG93bmxvYWRdIik7CiAgICBpZiAoYnRuKSBkb3du"
    "bG9hZENTVihidG4pOwogIH0pOwogIGFkZEV2ZW50TGlzdGVuZXIoInJlc2l6ZSIsIGhpZGVUaXApOwogIGRyYXcoKTsKfQo8L3Nj"
    "cmlwdD4KPC9ib2R5Pgo8L2h0bWw+Cg=="
])

if __name__ == "__main__":
    main()
